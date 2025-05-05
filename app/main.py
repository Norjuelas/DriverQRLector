################################################################################
## V: 1.0.0
################################################################################

import sys
import os
import platform
import uuid
import hashlib
import qrcode
import datetime
import traceback # Para detalles de error
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from PIL import ImageFont

from openpyxl import Workbook, load_workbook
import barcode
from barcode.writer import ImageWriter

from PySide2 import QtCore, QtGui, QtWidgets
from PySide2.QtCore import (QCoreApplication, QPropertyAnimation, QDate, QDateTime, 
                          QMetaObject, QObject, QPoint, QRect, QSize, QTime, QUrl, Qt, QEvent)
from PySide2.QtGui import (QBrush, QColor, QConicalGradient, QCursor, QFont, QFontDatabase, 
                         QIcon, QKeySequence, QLinearGradient, QPalette, QPainter, QPixmap, 
                         QRadialGradient)
from PySide2.QtWidgets import *

# GUI FILE
from app_modules import *

class MainWindow(QMainWindow):
    def __init__(self):
        QMainWindow.__init__(self)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.dragPos = None
        
        # Current code type (barcode or qr)
        self.current_code_type = "barcode"  # Default to barcode

        # Print system information
        print('System: ' + platform.system())
        print('Version: ' + platform.release())

        # Setup the window
        self.setup_window()
        
        # Setup menus
        self.setup_menus()
        
        # Setup barcode generator functionality
        self.setup_code_generator()
        
        # Setup code reader functionality
        self.setup_code_reader()
        
        # Setup employee management
        self.setup_employee_management()

        # Setup update reports button
        self.setup_update_button()

        # Show the window
        self.show()

    def setup_window(self):
        """Setup basic window properties and appearance"""
        # Remove standard title bar
        UIFunctions.removeTitleBar(True)
        
        # Set window title
        self.setWindowTitle('Gestor de Vales')
        UIFunctions.labelTitle(self, 'Thimoty')
        UIFunctions.labelDescription(self, '2025')
        
        # Set window size
        startSize = QSize(1000, 720)
        self.resize(startSize)
        self.setMinimumSize(startSize)
        # UIFunctions.enableMaximumSize(self, 500, 720)

    def setup_menus(self):
        """Setup menu functionality and navigation"""
        # Toggle menu size button
        self.ui.btn_toggle_menu.clicked.connect(lambda: UIFunctions.toggleMenu(self, 220, True))
        
        # Add custom menus
        self.ui.stackedWidget.setMinimumWidth(20)
        UIFunctions.addNewMenu(self, "Leer Vales", "btn_home", "url(:/16x16/icons/16x16/cil-home.png)", True)
        UIFunctions.addNewMenu(self, "Crear Vales", "btn_new_user", "url(:/16x16/icons/16x16/cil-user-follow.png)", True)
        UIFunctions.addNewMenu(self, "Configuracion", "btn_widgets", "url(:/16x16/icons/16x16/cil-equalizer.png)", False)
        
        # Select starting menu
        UIFunctions.selectStandardMenu(self, "btn_home")
        
        # Set starting page
        self.ui.stackedWidget.setCurrentWidget(self.ui.page_home)

    def setup_code_generator(self):
        """
        Setup the barcode and QR code generator functionality
        """
        # Define excel path and create barcode directory if needed
        self.excel_path = "trabajos_database.xlsx"
        self.vales_sheet_name = "Vales"  # Sheet name for vales data
        
        if not os.path.exists("codes"):
            os.makedirs("codes")
        
        # Create Excel file if it doesn't exist
        if not os.path.exists(self.excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Trabajos"
            ws.append(["Código Serial", "Tipo Trabajo", "Referencia", "Número Ticket", 
                      "Talla", "Color", "Valor", "Total Producido", "Tipo Código", "Ruta Imagen"])
            
            # Create Vales sheet
            vales_ws = wb.create_sheet(title=self.vales_sheet_name)
            vales_ws.append(["ID", "EmpleadoID", "Fecha", "Tipo Trabajo", "Número Ticket",
                           "Referencia", "Talla", "Color", "Valor", "Total Producido", "Código Serial"])
            
            wb.save(self.excel_path)
        else:
            # Check if Vales sheet exists, if not create it
            wb = load_workbook(self.excel_path)
            if self.vales_sheet_name not in wb.sheetnames:
                vales_ws = wb.create_sheet(title=self.vales_sheet_name)
                vales_ws.append(["ID", "EmpleadoID", "Fecha", "Tipo Trabajo", "Número Ticket",
                               "Referencia", "Talla", "Color", "Valor", "Total Producido", "Código Serial"])
                wb.save(self.excel_path)
        
        # Connect the save button
        self.ui.pushButtonGuardar.clicked.connect(self.on_save_button_clicked)
        
        # Add code type selector (radio buttons)
        self.setup_code_type_selector()
        
        # Ensure the QGraphicsView has a scene
        if not hasattr(self.ui, 'PreviwImage'):
            print("Warning: PreviwImage not found in UI")
        elif self.ui.PreviwImage.scene() is None:
            self.ui.PreviwImage.setScene(QtWidgets.QGraphicsScene())
    
    def setup_code_type_selector(self):
        """Setup UI elements to select between barcode and QR code"""
        # Create a frame for radio buttons if it doesn't exist
        if not hasattr(self.ui, 'codeTypeFrame'):
            self.ui.codeTypeFrame = QFrame()
            self.ui.codeTypeFrame.setFrameShape(QFrame.StyledPanel)
            self.ui.codeTypeFrame.setFrameShadow(QFrame.Raised)
            
            # Create radio buttons
            self.ui.radioBarcode = QRadioButton("Código de Barras")
            self.ui.radioQR = QRadioButton("Código QR")
            self.ui.radioBarcode.setChecked(True)  # Barcode is default
            
            # Create layout
            layout = QHBoxLayout(self.ui.codeTypeFrame)
            layout.addWidget(self.ui.radioBarcode)
            layout.addWidget(self.ui.radioQR)
            
            # Connect signals
            self.ui.radioBarcode.toggled.connect(self.on_code_type_changed)
            
            # Find a place to add the frame (this depends on your UI layout)
            # For example, if there's a verticalLayout in the form:
            if hasattr(self.ui, 'formLayout'):
                # Insert at position 0 (top)
                self.ui.formLayout.insertRow(0, "Tipo de Código:", self.ui.codeTypeFrame)
    
    def setup_code_reader(self):
        """Setup the code reader functionality and TableView inside WidgetTabla"""
        # Create a code reader input field if it doesn't exist
        if not hasattr(self.ui, 'codeReaderInput'):
            # Code reader input field
            self.ui.codeReaderInput = QLineEdit()
            self.ui.codeReaderInput.setPlaceholderText("Escanear o ingresar código aquí...")
            self.ui.codeReaderInput.setMinimumWidth(250)
            self.ui.codeReaderInput.returnPressed.connect(self.on_code_scanned)
            
            # Add to layout
            if hasattr(self.ui, 'verticalLayout'):
                # Create a label for the input
                label = QLabel("Lector de Códigos:")
                
                # Create horizontal layout for label and input
                h_layout = QHBoxLayout()
                h_layout.addWidget(label)
                h_layout.addWidget(self.ui.codeReaderInput)
                
                # Create a container widget
                container = QWidget()
                container.setLayout(h_layout)
                
                # Add to vertical layout
                self.ui.verticalLayout.addWidget(container)
        
        # Access the WidgetTabla from UI and setup table inside it
        if hasattr(self.ui, 'WidgetTabla'):
            # Create table layout
            table_layout = QVBoxLayout(self.ui.WidgetTabla)
            self.ui.WidgetTabla.setLayout(table_layout)
            
            # Create TableView
            self.ui.tableViewVale = QTableView(self.ui.WidgetTabla)
            self.ui.tableViewVale.setMinimumHeight(200)
            
            # Create the model for the table
            self.table_model = QtGui.QStandardItemModel()
            self.table_model.setHorizontalHeaderLabels([
                "ID", "Código Serial", "Tipo Trabajo", "Referencia", "Número Ticket", 
                "Talla", "Color", "Valor", "Total Producido", "Empleado"
            ])
            
            # Set model and adjust view
            self.ui.tableViewVale.setModel(self.table_model)
            self.ui.tableViewVale.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            
            # Add TableView to WidgetTabla's layout
            table_layout.addWidget(self.ui.tableViewVale)
        else:
            print("Warning: WidgetTabla not found in UI")
    def setup_employee_management(self):
        """Carga empleados en el ComboBox y conecta el botón Registrar Vale."""
        try:
            # --- Manejo de Archivo Excel ---
            try:
                wb = load_workbook(self.excel_path)
            except FileNotFoundError:
                print(f"Advertencia: Archivo Excel no encontrado en {self.excel_path}. Creando uno nuevo.")
                wb = Workbook()
                # Usa la hoja activa si es un libro nuevo
                if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
                    empleados_ws = wb.active
                    empleados_ws.title = "Empleados"
                else:
                     empleados_ws = wb.create_sheet(title="Empleados")
                # Añadir encabezados al crear
                empleados_ws.append(["Nombre", "Cedula", "Celular", "Correo", "EmpleadoId"])
                # Puedes añadir un empleado de ejemplo si quieres
                empleados_ws.append(["Juan Pérez (Ejemplo)", "1234567890", "3001234567", "juan@example.com", "E001"])
                wb.save(self.excel_path)
                print(f"Archivo Excel '{self.excel_path}' creado con hoja 'Empleados'.")

            if "Empleados" not in wb.sheetnames:
                print(f"Creando hoja 'Empleados' en archivo existente: {self.excel_path}")
                empleados_ws = wb.create_sheet(title="Empleados")
                empleados_ws.append(["Nombre", "Cedula", "Celular", "Correo", "EmpleadoId"])
                # Puedes añadir un empleado de ejemplo si quieres
                empleados_ws.append(["Juan Pérez (Ejemplo)", "1234567890", "3001234567", "juan@example.com", "E001"])
                wb.save(self.excel_path)
            else:
                empleados_ws = wb["Empleados"]

            # --- Configuración del ComboBox (Asumiendo que existe en self.ui) ---
            if hasattr(self.ui, 'EmpleadosBox'):
                self.ui.EmpleadosBox.clear() # Limpiar items previos
                self.ui.EmpleadosBox.setMinimumWidth(200) # Establecer ancho mínimo si es necesario

                employees_loaded = False
                # Iterar desde la fila 2 para saltar encabezados (ajusta si no tienes encabezados)
                for row in empleados_ws.iter_rows(min_row=2, max_col=5): # Leer solo las 5 columnas necesarias
                    # row es una tupla de celdas
                    if len(row) >= 5 and row[0].value: # Verificar que hay nombre y suficientes columnas
                        name = str(row[0].value) # Columna A (Nombre)
                        emp_id = str(row[4].value) if row[4].value else "SIN-ID" # Columna E (EmpleadoId)

                        # Añadir item: Texto visible y dato asociado (emp_id)
                        self.ui.EmpleadosBox.addItem(f"{name} ({emp_id})", emp_id)
                        employees_loaded = True

                if not employees_loaded:
                    self.ui.EmpleadosBox.addItem("Sin empleados registrados", "") # Opción por defecto
                    print("No se cargaron empleados desde la hoja 'Empleados'.")
                else:
                    print(f"Se cargaron {self.ui.EmpleadosBox.count()} empleados.")

            else:
                # Error crítico si el ComboBox no está en la UI
                print("ERROR: El QComboBox 'EmpleadosBox' no se encontró en la UI.")
                # Podrías deshabilitar el botón si el combo no existe
                if hasattr(self.ui, 'btnRegisterVale'):
                    self.ui.btnRegisterVale.setEnabled(False)
                return # Salir de la configuración si falta el combo

            # --- Conexión del Botón (Asumiendo que existe en self.ui) ---
# --- Conexión del Botón (Asumiendo que existe en self.ui) ---
            if hasattr(self.ui, 'btnRegisterVale'):
                # Verificar que el método al que conectar existe en esta clase
                if hasattr(self, 'register_vale'):
                    # Desconectar primero (de forma segura) por si esta función se llama múltiples veces
                    try:
                        # Intenta desconectar cualquier conexión previa de 'clicked'
                        self.ui.btnRegisterVale.clicked.disconnect()
                        # Opcional: puedes añadir un print aquí si quieres saber cuándo la desconexión SÍ funciona
                        # print("Desconexión previa de btnRegisterVale.clicked exitosa.")
                    except RuntimeError:
                        # Ignora el error si no había nada conectado previamente.
                        # print("Advertencia: No se pudo desconectar btnRegisterVale.clicked (probablemente no estaba conectado).")
                        pass
                    # Podrías añadir 'except TypeError: pass' también por si acaso, aunque el error fue RuntimeError

                    # Conectar la señal 'clicked' al método 'register_vale'
                    # ESTA ES LA LÍNEA IMPORTANTE AHORA
                    self.ui.btnRegisterVale.clicked.connect(self.register_vale)
                    print("Botón 'btnRegisterVale' conectado correctamente.") # <-- ¡Busca este mensaje!
                    self.ui.btnRegisterVale.setEnabled(True) # Asegurarse de que esté habilitado
                else:
                     # Error si el método no existe
                    print("ERROR: El método 'register_vale' no está definido en esta clase. El botón no funcionará.")
                    self.ui.btnRegisterVale.setEnabled(False) # Deshabilitar el botón
            else:
                # Error crítico si el botón no está en la UI
                print("ERROR: El QPushButton 'btnRegisterVale' no se encontró en la UI.")

        except FileNotFoundError:
             print(f"ERROR CRÍTICO: No se pudo encontrar ni crear el archivo Excel en {self.excel_path}")
        except Exception as e:
            print(f"Error inesperado durante setup_employee_management: {e}")
            traceback.print_exc() # Imprime el rastreo completo del error para depuración

    def update_employee_reports(self):
        """
        Crea o actualiza hojas individuales por cada empleado en el archivo Excel.
        Estas hojas contienen:
        1. Una tabla con todos los vales del empleado
        2. La suma total del campo Valor
        3. Consolidados semanales, mensuales y anuales
        """
        try:
            print("Actualizando reportes de empleados...")
            
            # Cargar el archivo Excel
            wb = load_workbook(self.excel_path)
            
            # Verificar si existen las hojas necesarias
            if "Vales" not in wb.sheetnames or "Empleados" not in wb.sheetnames:
                QMessageBox.warning(self, "Error", "No se encontraron las hojas 'Vales' o 'Empleados' en el archivo Excel.")
                return
            
            # Obtener datos de empleados
            empleados_ws = wb["Empleados"]
            empleados = {}
            for row in list(empleados_ws.rows)[1:]:  # Saltar el encabezado
                if row[0].value and row[4].value:  # Nombre y EmpleadoId
                    empleados[row[4].value] = {
                        "nombre": row[0].value,
                        "cedula": row[1].value,
                        "celular": row[2].value,
                        "correo": row[3].value
                    }
            
            if not empleados:
                QMessageBox.information(self, "Info", "No hay empleados registrados en la base de datos.")
                return
            
            # Obtener datos de vales
            vales_ws = wb["Vales"]
            vales_data = []
            for row in list(vales_ws.rows)[1:]:  # Saltar el encabezado
                # Verificar si la fila tiene suficientes columnas y datos necesarios
                if len(row) > 8 and row[1].value is not None:  # Al menos necesitamos EmpleadoID
                    try:
                        # Convertir valor a número si es posible
                        valor = None
                        if len(row) > 8 and row[8].value is not None:
                            if isinstance(row[8].value, (int, float)):
                                valor = float(row[8].value)
                            elif isinstance(row[8].value, str):
                                # Limpiar string de caracteres no numéricos
                                valor_str = ''.join(c for c in row[8].value if c.isdigit() or c == '.')
                                if valor_str:
                                    valor = float(valor_str) if '.' in valor_str else int(valor_str)
                        
                        vales_data.append({
                            "id": row[0].value if len(row) > 0 else None,
                            "empleado_id": row[1].value,  # Campo obligatorio
                            "fecha": row[2].value if len(row) > 2 else None,
                            "tipo_trabajo": row[3].value if len(row) > 3 else None,
                            "num_ticket": row[4].value if len(row) > 4 else None,
                            "referencia": row[5].value if len(row) > 5 else None,
                            "talla": row[6].value if len(row) > 6 else None,
                            "color": row[7].value if len(row) > 7 else None,
                            "valor": valor,  # Ya convertido a número
                            "total_producido": row[9].value if len(row) > 9 else None,
                            "codigo_serial": row[10].value if len(row) > 10 else None
                        })
                    except Exception as e:
                        print(f"Error al procesar fila de vale: {str(e)}")
            
            # Para cada empleado, crear o actualizar su hoja
            for empleado_id, empleado_info in empleados.items():
                # Obtener vales del empleado
                vales_empleado = [vale for vale in vales_data if vale["empleado_id"] == empleado_id]
                
                # Si el empleado ya tiene una hoja, eliminarla para recrearla
                sheet_name = f"Empleado_{empleado_id}"
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
                
                # Crear nueva hoja para el empleado
                emp_ws = wb.create_sheet(title=sheet_name)
                
                # Información del empleado en la parte superior
                emp_ws['A1'] = "INFORMACIÓN DEL EMPLEADO"
                emp_ws['A1'].font = Font(bold=True, size=14)
                emp_ws.merge_cells('A1:F1')
                
                emp_ws['A2'] = "Nombre:"
                emp_ws['B2'] = empleado_info["nombre"]
                emp_ws['A3'] = "Cédula:"
                emp_ws['B3'] = empleado_info["cedula"]
                emp_ws['A4'] = "Celular:"
                emp_ws['B4'] = empleado_info["celular"]
                emp_ws['A5'] = "Correo:"
                emp_ws['B5'] = empleado_info["correo"]
                emp_ws['A6'] = "ID Empleado:"
                emp_ws['B6'] = empleado_id
                
                # Dar formato a la información del empleado
                for row in range(2, 7):
                    emp_ws[f'A{row}'].font = Font(bold=True)
                
                # Espacio antes de la tabla de vales
                emp_ws['A8'] = "DETALLE DE VALES"
                emp_ws['A8'].font = Font(bold=True, size=12)
                emp_ws.merge_cells('A8:F8')
                
                # Encabezados tabla de vales
                headers = ["ID", "Fecha", "Tipo Trabajo", "# Ticket", "Referencia", 
                        "Talla", "Color", "Valor", "Total Producido", "Código Serial"]
                for col, header in enumerate(headers, start=1):
                    cell = emp_ws.cell(row=9, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                # Datos de vales
                valor_total = 0
                for row_idx, vale in enumerate(vales_empleado, start=10):
                    emp_ws.cell(row=row_idx, column=1).value = vale["id"]
                    
                    # Formato de fecha
                    if vale["fecha"] is not None:
                        if isinstance(vale["fecha"], datetime.datetime):
                            emp_ws.cell(row=row_idx, column=2).value = vale["fecha"].strftime("%Y-%m-%d")
                        else:
                            emp_ws.cell(row=row_idx, column=2).value = str(vale["fecha"])
                    else:
                        emp_ws.cell(row=row_idx, column=2).value = "Sin fecha"
                    
                    emp_ws.cell(row=row_idx, column=3).value = vale["tipo_trabajo"]
                    emp_ws.cell(row=row_idx, column=4).value = vale["num_ticket"]
                    emp_ws.cell(row=row_idx, column=5).value = vale["referencia"]
                    emp_ws.cell(row=row_idx, column=6).value = vale["talla"]
                    emp_ws.cell(row=row_idx, column=7).value = vale["color"]
                    
                    # Valor con formato
                    valor_cell = emp_ws.cell(row=row_idx, column=8)
                    valor_cell.value = vale["valor"]
                    valor_cell.number_format = '#,##0'
                    
                    emp_ws.cell(row=row_idx, column=9).value = vale["total_producido"]
                    emp_ws.cell(row=row_idx, column=10).value = vale["codigo_serial"]
                    
                    # Convertir valor a número antes de sumar (maneja tanto strings como números)
                    if vale["valor"] is not None:
                        try:
                            # Si es string, convertir a entero o float
                            if isinstance(vale["valor"], str):
                                # Eliminar cualquier caracter no numérico (como símbolos de moneda o comas)
                                valor_numerico = ''.join(c for c in vale["valor"] if c.isdigit() or c == '.')
                                if valor_numerico:
                                    if '.' in valor_numerico:
                                        valor_total += float(valor_numerico)
                                    else:
                                        valor_total += int(valor_numerico)
                            else:
                                # Si ya es número, sumar directamente
                                valor_total += float(vale["valor"])
                        except (ValueError, TypeError):
                            print(f"Error al convertir valor '{vale['valor']}' a número, se ignora para el total")
                
                last_row = len(vales_empleado) + 10
                
                # Total de vales
                emp_ws.cell(row=last_row, column=7).value = "TOTAL:"
                emp_ws.cell(row=last_row, column=7).font = Font(bold=True)
                emp_ws.cell(row=last_row, column=7).alignment = Alignment(horizontal='right')
                
                total_cell = emp_ws.cell(row=last_row, column=8)
                total_cell.value = valor_total
                total_cell.font = Font(bold=True)
                total_cell.number_format = '#,##0'
                
                # Ajustar anchos de columna
                for col in range(1, 11):
                    emp_ws.column_dimensions[chr(64 + col)].width = 15
                
                # Espacio antes de los consolidados
                consolidado_start_row = last_row + 3
                emp_ws.cell(row=consolidado_start_row, column=1).value = "CONSOLIDADO DE PAGOS"
                emp_ws.cell(row=consolidado_start_row, column=1).font = Font(bold=True, size=12)
                emp_ws.merge_cells(f'A{consolidado_start_row}:F{consolidado_start_row}')
                
                # Encabezados tablas de consolidado
                consolidado_start_row += 1
                
                # Tabla Semanal
                emp_ws.cell(row=consolidado_start_row, column=1).value = "CONSOLIDADO SEMANAL"
                emp_ws.cell(row=consolidado_start_row, column=1).font = Font(bold=True)
                emp_ws.merge_cells(f'A{consolidado_start_row}:C{consolidado_start_row}')
                
                consolidado_start_row += 1
                emp_ws.cell(row=consolidado_start_row, column=1).value = "Semana"
                emp_ws.cell(row=consolidado_start_row, column=2).value = "Fecha Inicio"
                emp_ws.cell(row=consolidado_start_row, column=3).value = "Fecha Fin"
                emp_ws.cell(row=consolidado_start_row, column=4).value = "Valor Total"
                
                # Estilo encabezados
                for col in range(1, 5):
                    cell = emp_ws.cell(row=consolidado_start_row, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                # Calcular consolidado semanal
                weekly_data = {}
                for vale in vales_empleado:
                    if vale["fecha"] is not None and isinstance(vale["fecha"], datetime.datetime):
                        try:
                            # Obtener el número de semana del año
                            year = vale["fecha"].year
                            week_num = vale["fecha"].isocalendar()[1]
                            week_key = f"{year}-W{week_num:02d}"
                            
                            # Calcular fecha inicio y fin de semana
                            # El lunes de la semana
                            start_of_week = vale["fecha"] - datetime.timedelta(days=vale["fecha"].weekday())
                            # El domingo de la semana
                            end_of_week = start_of_week + datetime.timedelta(days=6)
                        except Exception as e:
                            print(f"Error al procesar fecha para consolidado semanal: {str(e)}")
                            continue  # Salta este vale
                        
                        if week_key not in weekly_data:
                            weekly_data[week_key] = {
                                "start_date": start_of_week,
                                "end_date": end_of_week,
                                "total": 0
                            }
                        
                        # Convertir valor a número antes de sumar
                        if vale["valor"] is not None:
                            try:
                                # Si es string, convertir a número
                                if isinstance(vale["valor"], str):
                                    valor_numerico = ''.join(c for c in vale["valor"] if c.isdigit() or c == '.')
                                    if valor_numerico:
                                        if '.' in valor_numerico:
                                            weekly_data[week_key]["total"] += float(valor_numerico)
                                        else:
                                            weekly_data[week_key]["total"] += int(valor_numerico)
                                else:
                                    # Si ya es número, sumar directamente
                                    weekly_data[week_key]["total"] += float(vale["valor"])
                            except (ValueError, TypeError):
                                print(f"Error al convertir valor '{vale['valor']}' para suma semanal")
                
                # Llenar datos semanales
                row_idx = consolidado_start_row + 1
                for week_key, data in sorted(weekly_data.items()):
                    emp_ws.cell(row=row_idx, column=1).value = week_key
                    emp_ws.cell(row=row_idx, column=2).value = data["start_date"].strftime("%Y-%m-%d")
                    emp_ws.cell(row=row_idx, column=3).value = data["end_date"].strftime("%Y-%m-%d")
                    
                    valor_cell = emp_ws.cell(row=row_idx, column=4)
                    valor_cell.value = data["total"]
                    valor_cell.number_format = '#,##0'
                    
                    row_idx += 1
                
                # Espacio entre tablas
                row_idx += 2
                
                # Tabla Mensual
                emp_ws.cell(row=row_idx, column=1).value = "CONSOLIDADO MENSUAL"
                emp_ws.cell(row=row_idx, column=1).font = Font(bold=True)
                emp_ws.merge_cells(f'A{row_idx}:C{row_idx}')
                
                row_idx += 1
                emp_ws.cell(row=row_idx, column=1).value = "Año"
                emp_ws.cell(row=row_idx, column=2).value = "Mes"
                emp_ws.cell(row=row_idx, column=3).value = "Valor Total"
                
                # Estilo encabezados
                for col in range(1, 4):
                    cell = emp_ws.cell(row=row_idx, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                # Calcular consolidado mensual
                monthly_data = {}
                for vale in vales_empleado:
                    if vale["fecha"] is not None and isinstance(vale["fecha"], datetime.datetime):
                        try:
                            year = vale["fecha"].year
                            month = vale["fecha"].month
                            month_key = f"{year}-{month:02d}"
                        except Exception as e:
                            print(f"Error al procesar fecha para consolidado mensual: {str(e)}")
                            continue  # Salta este vale
                        
                        if month_key not in monthly_data:
                            monthly_data[month_key] = {
                                "year": year,
                                "month": month,
                                "total": 0
                            }
                        
                        # Convertir valor a número antes de sumar
                        if vale["valor"] is not None:
                            try:
                                # Si es string, convertir a número
                                if isinstance(vale["valor"], str):
                                    valor_numerico = ''.join(c for c in vale["valor"] if c.isdigit() or c == '.')
                                    if valor_numerico:
                                        if '.' in valor_numerico:
                                            monthly_data[month_key]["total"] += float(valor_numerico)
                                        else:
                                            monthly_data[month_key]["total"] += int(valor_numerico)
                                else:
                                    # Si ya es número, sumar directamente
                                    monthly_data[month_key]["total"] += float(vale["valor"])
                            except (ValueError, TypeError):
                                print(f"Error al convertir valor '{vale['valor']}' para suma mensual")
                
                # Nombres de meses
                month_names = [
                    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
                ]
                
                # Llenar datos mensuales
                row_idx += 1
                for month_key, data in sorted(monthly_data.items()):
                    emp_ws.cell(row=row_idx, column=1).value = data["year"]
                    emp_ws.cell(row=row_idx, column=2).value = month_names[data["month"] - 1]
                    
                    valor_cell = emp_ws.cell(row=row_idx, column=3)
                    valor_cell.value = data["total"]
                    valor_cell.number_format = '#,##0'
                    
                    row_idx += 1
                
                # Espacio entre tablas
                row_idx += 2
                
                # Tabla Anual
                emp_ws.cell(row=row_idx, column=1).value = "CONSOLIDADO ANUAL"
                emp_ws.cell(row=row_idx, column=1).font = Font(bold=True)
                emp_ws.merge_cells(f'A{row_idx}:B{row_idx}')
                
                row_idx += 1
                emp_ws.cell(row=row_idx, column=1).value = "Año"
                emp_ws.cell(row=row_idx, column=2).value = "Valor Total"
                
                # Estilo encabezados
                for col in range(1, 3):
                    cell = emp_ws.cell(row=row_idx, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                # Calcular consolidado anual
                yearly_data = {}
                for vale in vales_empleado:
                    if vale["fecha"] is not None and isinstance(vale["fecha"], datetime.datetime):
                        try:
                            year = vale["fecha"].year
                        except Exception as e:
                            print(f"Error al procesar fecha para consolidado anual: {str(e)}")
                            continue  # Salta este vale
                        
                        if year not in yearly_data:
                            yearly_data[year] = 0
                        
                        # Convertir valor a número antes de sumar
                        if vale["valor"] is not None:
                            try:
                                # Si es string, convertir a número
                                if isinstance(vale["valor"], str):
                                    valor_numerico = ''.join(c for c in vale["valor"] if c.isdigit() or c == '.')
                                    if valor_numerico:
                                        if '.' in valor_numerico:
                                            yearly_data[year] += float(valor_numerico)
                                        else:
                                            yearly_data[year] += int(valor_numerico)
                                else:
                                    # Si ya es número, sumar directamente
                                    yearly_data[year] += float(vale["valor"])
                            except (ValueError, TypeError):
                                print(f"Error al convertir valor '{vale['valor']}' para suma anual")
                
                # Llenar datos anuales
                row_idx += 1
                for year, total in sorted(yearly_data.items()):
                    emp_ws.cell(row=row_idx, column=1).value = year
                    
                    valor_cell = emp_ws.cell(row=row_idx, column=2)
                    valor_cell.value = total
                    valor_cell.number_format = '#,##0'
                    
                    row_idx += 1
            
            # Guardar el archivo Excel
            wb.save(self.excel_path)
            
            QMessageBox.information(self, "Éxito", "Se han actualizado los reportes de todos los empleados correctamente.")
            print("Reportes de empleados actualizados con éxito.")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al actualizar reportes: {str(e)}")
            print(f"Error al actualizar reportes: {str(e)}")


    def setup_update_button(self):
        """
        Agrega un botón ActualizarDB a la interfaz para actualizar los reportes de empleados
        """
        if not hasattr(self.ui, 'btnActualizarDB'):
            # Crear el botón
            self.ui.btnActualizarDB = QPushButton("ActualizarDB")
            self.ui.btnActualizarDB.clicked.connect(self.update_employee_reports)
            
            # Añadir estilos al botón (opcional)
            self.ui.btnActualizarDB.setMinimumWidth(120)
            self.ui.btnActualizarDB.setStyleSheet("""
                QPushButton {
                    background-color: #3498db;
                    color: white;
                    border-radius: 5px;
                    padding: 5px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #2980b9;
                }
            """)
            
            # Añadir el botón a la interfaz
            # Puedes modificar esta parte según el layout de tu interfaz
            if hasattr(self.ui, 'verticalLayout'):
                # Crear un contenedor para el botón
                button_container = QWidget()
                button_layout = QHBoxLayout(button_container)
                button_layout.addStretch()
                button_layout.addWidget(self.ui.btnActualizarDB)
                
                # Añadir el contenedor al layout principal
                # Puedes ajustar la posición según necesites
                self.ui.verticalLayout.addWidget(button_container)
                
        # Asegurarse de que se importen los módulos necesarios
        # Añadir estos imports al inicio del archivo si no están ya
        # from openpyxl.styles import Font, PatternFill, Alignment
    def on_code_type_changed(self, checked):
        """Handle change in code type selection"""
        if checked:
            self.current_code_type = "barcode"
        else:
            self.current_code_type = "qr"
            
    def shorten_serial_code(self, code: str, length: int = 8) -> str:
        """Return a short hash from a given code"""
        hash_object = hashlib.sha1(code.encode())
        short_hash = hash_object.hexdigest()[:length]
        return short_hash.upper()
        
    def generate_serial_code(self):
        """Generate a unique serial code based on form fields"""
        # Get field values
        tipo_trabajo = self.ui.CampoTipoTrabajo.text().strip() if hasattr(self.ui, 'CampoTipoTrabajo') else ""
        referencia = self.ui.CampoReferenciaTrabajo.text().strip() if hasattr(self.ui, 'CampoReferenciaTrabajo') else ""
        num_ticket = self.ui.CampoNumeroTicket.text().strip() if hasattr(self.ui, 'CampoNumeroTicket') else ""
        talla = self.ui.CampoTalla.text().strip() if hasattr(self.ui, 'CampoTalla') else ""
        color = self.ui.CampoColor.text().strip() if hasattr(self.ui, 'CampoColor') else ""
        
        # Generate a UUID and take first 6 characters
        unique_id = str(uuid.uuid4())[:6]
        
        # Create the compact serial code
        serial_code = f"{tipo_trabajo[:1]}{referencia[:2]}{num_ticket[:3]}{talla[:1]}{color[:1]}{unique_id}"
        return serial_code.upper()
    

    def generate_barcode(self, serial_code):
        try:
            # Ruta absoluta a fuente embebida (úsala como recurso)
            font_path = os.path.join(os.path.dirname(__file__), "app/fonts/DejaVuSansMono.ttf")
            font = ImageFont.truetype(font_path, size=7)

            barcode_class = barcode.get_barcode_class('code128')
            writer = ImageWriter()

            writer.font = font  # asignación directa

            options = {
                'module_height': 8.0,
                'module_width': 0.2,
                'quiet_zone': 1.0,
                'font_size': 7,
                'text_distance': 1.0,
            }

            for key, val in options.items():
                if hasattr(writer, key):
                    setattr(writer, key, val)

            barcode_instance = barcode_class(serial_code, writer=writer)
            barcode_path = os.path.join("codes", f"barcode_{serial_code}")
            barcode_filename = barcode_instance.save(barcode_path)

            return barcode_filename
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al generar el código de barras: {str(e)}")
            return None

    
    def generate_qr_code(self, serial_code):
        """Generate and save a QR code image"""
        try:
            # Create QR code instance
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            
            # Add data to QR code
            qr.add_data(serial_code)
            qr.make(fit=True)
            
            # Create QR code image
            img = qr.make_image(fill_color="black", back_color="white")
            
            # Save QR code as PNG
            qr_path = f"codes/qr_{serial_code}.png"
            img.save(qr_path)
            
            return qr_path
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al generar el código QR: {str(e)}")
            return None
    
    def display_code_image(self, image_path):
        """Display the code image in the QGraphicsView"""
        try:
            if not hasattr(self.ui, 'PreviwImage'):
                print("Error: PreviwImage not found in UI")
                return False
                
            # Load image
            pixmap = QPixmap(image_path)
            
            # Get scene from QGraphicsView
            scene = self.ui.PreviwImage.scene()
            if scene is None:
                scene = QtWidgets.QGraphicsScene()
                self.ui.PreviwImage.setScene(scene)
            else:
                scene.clear()
            
            # Add image to scene
            scene.addPixmap(pixmap)
            
            # Fit view
            self.ui.PreviwImage.fitInView(scene.sceneRect(), Qt.KeepAspectRatio)
            
            return True
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al mostrar la imagen: {str(e)}")
            return False
    
    def save_to_excel(self, serial_code, code_path):
        """Save data to Excel file"""
        try:
            # Get field values
            tipo_trabajo = self.ui.CampoTipoTrabajo.text().strip() if hasattr(self.ui, 'CampoTipoTrabajo') else ""
            referencia = self.ui.CampoReferenciaTrabajo.text().strip() if hasattr(self.ui, 'CampoReferenciaTrabajo') else ""
            num_ticket = self.ui.CampoNumeroTicket.text().strip() if hasattr(self.ui, 'CampoNumeroTicket') else ""
            talla = self.ui.CampoTalla.text().strip() if hasattr(self.ui, 'CampoTalla') else ""
            color = self.ui.CampoColor.text().strip() if hasattr(self.ui, 'CampoColor') else ""
            valor = self.ui.CampoValor.text().strip() if hasattr(self.ui, 'CampoValor') else ""
            total_producido = self.ui.CampoTotaProducido.text().strip() if hasattr(self.ui, 'CampoTotaProducido') else ""
            
            # Load Excel file
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            # Add new row
            ws.append([serial_code, tipo_trabajo, referencia, num_ticket, 
                      talla, color, valor, total_producido, self.current_code_type, code_path])
            
            # Save Excel file
            wb.save(self.excel_path)
            
            return True
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al guardar en Excel: {str(e)}")
            return False
    
    def find_code_data(self, serial_code):
        """Find data related to a specific serial code in the Excel file"""
        try:
            # Load Excel file
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            # Search for the serial code in the first column
            for row in ws.iter_rows(min_row=2):  # Start from row 2 to skip header
                if row[0].value == serial_code:
                    # Return all data for this row
                    return {
                        "serial_code": row[0].value,
                        "tipo_trabajo": row[1].value,
                        "referencia": row[2].value,
                        "num_ticket": row[3].value,
                        "talla": row[4].value,
                        "color": row[5].value,
                        "valor": row[6].value,
                        "total_producido": row[7].value,
                        "tipo_codigo": row[8].value,
                        "ruta_imagen": row[9].value
                    }
            
            return None
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al buscar datos: {str(e)}")
            return None
    
    def on_code_scanned(self):
        """Handle barcode/QR code scanning"""
        # Get scanned code
        scanned_code = self.ui.codeReaderInput.text().strip()
        if not scanned_code:
            return
        
        # Find data for this code
        data = self.find_code_data(scanned_code)
        if not data:
            QMessageBox.warning(self, "Código No Encontrado", 
                               f"No se encontraron datos para el código: {scanned_code}")
            self.ui.codeReaderInput.clear()
            return
        
        # Generate a unique ID for this vale
        vale_id = f"V-{datetime.datetime.now().strftime('%Y%m%d')}-{str(uuid.uuid4())[:6]}"
        
        # Add data to table view
        self.add_to_table_view(vale_id, data)
        
        # Clear input field
        self.ui.codeReaderInput.clear()
        
        # Display code image if available
        if data["ruta_imagen"] and os.path.exists(data["ruta_imagen"]):
            self.display_code_image(data["ruta_imagen"])
            
        # Show success message
        QMessageBox.information(self, "Código Encontrado", 
                               f"Código escaneado: {scanned_code}\nTipo de trabajo: {data['tipo_trabajo']}\nVale ID: {vale_id}")
    
    def add_to_table_view(self, vale_id, data):
        """Add scanned code data to the TableView"""
        # Get employee info
        employee_text = self.ui.EmpleadosBox.currentText()
        employee_id = self.ui.EmpleadosBox.currentData()
        
        # Create row items
        items = [
            QtGui.QStandardItem(vale_id),
            QtGui.QStandardItem(data["serial_code"]),
            QtGui.QStandardItem(str(data["tipo_trabajo"])),
            QtGui.QStandardItem(str(data["referencia"])),
            QtGui.QStandardItem(str(data["num_ticket"])),
            QtGui.QStandardItem(str(data["talla"])),
            QtGui.QStandardItem(str(data["color"])),
            QtGui.QStandardItem(str(data["valor"])),
            QtGui.QStandardItem(str(data["total_producido"])),
            QtGui.QStandardItem(employee_text)
        ]
        
        # Add row to model
        self.table_model.appendRow(items)
        
        # Store data for later registration
        data["vale_id"] = vale_id
        data["employee_id"] = employee_id
        data["employee_text"] = employee_text
        
        # Add to pending vales list if not already exists
        if not hasattr(self, 'pending_vales'):
            self.pending_vales = []
        
        self.pending_vales.append(data)
    
    def register_vale(self):
        """Register the scanned vales to Excel"""
        if not hasattr(self, 'pending_vales') or not self.pending_vales:
            QMessageBox.warning(self, "Sin Vales", "No hay vales pendientes para registrar.")
            return
        
        try:
            # Get current date
            current_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Load Excel file
            wb = load_workbook(self.excel_path)
            
            # Get vales sheet or create if not exists
            if self.vales_sheet_name in wb.sheetnames:
                ws = wb[self.vales_sheet_name]
            else:
                ws = wb.create_sheet(title=self.vales_sheet_name)
                ws.append(["ID", "EmpleadoID", "Fecha", "Tipo Trabajo", "Número Ticket",
                          "Referencia", "Talla", "Color", "Valor", "Total Producido", "Código Serial"])
            
            # Add all pending vales
            for vale in self.pending_vales:
                ws.append([
                    vale["vale_id"],
                    vale["employee_id"],
                    current_date,
                    vale["tipo_trabajo"],
                    vale["num_ticket"],
                    vale["referencia"],
                    vale["talla"],
                    vale["color"],
                    vale["valor"],
                    vale["total_producido"],
                    vale["serial_code"]
                ])
            
            # Save Excel file
            wb.save(self.excel_path)
            
            # Clear pending vales and table view
            self.pending_vales = []
            self.table_model.removeRows(0, self.table_model.rowCount())
            
            # Show success message
            QMessageBox.information(self, "Vales Registrados", 
                                   f"Se han registrado {len(self.pending_vales)} vales correctamente.")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al registrar vales: {str(e)}")
    
    def on_save_button_clicked(self):
        """Handler for save button click"""
        # Define required fields and their UI elements
        required_fields = []
        
        # Only add fields that exist in the UI
        if hasattr(self.ui, 'CampoTipoTrabajo'):
            required_fields.append((self.ui.CampoTipoTrabajo, "Tipo de Trabajo"))
        if hasattr(self.ui, 'CampoReferenciaTrabajo'):
            required_fields.append((self.ui.CampoReferenciaTrabajo, "Referencia"))
        if hasattr(self.ui, 'CampoNumeroTicket'):
            required_fields.append((self.ui.CampoNumeroTicket, "Número de Ticket"))
        if hasattr(self.ui, 'CampoTalla'):
            required_fields.append((self.ui.CampoTalla, "Talla"))
        if hasattr(self.ui, 'CampoColor'):
            required_fields.append((self.ui.CampoColor, "Color"))
        if hasattr(self.ui, 'CampoValor'):
            required_fields.append((self.ui.CampoValor, "Valor"))
        if hasattr(self.ui, 'CampoTotaProducido'):
            required_fields.append((self.ui.CampoTotaProducido, "Total Producido"))
        
        # Validate required fields
        for field, name in required_fields:
            if not field.text().strip():
                QMessageBox.warning(self, "Campos Incompletos", f"El campo {name} es obligatorio.")
                field.setFocus()
                return
        
        # Generate unique serial code
        serial_code = self.generate_serial_code()
        
        # Generate the selected code type
        if self.current_code_type == "barcode":
            code_path = self.generate_barcode(serial_code)
        else:
            code_path = self.generate_qr_code(serial_code)
            
        if not code_path:
            return
        
        # Display the code
        if not self.display_code_image(code_path):
            return
        
        # Save data to Excel
        if self.save_to_excel(serial_code, code_path):
            QMessageBox.information(
                self, 
                "Operación Exitosa", 
                f"Código generado: {serial_code}\nTipo: {self.current_code_type.upper()}\nLos datos se han guardado correctamente."
            )
            
            # Clear fields
            for field, _ in required_fields:
                field.clear()
            
            # Focus first field
            if required_fields:
                required_fields[0][0].setFocus()

    def Button(self):
        """Handler for menu button clicks"""
        # Get clicked button
        btnWidget = self.sender()

        # Handle different pages
        if btnWidget.objectName() == "btn_home":
            self.ui.stackedWidget.setCurrentWidget(self.ui.page_home)
            UIFunctions.resetStyle(self, "btn_home")
            btnWidget.setStyleSheet(UIFunctions.selectMenu(btnWidget.styleSheet()))
        elif btnWidget.objectName() == "btn_new_user":
            self.ui.stackedWidget.setCurrentWidget(self.ui.page_widgets)
            UIFunctions.resetStyle(self, "btn_widgets")
            btnWidget.setStyleSheet(UIFunctions.selectMenu(btnWidget.styleSheet()))
        elif btnWidget.objectName() == "btn_widgets":
            self.ui.stackedWidget.setCurrentWidget(self.ui.page_widgets)
            UIFunctions.resetStyle(self, "btn_widgets")
            btnWidget.setStyleSheet(UIFunctions.selectMenu(btnWidget.styleSheet()))

    # Window event handlers
    def moveWindow(self, event):
        """Handle window movement"""
        # If maximized change to normal
        if UIFunctions.returStatus() == 1:
            UIFunctions.maximize_restore(self)

        # Move window
        if event.buttons() == Qt.LeftButton:
            self.move(self.pos() + event.globalPos() - self.dragPos)
            self.dragPos = event.globalPos()
            event.accept()

    def eventFilter(self, watched, event):
        """Event filter for double-click events"""
        if hasattr(self, 'le') and watched == self.le and event.type() == QtCore.QEvent.MouseButtonDblClick:
            print("pos: ", event.pos())
        return super().eventFilter(watched, event)

    def mousePressEvent(self, event):
        """Handle mouse press events"""
        self.dragPos = event.globalPos()
        if event.buttons() == Qt.LeftButton:
            print('Mouse click: LEFT CLICK')
        if event.buttons() == Qt.RightButton:
            print('Mouse click: RIGHT CLICK')
        if event.buttons() == Qt.MidButton:
            print('Mouse click: MIDDLE BUTTON')

    def keyPressEvent(self, event):
        """Handle key press events"""
        print('Key: ' + str(event.key()) + ' | Text Press: ' + str(event.text()))

    def resizeEvent(self, event):
        """Handle resize events"""
        self.resizeFunction()
        return super(MainWindow, self).resizeEvent(event)

    def resizeFunction(self):
        """Log window size on resize"""
        print('Height: ' + str(self.height()) + ' | Width: ' + str(self.width()))



if __name__ == "__main__":
    app = QApplication(sys.argv)
    QtGui.QFontDatabase.addApplicationFont('fonts/segoeui.ttf')
    QtGui.QFontDatabase.addApplicationFont('fonts/segoeuib.ttf')
    window = MainWindow()
    sys.exit(app.exec_())