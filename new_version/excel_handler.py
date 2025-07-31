import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from PySide2.QtWidgets import QMessageBox

class ExcelHandler:
    def __init__(self, excel_path="trabajos_database.xlsx"):
        self.excel_path = excel_path
        self.trabajos_sheet_name = "Trabajos"
        self.vales_sheet_name = "Vales"
        self.empleados_sheet_name = "Empleados"

    def inicializar_database(self, trabajos_headers, vales_headers, empleados_headers):
        """
        Verifica que el archivo Excel y las hojas necesarias existan.
        Si no existen, los crea con sus cabeceras.
        """
        if not os.path.exists(self.excel_path):
            wb = Workbook()
            # Hoja de Trabajos
            ws_trabajos = wb.active
            ws_trabajos.title = self.trabajos_sheet_name
            ws_trabajos.append(trabajos_headers)
            # Hoja de Vales
            ws_vales = wb.create_sheet(title=self.vales_sheet_name)
            ws_vales.append(vales_headers)
            # Hoja de Empleados
            ws_empleados = wb.create_sheet(title=self.empleados_sheet_name)
            ws_empleados.append(empleados_headers)
            wb.save(self.excel_path)
            print(f"Archivo '{self.excel_path}' creado con todas las hojas.")
        else:
            # Verificar y crear hojas si faltan en un archivo existente
            wb = load_workbook(self.excel_path)
            sheets_to_check = {
                self.trabajos_sheet_name: trabajos_headers,
                self.vales_sheet_name: vales_headers,
                self.empleados_sheet_name: empleados_headers,
            }
            for sheet_name, headers in sheets_to_check.items():
                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(title=sheet_name)
                    ws.append(headers)
                    print(f"Hoja '{sheet_name}' creada.")
            wb.save(self.excel_path)

    def guardar_trabajo(self, row_data):
        """Guarda una nueva fila en la hoja 'Trabajos'."""
        try:
            wb = load_workbook(self.excel_path)
            ws = wb[self.trabajos_sheet_name]
            ws.append(row_data)
            wb.save(self.excel_path)
            return True
        except PermissionError:
            QMessageBox.critical(None, "Error al Guardar", f"Permiso denegado para escribir en '{self.excel_path}'. Asegúrese de que no esté abierto.")
            return False
        except Exception as e:
            QMessageBox.critical(None, "Error de Excel", f"Error al guardar trabajo: {e}")
            return False

    def guardar_vale(self, vale_data):
        """Guarda un nuevo vale en la hoja 'Vales'."""
        try:
            wb = load_workbook(self.excel_path)
            ws = wb[self.vales_sheet_name]
            ws.append(vale_data)
            wb.save(self.excel_path)
            return True
        except Exception as e:
            QMessageBox.critical(None, "Error de Excel", f"Error al guardar vale: {e}")
            return False

    def guardar_empleado(self, empleado_data):
        """Guarda un nuevo empleado en la hoja 'Empleados'."""
        try:
            wb = load_workbook(self.excel_path)
            ws = wb[self.empleados_sheet_name]
            ws.append(empleado_data)
            wb.save(self.excel_path)
            return True
        except Exception as e:
            QMessageBox.critical(None, "Error de Excel", f"Error al guardar empleado: {e}")
            return False
            
    def leer_empleados(self):
        """Lee todos los empleados de la hoja 'Empleados'."""
        try:
            wb = load_workbook(self.excel_path, read_only=True)
            if self.empleados_sheet_name not in wb.sheetnames:
                return []
            ws = wb[self.empleados_sheet_name]
            empleados = []
            # min_row=2 para saltar la cabecera
            for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
                 if row[0] and row[4]: # Asegurarse que Nombre y ID existan
                    empleados.append({"nombre": row[0], "id": row[4]})
            return empleados
        except Exception as e:
            print(f"Error al leer empleados: {e}")
            return []

    # ... Aquí irían más funciones de lectura como find_code_data, is_vale_registered, etc.
    # Por ejemplo:
    def buscar_trabajo_por_codigo(self, scanned_code, work_type_columns):
        """Busca un trabajo por su código serial en cualquiera de las columnas de tipo de trabajo."""
        try:
            wb = load_workbook(self.excel_path, read_only=True)
            ws = wb[self.trabajos_sheet_name]
            headers = [cell.value for cell in ws[1]]
            
            for row_values in ws.iter_rows(min_row=2, values_only=True):
                for work_type, column_name in work_type_columns.items():
                    try:
                        col_idx = headers.index(column_name)
                        if row_values[col_idx] == scanned_code:
                            # Encontrado! Retornamos la fila y el tipo de trabajo.
                            return dict(zip(headers, row_values)), work_type
                    except (ValueError, IndexError):
                        continue # La columna no existe o la fila es corta, seguir buscando
            return None, None # No encontrado
        except Exception as e:
            QMessageBox.critical(None, "Error de Búsqueda", f"Error al buscar código: {e}")
            return None, None