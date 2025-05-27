################################################################################
## V: 1.0.0
################################################################################

import sys
import os
import platform
import uuid
import hashlib
import traceback
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

import qrcode
import barcode
from barcode.writer import ImageWriter
from PIL import ImageFont

from PySide2 import QtCore, QtGui, QtWidgets
from PySide2.QtCore import (
    QCoreApplication, QPropertyAnimation, QDate, QDateTime, QMetaObject, 
    QObject, QPoint, QRect, QSize, QTime, QThread, QUrl, Qt, QEvent
)
from PySide2.QtGui import (
    QBrush, QColor, QConicalGradient, QCursor, QFont, QFontDatabase, 
    QIcon, QKeySequence, QLinearGradient, QPalette, QPainter, QPixmap, QRadialGradient
)
from PySide2.QtWidgets import *

import shutil
import zipfile

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from reportlab.lib.units import inch, mm
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

# GUI FILE
from app_modules import *


class MainWindow(QMainWindow):
    def __init__(self):
        QMainWindow.__init__(self)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.dragPos = None
        
        # Current code type (barcode or qr)
        self.current_code_type = "barcode"  # Default to barcode

        # Print system information
        print('System: ' + platform.system())
        print('Version: ' + platform.release())

        # Setup the window
        self.setup_window()
        
        # Setup menus
        self.setup_menus()
        
        # Setup barcode generator functionality
        self.setup_code_generator()
        
        # Setup code reader functionality
        self.setup_code_reader()
        
        # Setup employee management
        self.setup_employee_management()

        # Conectar el botón de eliminarTODO (agregar esta línea)
        self.setup_eliminar_todo_button()

        # Setup update reports button
        self.setup_update_button()
        
        # Setup add employee button
        self.setup_add_employee_button()

        self.autocompletado_manager = AutocompletadoManager(excel_path=self.excel_path, sheet_name="Trabajos")
        self.setup_autocompletado_fields()

        # Show the window
        self.show()

    def setup_window(self):
        """Setup basic window properties and appearance"""
        # Remove standard title bar
        UIFunctions.removeTitleBar(True)
        
        # Set window title
        self.setWindowTitle('Gestor de Vales')
        UIFunctions.labelTitle(self, 'Thimoty')
        UIFunctions.labelDescription(self, '2025')
        
        # Set window size
        startSize = QSize(1000, 720)
        self.resize(startSize)
        self.setMinimumSize(startSize)
        # UIFunctions.enableMaximumSize(self, 500, 720)

    def setup_menus(self):
        """Setup menu functionality and navigation"""
        # Toggle menu size button
        self.ui.btn_toggle_menu.clicked.connect(lambda: UIFunctions.toggleMenu(self, 220, True))
        
        # Add custom menus
        self.ui.stackedWidget.setMinimumWidth(20)
        UIFunctions.addNewMenu(self, "Leer Vales", "btn_home", "url(:/16x16/icons/16x16/cil-home.png)", True)
        UIFunctions.addNewMenu(self, "Crear Vales", "btn_new_user", "url(:/16x16/icons/16x16/cil-user-follow.png)", True)
        UIFunctions.addNewMenu(self, "Configuracion", "btn_widgets", "url(:/16x16/icons/16x16/cil-equalizer.png)", False)
        
        # Select starting menu
        UIFunctions.selectStandardMenu(self, "btn_home")
        
        # Set starting page
        self.ui.stackedWidget.setCurrentWidget(self.ui.page_home)

    def setup_eliminar_todo_button(self):
        """Conecta el botón EliminarTODO con la función de eliminación"""
        self.ui.EliminarTODO.clicked.connect(self.eliminar_todo_con_backup)

    def eliminar_todo_con_backup(self):
        """
        Elimina la base de datos Excel y el directorio codes después de crear un backup
        Versión con threading para UI más responsiva
        """
        try:
            # Confirmar acción con el usuario
            reply = QMessageBox.question(
                self, 
                'Confirmar Eliminación', 
                '¿Está seguro de que desea eliminar toda la base de datos?\n\n'
                'Se creará un backup automáticamente antes de eliminar.',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
                
            # Obtener rutas usando BackupManager
            excel_path, codes_dir = BackupManager.get_paths()
            
            # Verificar si hay algo que respaldar
            has_files, excel_exists, codes_exists = BackupManager.check_files_exist(excel_path, codes_dir)
            
            if not has_files:
                QMessageBox.information(
                    self, 
                    'Información', 
                    'No hay archivos para eliminar. La base de datos ya está limpia.'
                )
                return
            
            # Deshabilitar el botón mientras se procesa
            self.ui.EliminarTODO.setEnabled(False)
            self.ui.EliminarTODO.setText("Procesando...")
            QApplication.processEvents()
            
            # Crear backup en thread separado
            self.backup_thread = BackupThread(excel_path, codes_dir)
            self.backup_thread.finished.connect(self.on_backup_finished)
            self.backup_thread.progress.connect(self.on_backup_progress)
            self.backup_thread.start()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al iniciar eliminación: {str(e)}")
            self.ui.EliminarTODO.setEnabled(True)
            self.ui.EliminarTODO.setText("Eliminar TODO")
    
    def eliminar_todo_simple(self):
        """
        Versión simple sin threading - buena para pocos archivos
        """
        try:
            # Confirmar acción
            reply = QMessageBox.question(
                self, 
                'Confirmar Eliminación', 
                '¿Está seguro de que desea eliminar toda la base de datos?\n\n'
                'Se creará un backup automáticamente.',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
            
            # Cambiar botón para mostrar progreso
            self.ui.EliminarTODO.setEnabled(False)
            original_text = self.ui.EliminarTODO.text()
            
            def update_progress(message):
                self.ui.EliminarTODO.setText(message)
                QApplication.processEvents()
            
            # Obtener rutas
            excel_path, codes_dir = BackupManager.get_paths()
            
            # Verificar archivos
            has_files, excel_exists, codes_exists = BackupManager.check_files_exist(excel_path, codes_dir)
            
            if not has_files:
                QMessageBox.information(self, 'Información', 'No hay archivos para eliminar.')
                return
            
            # Crear backup
            update_progress("Creando backup...")
            backup_name = BackupManager.create_backup_sync(excel_path, codes_dir, update_progress)
            
            # Eliminar archivos originales
            update_progress("Eliminando archivos...")
            deleted_files = BackupManager.delete_files(excel_path, codes_dir)
            
            # Recrear estructura
            update_progress("Recreando estructura...")
            self.setup_code_generator()
            
            QMessageBox.information(
                self, 
                'Completado', 
                f'Backup creado: {backup_name}\n'
                f'Archivos eliminados: {", ".join(deleted_files)}'
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error durante la eliminación: {str(e)}")
        
        finally:
            # Restaurar botón
            self.ui.EliminarTODO.setEnabled(True)
            self.ui.EliminarTODO.setText("Eliminar TODO")
    
    def on_backup_progress(self, message):
        """Callback para actualizar progreso del backup"""
        self.ui.EliminarTODO.setText(message)
    
    def on_backup_finished(self, success, message):
        """Callback cuando termina el backup"""
        try:
            if success:
                # Backup exitoso, proceder con eliminación
                self.ui.EliminarTODO.setText("Eliminando archivos...")
                QApplication.processEvents()
                
                excel_path, codes_dir = BackupManager.get_paths()
                deleted_files = BackupManager.delete_files(excel_path, codes_dir)
                
                # Recrear estructura
                self.ui.EliminarTODO.setText("Recreando estructura...")
                QApplication.processEvents()
                self.setup_code_generator()
                
                QMessageBox.information(
                    self, 
                    'Completado', 
                    f'{message}\n'
                    f'Archivos eliminados: {", ".join(deleted_files)}'
                )
            else:
                # Error en backup
                QMessageBox.critical(self, "Error en Backup", message)
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error en proceso de eliminación: {str(e)}")
        
        finally:
            # Rehabilitar botón
            self.ui.EliminarTODO.setEnabled(True)
            self.ui.EliminarTODO.setText("Eliminar TODO")

    def setup_code_generator(self):
        """
        Setup the barcode and QR code generator functionality
        """
        # Define excel path and create barcode directory if needed
        self.excel_path = "trabajos_database.xlsx"
        self.vales_sheet_name = "Vales"  # Sheet name for vales data
        
        if not os.path.exists("codes"):
            os.makedirs("codes")
        
        # Create Excel file if it doesn't exist
        if not os.path.exists(self.excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Trabajos"
            ws.append(["Código Serial", "Tipo Trabajo", "Referencia", "Número Ticket",
                     "Talla", "Color", "Valor", "Total Producido", "Tipo Código", "Ruta Imagen"])
            
            # Create Vales sheet
            vales_ws = wb.create_sheet(title=self.vales_sheet_name)
            vales_ws.append(["ID", "EmpleadoID", "Fecha", "Tipo Trabajo", "Número Ticket",
                          "Referencia", "Talla", "Color", "Valor", "Total Producido", "Código Serial"])
            
            wb.save(self.excel_path)
        else:
            # Check if Vales sheet exists, if not create it
            wb = load_workbook(self.excel_path)
            if self.vales_sheet_name not in wb.sheetnames:
                vales_ws = wb.create_sheet(title=self.vales_sheet_name)
                vales_ws.append(["ID", "EmpleadoID", "Fecha", "Tipo Trabajo", "Número Ticket",
                              "Referencia", "Talla", "Color", "Valor", "Total Producido", "Código Serial"])
                wb.save(self.excel_path)
        
        # Connect the save button
        self.ui.pushButtonGuardar.clicked.connect(self.on_save_button_clicked)
        
        # Add code type selector (radio buttons)
        self.setup_code_type_selector()
        
        # Ensure the QGraphicsView has a scene
        if not hasattr(self.ui, 'PreviwImage'):
            print("Warning: PreviwImage not found in UI")
        elif self.ui.PreviwImage.scene() is None:
            self.ui.PreviwImage.setScene(QtWidgets.QGraphicsScene())
    
    def setup_code_type_selector(self):
        """Setup UI elements to select between barcode and QR code"""
        # Create a frame for radio buttons if it doesn't exist
        if not hasattr(self.ui, 'codeTypeFrame'):
            self.ui.codeTypeFrame = QFrame()
            self.ui.codeTypeFrame.setFrameShape(QFrame.StyledPanel)
            self.ui.codeTypeFrame.setFrameShadow(QFrame.Raised)
            
            # Create radio buttons
            self.ui.radioBarcode = QRadioButton("Código de Barras")
            self.ui.radioQR = QRadioButton("Código QR")
            self.ui.radioBarcode.setChecked(True)  # Barcode is default
            
            # Create layout
            layout = QHBoxLayout(self.ui.codeTypeFrame)
            layout.addWidget(self.ui.radioBarcode)
            layout.addWidget(self.ui.radioQR)
            
            # Connect signals
            self.ui.radioBarcode.toggled.connect(self.on_code_type_changed)
            
            # Find a place to add the frame (this depends on your UI layout)
            # For example, if there's a verticalLayout in the form:
            if hasattr(self.ui, 'formLayout'):
                # Insert at position 0 (top)
                self.ui.formLayout.insertRow(0, "Tipo de Código:", self.ui.codeTypeFrame)
    
    def setup_code_reader(self):
        """Setup the code reader functionality and TableView inside WidgetTabla"""
        # Create a code reader input field if it doesn't exist
        if not hasattr(self.ui, 'codeReaderInput'):
            # Code reader input field
            self.ui.codeReaderInput = QLineEdit()
            self.ui.codeReaderInput.setPlaceholderText("Escanear o ingresar código aquí...")
            self.ui.codeReaderInput.setMinimumWidth(250)
            self.ui.codeReaderInput.returnPressed.connect(self.on_code_scanned)
            
            # Add to layout
            if hasattr(self.ui, 'verticalLayout'):
                # Create a label for the input
                label = QLabel("Lector de Códigos:")
                
                # Create horizontal layout for label and input
                h_layout = QHBoxLayout()
                h_layout.addWidget(label)
                h_layout.addWidget(self.ui.codeReaderInput)
                
                # Create a container widget
                container = QWidget()
                container.setLayout(h_layout)
                
                # Add to vertical layout
                self.ui.verticalLayout.addWidget(container)
        
        # Access the WidgetTabla from UI and setup table inside ac
        if hasattr(self.ui, 'WidgetTabla'):
            # Create table layout
            table_layout = QVBoxLayout(self.ui.WidgetTabla)
            self.ui.WidgetTabla.setLayout(table_layout)
            
            # Create TableView
            self.ui.tableViewVale = QTableView(self.ui.WidgetTabla)
            self.ui.tableViewVale.setMinimumHeight(200)
            
            # Create the model for the table
            self.table_model = QtGui.QStandardItemModel()
            self.table_model.setHorizontalHeaderLabels([
                "ID", "Código Serial", "Tipo Trabajo", "Referencia", "Número Ticket", 
                "Talla", "Color", "Valor", "Total Producido", "Empleado"
            ])
            
            # Set model and adjust view
            self.ui.tableViewVale.setModel(self.table_model)
            self.ui.tableViewVale.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            
            # Add TableView to WidgetTabla's layout
            table_layout.addWidget(self.ui.tableViewVale)
        else:
            print("Warning: WidgetTabla not found in UI")


    def setup_add_employee_button(self):
        """Conecta el botón de agregar empleado (si lo tienes)"""
        if hasattr(self.ui, 'btnAgregarEmpleado'):
            self.ui.btnAgregarEmpleado.clicked.connect(self.add_employee)
            print("Botón 'btnAgregarEmpleado' conectado correctamente.")
        else:
            print("Botón 'btnAgregarEmpleado' no encontrado en la UI.")

    # Función adicional para validar cédula (opcional)
    def validate_cedula(self, cedula):
        """Valida que la cédula tenga formato correcto"""
        # Remover espacios y guiones
        cedula_clean = cedula.replace(" ", "").replace("-", "")
        
        # Verificar que solo contenga números y tenga longitud apropiada
        if not cedula_clean.isdigit() or len(cedula_clean) < 6 or len(cedula_clean) > 12:
            return False
        
        return True

    # Versión mejorada con validación
    def add_employee(self):
        """Agrega empleado con validaciones adicionales"""
        try:
            # Obtener valores de los campos
            nombre = self.ui.Nombre_Empleado.text().strip() if hasattr(self.ui, 'Nombre_Empleado') else ""
            cedula = self.ui.Cedula_Empleado.text().strip() if hasattr(self.ui, 'Cedula_Empleado') else ""
            celular = self.ui.Celular_Empleado.text().strip() if hasattr(self.ui, 'Celular_Empleado') else ""
            correo = self.ui.Correo_Empleado.text().strip() if hasattr(self.ui, 'Correo_Empleado') else ""
            
            # Validaciones
            if not nombre:
                QMessageBox.warning(self, "Campo Requerido", "El nombre es obligatorio.")
                self.ui.Nombre_Empleado.setFocus()
                return
                
            if not cedula:
                QMessageBox.warning(self, "Campo Requerido", "La cédula es obligatoria.")
                self.ui.Cedula_Empleado.setFocus()
                return
            
            # Validar formato de cédula
            if not self.validate_cedula(cedula):
                QMessageBox.warning(self, "Cédula Inválida", "La cédula debe contener solo números (6-12 dígitos).")
                self.ui.Cedula_Empleado.setFocus()
                return
            
            # Validar email si se proporciona
            if correo and "@" not in correo:
                QMessageBox.warning(self, "Email Inválido", "El formato del correo electrónico no es válido.")
                self.ui.Correo_Empleado.setFocus()
                return
            
            # Generar ID único del empleado
            unique_id = str(uuid.uuid4())[:4]
            empleado_id = f"E{cedula[-4:]}{unique_id}".upper()
            
            # Cargar Excel y verificar duplicados
            wb = load_workbook(self.excel_path)
            
            if "Empleados" not in wb.sheetnames:
                empleados_ws = wb.create_sheet(title="Empleados")
                empleados_ws.append(["Nombre", "Cedula", "Celular", "Correo", "EmpleadoId"])
            else:
                empleados_ws = wb["Empleados"]
            
            # Verificar si la cédula ya existe
            for row in empleados_ws.iter_rows(min_row=2, max_col=5):
                if row[1].value and str(row[1].value) == cedula:
                    QMessageBox.warning(self, "Empleado Existente", f"Ya existe un empleado con cédula {cedula}")
                    return
            
            # Agregar empleado
            empleados_ws.append([nombre, cedula, celular, correo, empleado_id])
            wb.save(self.excel_path)
            
            # Limpiar campos
            self.ui.Nombre_Empleado.clear()
            self.ui.Cedula_Empleado.clear()
            self.ui.Celular_Empleado.clear()
            self.ui.Correo_Empleado.clear()
            
            # Actualizar ComboBox
            if hasattr(self, 'setup_employee_management'):
                self.setup_employee_management()
            
            # Mensaje de éxito
            QMessageBox.information(self, "Empleado Agregado", f"Empleado: {nombre}\nCédula: {cedula}\nID: {empleado_id}")
            
            print(f"Empleado agregado exitosamente: {nombre} ({empleado_id})")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al agregar empleado: {str(e)}")
            print(f"Error en add_employee_with_validation: {e}")

    def setup_employee_management(self):
        """Carga empleados en el ComboBox y conecta el botón Registrar Vale."""
        try:
            # --- Manejo de Archivo Excel ---
            try:
                wb = load_workbook(self.excel_path)
            except FileNotFoundError:
                print(f"Advertencia: Archivo Excel no encontrado en {self.excel_path}. Creando uno nuevo.")
                wb = Workbook()
                # Usa la hoja activa si es un libro nuevo
                if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
                    empleados_ws = wb.active
                    empleados_ws.title = "Empleados"
                else:
                     empleados_ws = wb.create_sheet(title="Empleados")
                # Añadir encabezados al crear
                empleados_ws.append(["Nombre", "Cedula", "Celular", "Correo", "EmpleadoId"])
                # Puedes añadir un empleado de ejemplo si quieres
                empleados_ws.append(["Juan Pérez (Ejemplo)", "1234567890", "3001234567", "juan@example.com", "E001"])
                wb.save(self.excel_path)
                print(f"Archivo Excel '{self.excel_path}' creado con hoja 'Empleados'.")

            if "Empleados" not in wb.sheetnames:
                print(f"Creando hoja 'Empleados' en archivo existente: {self.excel_path}")
                empleados_ws = wb.create_sheet(title="Empleados")
                empleados_ws.append(["Nombre", "Cedula", "Celular", "Correo", "EmpleadoId"])
                # Puedes añadir un empleado de ejemplo si quieres
                empleados_ws.append(["Juan Pérez (Ejemplo)", "1234567890", "3001234567", "juan@example.com", "E001"])
                wb.save(self.excel_path)
            else:
                empleados_ws = wb["Empleados"]

            # --- Configuración del ComboBox (Asumiendo que existe en self.ui) ---
            if hasattr(self.ui, 'EmpleadosBox'):
                self.ui.EmpleadosBox.clear() # Limpiar items previos
                self.ui.EmpleadosBox.setMinimumWidth(200) # Establecer ancho mínimo si es necesario

                employees_loaded = False
                # Iterar desde la fila 2 para saltar encabezados (ajusta si no tienes encabezados)
                for row in empleados_ws.iter_rows(min_row=2, max_col=5): # Leer solo las 5 columnas necesarias
                    # row es una tupla de celdas
                    if len(row) >= 5 and row[0].value: # Verificar que hay nombre y suficientes columnas
                        name = str(row[0].value) # Columna A (Nombre)
                        emp_id = str(row[4].value) if row[4].value else "SIN-ID" # Columna E (EmpleadoId)

                        # Añadir item: Texto visible y dato asociado (emp_id)
                        self.ui.EmpleadosBox.addItem(f"{name} ({emp_id})", emp_id)
                        employees_loaded = True

                if not employees_loaded:
                    self.ui.EmpleadosBox.addItem("Sin empleados registrados", "") # Opción por defecto
                    print("No se cargaron empleados desde la hoja 'Empleados'.")
                else:
                    print(f"Se cargaron {self.ui.EmpleadosBox.count()} empleados.")

            else:
                # Error crítico si el ComboBox no está en la UI
                print("ERROR: El QComboBox 'EmpleadosBox' no se encontró en la UI.")
                # Podrías deshabilitar el botón si el combo no existe
                if hasattr(self.ui, 'btnRegisterVale'):
                    self.ui.btnRegisterVale.setEnabled(False)
                return # Salir de la configuración si falta el combo

            # --- Conexión del Botón (Asumiendo que existe en self.ui) ---
# --- Conexión del Botón (Asumiendo que existe en self.ui) ---
            if hasattr(self.ui, 'btnRegisterVale'):
                # Verificar que el método al que conectar existe en esta clase
                if hasattr(self, 'register_vale'):
                    # Desconectar primero (de forma segura) por si esta función se llama múltiples veces
                    try:
                        # Intenta desconectar cualquier conexión previa de 'clicked'
                        self.ui.btnRegisterVale.clicked.disconnect()
                        # Opcional: puedes añadir un print aquí si quieres saber cuándo la desconexión SÍ funciona
                        # print("Desconexión previa de btnRegisterVale.clicked exitosa.")
                    except RuntimeError:
                        # Ignora el error si no había nada conectado previamente.
                        # print("Advertencia: No se pudo desconectar btnRegisterVale.clicked (probablemente no estaba conectado).")
                        pass
                    # Podrías añadir 'except TypeError: pass' también por si acaso, aunque el error fue RuntimeError

                    # Conectar la señal 'clicked' al método 'register_vale'
                    # ESTA ES LA LÍNEA IMPORTANTE AHORA
                    self.ui.btnRegisterVale.clicked.connect(self.register_vale)
                    print("Botón 'btnRegisterVale' conectado correctamente.") # <-- ¡Busca este mensaje!
                    self.ui.btnRegisterVale.setEnabled(True) # Asegurarse de que esté habilitado
                else:
                     # Error si el método no existe
                    print("ERROR: El método 'register_vale' no está definido en esta clase. El botón no funcionará.")
                    self.ui.btnRegisterVale.setEnabled(False) # Deshabilitar el botón
            else:
                # Error crítico si el botón no está en la UI
                print("ERROR: El QPushButton 'btnRegisterVale' no se encontró en la UI.")

        except FileNotFoundError:
             print(f"ERROR CRÍTICO: No se pudo encontrar ni crear el archivo Excel en {self.excel_path}")
        except Exception as e:
            print(f"Error inesperado durante setup_employee_management: {e}")
            traceback.print_exc() # Imprime el rastreo completo del error para depuración

    def update_employee_reports(self):
        """
        Crea o actualiza hojas individuales por cada empleado en el archivo Excel.
        Estas hojas contienen:
        1. Una tabla con todos los vales del empleado
        2. La suma total del campo Valor
        3. Consolidados semanales, mensuales y anuales
        """
        try:
            print("Actualizando reportes de empleados...")
            
            # Cargar el archivo Excel
            wb = load_workbook(self.excel_path)
            
            # Verificar si existen las hojas necesarias
            if "Vales" not in wb.sheetnames or "Empleados" not in wb.sheetnames:
                QMessageBox.warning(self, "Error", "No se encontraron las hojas 'Vales' o 'Empleados' en el archivo Excel.")
                return
            
            # Obtener datos de empleados
            empleados_ws = wb["Empleados"]
            empleados = {}
            for row in list(empleados_ws.rows)[1:]:  # Saltar el encabezado
                if row[0].value and row[4].value:  # Nombre y EmpleadoId
                    empleados[row[4].value] = {
                        "nombre": row[0].value,
                        "cedula": row[1].value,
                        "celular": row[2].value,
                        "correo": row[3].value
                    }
            
            if not empleados:
                QMessageBox.information(self, "Info", "No hay empleados registrados en la base de datos.")
                return
            
            # Obtener datos de vales
            vales_ws = wb["Vales"]
            vales_data = []
            for row in list(vales_ws.rows)[1:]:  # Saltar el encabezado
                # Verificar si la fila tiene suficientes columnas y datos necesarios
                if len(row) > 8 and row[1].value is not None:  # Al menos necesitamos EmpleadoID
                    try:
                        # Convertir valor a número si es posible
                        valor = None
                        if len(row) > 8 and row[8].value is not None:
                            if isinstance(row[8].value, (int, float)):
                                valor = float(row[8].value)
                            elif isinstance(row[8].value, str):
                                # Limpiar string de caracteres no numéricos
                                valor_str = ''.join(c for c in row[8].value if c.isdigit() or c == '.')
                                if valor_str:
                                    valor = float(valor_str) if '.' in valor_str else int(valor_str)
                        
                        vales_data.append({
                            "id": row[0].value if len(row) > 0 else None,
                            "empleado_id": row[1].value,  # Campo obligatorio
                            "fecha": row[2].value if len(row) > 2 else None,
                            "tipo_trabajo": row[3].value if len(row) > 3 else None,
                            "num_ticket": row[4].value if len(row) > 4 else None,
                            "referencia": row[5].value if len(row) > 5 else None,
                            "talla": row[6].value if len(row) > 6 else None,
                            "color": row[7].value if len(row) > 7 else None,
                            "valor": valor,  # Ya convertido a número
                            "total_producido": row[9].value if len(row) > 9 else None,
                            "codigo_serial": row[10].value if len(row) > 10 else None
                        })
                    except Exception as e:
                        print(f"Error al procesar fila de vale: {str(e)}")
            
            # Para cada empleado, crear o actualizar su hoja
            for empleado_id, empleado_info in empleados.items():
                # Obtener vales del empleado
                vales_empleado = [vale for vale in vales_data if vale["empleado_id"] == empleado_id]
                
                # Si el empleado ya tiene una hoja, eliminarla para recrearla
                sheet_name = f"Empleado_{empleado_id}"
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
                
                # Crear nueva hoja para el empleado
                emp_ws = wb.create_sheet(title=sheet_name)
                
                # Información del empleado en la parte superior
                emp_ws['A1'] = "INFORMACIÓN DEL EMPLEADO"
                emp_ws['A1'].font = Font(bold=True, size=14)
                emp_ws.merge_cells('A1:F1')
                
                emp_ws['A2'] = "Nombre:"
                emp_ws['B2'] = empleado_info["nombre"]
                emp_ws['A3'] = "Cédula:"
                emp_ws['B3'] = empleado_info["cedula"]
                emp_ws['A4'] = "Celular:"
                emp_ws['B4'] = empleado_info["celular"]
                emp_ws['A5'] = "Correo:"
                emp_ws['B5'] = empleado_info["correo"]
                emp_ws['A6'] = "ID Empleado:"
                emp_ws['B6'] = empleado_id
                
                # Dar formato a la información del empleado
                for row in range(2, 7):
                    emp_ws[f'A{row}'].font = Font(bold=True)
                
                # Espacio antes de la tabla de vales
                emp_ws['A8'] = "DETALLE DE VALES"
                emp_ws['A8'].font = Font(bold=True, size=12)
                emp_ws.merge_cells('A8:F8')
                
                # Encabezados tabla de vales
                headers = ["ID", "Fecha", "Tipo Trabajo", "# Ticket", "Referencia", 
                        "Talla", "Color", "Valor", "Total Producido", "Código Serial"]
                for col, header in enumerate(headers, start=1):
                    cell = emp_ws.cell(row=9, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                # Datos de vales
                valor_total = 0
                for row_idx, vale in enumerate(vales_empleado, start=10):
                    emp_ws.cell(row=row_idx, column=1).value = vale["id"]
                    
                    # Formato de fecha
                    if vale["fecha"] is not None:
                        if isinstance(vale["fecha"], datetime.datetime):
                            emp_ws.cell(row=row_idx, column=2).value = vale["fecha"].strftime("%Y-%m-%d")
                        else:
                            emp_ws.cell(row=row_idx, column=2).value = str(vale["fecha"])
                    else:
                        emp_ws.cell(row=row_idx, column=2).value = "Sin fecha"
                    
                    emp_ws.cell(row=row_idx, column=3).value = vale["tipo_trabajo"]
                    emp_ws.cell(row=row_idx, column=4).value = vale["num_ticket"]
                    emp_ws.cell(row=row_idx, column=5).value = vale["referencia"]
                    emp_ws.cell(row=row_idx, column=6).value = vale["talla"]
                    emp_ws.cell(row=row_idx, column=7).value = vale["color"]
                    
                    # Valor con formato
                    valor_cell = emp_ws.cell(row=row_idx, column=8)
                    valor_cell.value = vale["valor"]
                    valor_cell.number_format = '#,##0'
                    
                    emp_ws.cell(row=row_idx, column=9).value = vale["total_producido"]
                    emp_ws.cell(row=row_idx, column=10).value = vale["codigo_serial"]
                    
                    # Convertir valor a número antes de sumar (maneja tanto strings como números)
                    if vale["valor"] is not None:
                        try:
                            # Si es string, convertir a entero o float
                            if isinstance(vale["valor"], str):
                                # Eliminar cualquier caracter no numérico (como símbolos de moneda o comas)
                                valor_numerico = ''.join(c for c in vale["valor"] if c.isdigit() or c == '.')
                                if valor_numerico:
                                    if '.' in valor_numerico:
                                        valor_total += float(valor_numerico)
                                    else:
                                        valor_total += int(valor_numerico)
                            else:
                                # Si ya es número, sumar directamente
                                valor_total += float(vale["valor"])
                        except (ValueError, TypeError):
                            print(f"Error al convertir valor '{vale['valor']}' a número, se ignora para el total")
                
                last_row = len(vales_empleado) + 10
                
                # Total de vales
                emp_ws.cell(row=last_row, column=7).value = "TOTAL:"
                emp_ws.cell(row=last_row, column=7).font = Font(bold=True)
                emp_ws.cell(row=last_row, column=7).alignment = Alignment(horizontal='right')
                
                total_cell = emp_ws.cell(row=last_row, column=8)
                total_cell.value = valor_total
                total_cell.font = Font(bold=True)
                total_cell.number_format = '#,##0'
                
                # Ajustar anchos de columna
                for col in range(1, 11):
                    emp_ws.column_dimensions[chr(64 + col)].width = 15
                
                # Espacio antes de los consolidados
                consolidado_start_row = last_row + 3
                emp_ws.cell(row=consolidado_start_row, column=1).value = "CONSOLIDADO DE PAGOS"
                emp_ws.cell(row=consolidado_start_row, column=1).font = Font(bold=True, size=12)
                emp_ws.merge_cells(f'A{consolidado_start_row}:F{consolidado_start_row}')
                
                # Encabezados tablas de consolidado
                consolidado_start_row += 1
                
                # Tabla Semanal
                emp_ws.cell(row=consolidado_start_row, column=1).value = "CONSOLIDADO SEMANAL"
                emp_ws.cell(row=consolidado_start_row, column=1).font = Font(bold=True)
                emp_ws.merge_cells(f'A{consolidado_start_row}:C{consolidado_start_row}')
                
                consolidado_start_row += 1
                emp_ws.cell(row=consolidado_start_row, column=1).value = "Semana"
                emp_ws.cell(row=consolidado_start_row, column=2).value = "Fecha Inicio"
                emp_ws.cell(row=consolidado_start_row, column=3).value = "Fecha Fin"
                emp_ws.cell(row=consolidado_start_row, column=4).value = "Valor Total"
                
                # Estilo encabezados
                for col in range(1, 5):
                    cell = emp_ws.cell(row=consolidado_start_row, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                # Calcular consolidado semanal
                weekly_data = {}
                for vale in vales_empleado:
                    if vale["fecha"] is not None and isinstance(vale["fecha"], datetime.datetime):
                        try:
                            # Obtener el número de semana del año
                            year = vale["fecha"].year
                            week_num = vale["fecha"].isocalendar()[1]
                            week_key = f"{year}-W{week_num:02d}"
                            
                            # Calcular fecha inicio y fin de semana
                            # El lunes de la semana
                            start_of_week = vale["fecha"] - datetime.timedelta(days=vale["fecha"].weekday())
                            # El domingo de la semana
                            end_of_week = start_of_week + datetime.timedelta(days=6)
                        except Exception as e:
                            print(f"Error al procesar fecha para consolidado semanal: {str(e)}")
                            continue  # Salta este vale
                        
                        if week_key not in weekly_data:
                            weekly_data[week_key] = {
                                "start_date": start_of_week,
                                "end_date": end_of_week,
                                "total": 0
                            }
                        
                        # Convertir valor a número antes de sumar
                        if vale["valor"] is not None:
                            try:
                                # Si es string, convertir a número
                                if isinstance(vale["valor"], str):
                                    valor_numerico = ''.join(c for c in vale["valor"] if c.isdigit() or c == '.')
                                    if valor_numerico:
                                        if '.' in valor_numerico:
                                            weekly_data[week_key]["total"] += float(valor_numerico)
                                        else:
                                            weekly_data[week_key]["total"] += int(valor_numerico)
                                else:
                                    # Si ya es número, sumar directamente
                                    weekly_data[week_key]["total"] += float(vale["valor"])
                            except (ValueError, TypeError):
                                print(f"Error al convertir valor '{vale['valor']}' para suma semanal")
                
                # Llenar datos semanales
                row_idx = consolidado_start_row + 1
                for week_key, data in sorted(weekly_data.items()):
                    emp_ws.cell(row=row_idx, column=1).value = week_key
                    emp_ws.cell(row=row_idx, column=2).value = data["start_date"].strftime("%Y-%m-%d")
                    emp_ws.cell(row=row_idx, column=3).value = data["end_date"].strftime("%Y-%m-%d")
                    
                    valor_cell = emp_ws.cell(row=row_idx, column=4)
                    valor_cell.value = data["total"]
                    valor_cell.number_format = '#,##0'
                    
                    row_idx += 1
                
                # Espacio entre tablas
                row_idx += 2
                
                # Tabla Mensual
                emp_ws.cell(row=row_idx, column=1).value = "CONSOLIDADO MENSUAL"
                emp_ws.cell(row=row_idx, column=1).font = Font(bold=True)
                emp_ws.merge_cells(f'A{row_idx}:C{row_idx}')
                
                row_idx += 1
                emp_ws.cell(row=row_idx, column=1).value = "Año"
                emp_ws.cell(row=row_idx, column=2).value = "Mes"
                emp_ws.cell(row=row_idx, column=3).value = "Valor Total"
                
                # Estilo encabezados
                for col in range(1, 4):
                    cell = emp_ws.cell(row=row_idx, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                # Calcular consolidado mensual
                monthly_data = {}
                for vale in vales_empleado:
                    if vale["fecha"] is not None and isinstance(vale["fecha"], datetime.datetime):
                        try:
                            year = vale["fecha"].year
                            month = vale["fecha"].month
                            month_key = f"{year}-{month:02d}"
                        except Exception as e:
                            print(f"Error al procesar fecha para consolidado mensual: {str(e)}")
                            continue  # Salta este vale
                        
                        if month_key not in monthly_data:
                            monthly_data[month_key] = {
                                "year": year,
                                "month": month,
                                "total": 0
                            }
                        
                        # Convertir valor a número antes de sumar
                        if vale["valor"] is not None:
                            try:
                                # Si es string, convertir a número
                                if isinstance(vale["valor"], str):
                                    valor_numerico = ''.join(c for c in vale["valor"] if c.isdigit() or c == '.')
                                    if valor_numerico:
                                        if '.' in valor_numerico:
                                            monthly_data[month_key]["total"] += float(valor_numerico)
                                        else:
                                            monthly_data[month_key]["total"] += int(valor_numerico)
                                else:
                                    # Si ya es número, sumar directamente
                                    monthly_data[month_key]["total"] += float(vale["valor"])
                            except (ValueError, TypeError):
                                print(f"Error al convertir valor '{vale['valor']}' para suma mensual")
                
                # Nombres de meses
                month_names = [
                    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
                ]
                
                # Llenar datos mensuales
                row_idx += 1
                for month_key, data in sorted(monthly_data.items()):
                    emp_ws.cell(row=row_idx, column=1).value = data["year"]
                    emp_ws.cell(row=row_idx, column=2).value = month_names[data["month"] - 1]
                    
                    valor_cell = emp_ws.cell(row=row_idx, column=3)
                    valor_cell.value = data["total"]
                    valor_cell.number_format = '#,##0'
                    
                    row_idx += 1
                
                # Espacio entre tablas
                row_idx += 2
                
                # Tabla Anual
                emp_ws.cell(row=row_idx, column=1).value = "CONSOLIDADO ANUAL"
                emp_ws.cell(row=row_idx, column=1).font = Font(bold=True)
                emp_ws.merge_cells(f'A{row_idx}:B{row_idx}')
                
                row_idx += 1
                emp_ws.cell(row=row_idx, column=1).value = "Año"
                emp_ws.cell(row=row_idx, column=2).value = "Valor Total"
                
                # Estilo encabezados
                for col in range(1, 3):
                    cell = emp_ws.cell(row=row_idx, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                # Calcular consolidado anual
                yearly_data = {}
                for vale in vales_empleado:
                    if vale["fecha"] is not None and isinstance(vale["fecha"], datetime.datetime):
                        try:
                            year = vale["fecha"].year
                        except Exception as e:
                            print(f"Error al procesar fecha para consolidado anual: {str(e)}")
                            continue  # Salta este vale
                        
                        if year not in yearly_data:
                            yearly_data[year] = 0
                        
                        # Convertir valor a número antes de sumar
                        if vale["valor"] is not None:
                            try:
                                # Si es string, convertir a número
                                if isinstance(vale["valor"], str):
                                    valor_numerico = ''.join(c for c in vale["valor"] if c.isdigit() or c == '.')
                                    if valor_numerico:
                                        if '.' in valor_numerico:
                                            yearly_data[year] += float(valor_numerico)
                                        else:
                                            yearly_data[year] += int(valor_numerico)
                                else:
                                    # Si ya es número, sumar directamente
                                    yearly_data[year] += float(vale["valor"])
                            except (ValueError, TypeError):
                                print(f"Error al convertir valor '{vale['valor']}' para suma anual")
                
                # Llenar datos anuales
                row_idx += 1
                for year, total in sorted(yearly_data.items()):
                    emp_ws.cell(row=row_idx, column=1).value = year
                    
                    valor_cell = emp_ws.cell(row=row_idx, column=2)
                    valor_cell.value = total
                    valor_cell.number_format = '#,##0'
                    
                    row_idx += 1
            
            # Guardar el archivo Excel
            wb.save(self.excel_path)
            
            QMessageBox.information(self, "Éxito", "Se han actualizado los reportes de todos los empleados correctamente.")
            print("Reportes de empleados actualizados con éxito.")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al actualizar reportes: {str(e)}")
            print(f"Error al actualizar reportes: {str(e)}")


    def setup_update_button(self):
        """
        Agrega un botón ActualizarDB a la interfaz para actualizar los reportes de empleados
        """
        self.ui.btnActualizarDB.clicked.connect(self.update_employee_reports)

    def on_code_type_changed(self, checked):
        """Handle change in code type selection"""
        if checked:
            self.current_code_type = "barcode"
        else:
            self.current_code_type = "qr"
            
    def shorten_serial_code(self, code: str, length: int = 8) -> str:
        """Return a short hash from a given code"""
        hash_object = hashlib.sha1(code.encode())
        short_hash = hash_object.hexdigest()[:length]
        return short_hash.upper()
        
    def generate_serial_code(self):
        """Generate a unique serial code based on form fields"""
        # Get field values
        tipo_trabajo = self.ui.CampoTipoTrabajo.text().strip() if hasattr(self.ui, 'CampoTipoTrabajo') else ""
        referencia = self.ui.CampoReferenciaTrabajo.text().strip() if hasattr(self.ui, 'CampoReferenciaTrabajo') else ""
        num_ticket = self.ui.CampoNumeroTicket.text().strip() if hasattr(self.ui, 'CampoNumeroTicket') else ""
        talla = self.ui.CampoTalla.text().strip() if hasattr(self.ui, 'CampoTalla') else ""
        color = self.ui.CampoColor.text().strip() if hasattr(self.ui, 'CampoColor') else ""
        
        # Generate a UUID and take first 6 characters
        unique_id = str(uuid.uuid4())[:6]
        
        # Create the compact serial code
        serial_code = f"{tipo_trabajo[:1]}{referencia[:2]}{num_ticket[:3]}{talla[:1]}{color[:1]}{unique_id}"
        return serial_code.upper()
    


    def generate_barcode(self, serial_code):
        try:
            barcode_class = barcode.get_barcode_class('code128')

            class CustomImageWriter(ImageWriter):
                def __init__(self):
                    super().__init__()
                    # Ruta base para recursos (fuente)
                    if getattr(sys, 'frozen', False):
                        base_path = sys._MEIPASS
                    else:
                        base_path = os.path.abspath(".")

                    # Lista de posibles ubicaciones de la fuente
                    font_paths = [
                        os.path.join(base_path, "fonts", "segoeui.ttf"),
                        os.path.join(base_path, "segoeui.ttf")
                    ]
                    self.font_path = next((p for p in font_paths if os.path.exists(p)), None)
                    if not self.font_path:
                        print("ADVERTENCIA: No se encontró la fuente, usando valores por defecto")

                def _paint_text(self, xpos, ypos):
                    """
                    Invocado por python-barcode como paint_text(xpos, ypos).
                    El texto a pintar está en self.text.
                    """
                    text = self.text  # el código (string) a pintar
                    # Intentar usar la fuente personalizada
                    if self.font_path:
                        try:
                            size = int(getattr(self, "font_size", 10))
                            font = ImageFont.truetype(self.font_path, size)
                            self._draw.text((xpos, ypos), text, fill=self.foreground, font=font)
                            return
                        except Exception as e:
                            print(f"Error al usar fuente personalizada: {e}")
                    # Fallback al método base (sin truetype)
                    self._draw.text((xpos, ypos), text, fill=self.foreground)

            # Instanciar writer y aplicar opciones
            writer = CustomImageWriter()
            options = {
                'module_height': 8.0,
                'module_width': 0.2,
                'quiet_zone': 1.0,
                'font_size': 7,
                'text_distance': 1.0,
            }
            for k, v in options.items():
                if hasattr(writer, k):
                    setattr(writer, k, v)

            # Determinar directorio de salida
            if getattr(sys, 'frozen', False):
                application_path = os.path.dirname(sys.executable)
            else:
                application_path = os.path.dirname(os.path.abspath(__file__))
            codes_dir = os.path.join(application_path, "codes")
            os.makedirs(codes_dir, exist_ok=True)

            # Generar y guardar
            barcode_path = os.path.join(codes_dir, f"barcode_{serial_code}")
            instance = barcode_class(serial_code, writer=writer)
            filename = instance.save(barcode_path)
            return filename

        except Exception as e:
            print(f"Error al generar código de barras: {e}")
            traceback.print_exc()
            QMessageBox.critical(self, "Error", f"Error al generar el código de barras:\n{e}")
            return None

    def setup_code_generator(self):
        """
        Setup the barcode and QR code generator functionality
        """
        # Define excel path and create barcode directory if needed
        self.excel_path = "trabajos_database.xlsx"
        self.vales_sheet_name = "Vales"  # Sheet name for vales data
        
        if not os.path.exists("codes"):
            os.makedirs("codes")
        
        # Create Excel file if it doesn't exist
        if not os.path.exists(self.excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Trabajos"
            ws.append(["Código Serial", "Tipo Trabajo", "Referencia", "Número Ticket",
                     "Talla", "Color", "Valor", "Total Producido", "Tipo Código", "Ruta Imagen"])
            
            # Create Vales sheet
            vales_ws = wb.create_sheet(title=self.vales_sheet_name)
            vales_ws.append(["ID", "EmpleadoID", "Fecha", "Tipo Trabajo", "Número Ticket",
                          "Referencia", "Talla", "Color", "Valor", "Total Producido", "Código Serial"])
            
            wb.save(self.excel_path)
        else:
            # Check if Vales sheet exists, if not create it
            wb = load_workbook(self.excel_path)
            if self.vales_sheet_name not in wb.sheetnames:
                vales_ws = wb.create_sheet(title=self.vales_sheet_name)
                vales_ws.append(["ID", "EmpleadoID", "Fecha", "Tipo Trabajo", "Número Ticket",
                              "Referencia", "Talla", "Color", "Valor", "Total Producido", "Código Serial"])
                wb.save(self.excel_path)
        
        # Connect the save button
        self.ui.pushButtonGuardar.clicked.connect(self.on_save_button_clicked)
        
        # Add code type selector (radio buttons)
        self.setup_code_type_selector()
        
        # Ensure the QGraphicsView has a scene
        if not hasattr(self.ui, 'PreviwImage'):
            print("Warning: PreviwImage not found in UI")
        elif self.ui.PreviwImage.scene() is None:
            self.ui.PreviwImage.setScene(QtWidgets.QGraphicsScene())
    
    def setup_autocompletado_fields(self):
        """
        Configura el autocompletado para los campos QLineEdit relevantes.
        """
        # Define los campos QLineEdit de tu UI y las columnas de Excel correspondientes
        # El formato es: 'identificador_unico': {'line_edit': self.ui.TuLineEdit, 'columna': 'NombreDeColumnaEnExcel'}
        
        campos_a_configurar = {}

        # Asegúrate de que estos QLineEdit existen en tu self.ui y las columnas en tu Excel
        if hasattr(self.ui, 'CampoTipoTrabajo'):
            campos_a_configurar['tipo_trabajo'] = {
                'line_edit': self.ui.CampoTipoTrabajo, 
                'columna': 'Tipo Trabajo'
            }
        if hasattr(self.ui, 'CampoReferenciaTrabajo'):
            campos_a_configurar['referencia_trabajo'] = {
                'line_edit': self.ui.CampoReferenciaTrabajo, 
                'columna': 'Referencia'
            }
        if hasattr(self.ui, 'CampoNumeroTicket'):
            campos_a_configurar['numero_ticket'] = {
                'line_edit': self.ui.CampoNumeroTicket, 
                'columna': 'Número Ticket'
            }
        if hasattr(self.ui, 'CampoTalla'):
            campos_a_configurar['talla'] = {
                'line_edit': self.ui.CampoTalla, 
                'columna': 'Talla'
            }
        if hasattr(self.ui, 'CampoColor'):
            campos_a_configurar['color'] = {
                'line_edit': self.ui.CampoColor, 
                'columna': 'Color'
            }
        # Añade más campos si es necesario, por ejemplo:
        # if hasattr(self.ui, 'CampoValor'):
        #     campos_a_configurar['valor'] = {
        #         'line_edit': self.ui.CampoValor,
        #         'columna': 'Valor' # Asegúrate que esta columna exista y tenga sentido para autocompletar
        #     }

        if campos_a_configurar:
            self.autocompletado_manager.configurar_multiples_campos(campos_a_configurar)
            print("Autocompletado configurado para los campos.")
        else:
            print("No se encontraron campos para configurar el autocompletado.")

    def on_save_button_clicked(self):
        """Handler for save button click"""
        # Define required fields and their UI elements
        required_fields = []
        
        # Only add fields that exist in the UI
        if hasattr(self.ui, 'CampoTipoTrabajo'):
            required_fields.append((self.ui.CampoTipoTrabajo, "Tipo de Trabajo"))
        if hasattr(self.ui, 'CampoReferenciaTrabajo'):
            required_fields.append((self.ui.CampoReferenciaTrabajo, "Referencia"))
        if hasattr(self.ui, 'CampoNumeroTicket'):
            required_fields.append((self.ui.CampoNumeroTicket, "Número de Ticket"))
        if hasattr(self.ui, 'CampoTalla'):
            required_fields.append((self.ui.CampoTalla, "Talla"))
        if hasattr(self.ui, 'CampoColor'):
            required_fields.append((self.ui.CampoColor, "Color"))
        if hasattr(self.ui, 'CampoValor'):
            required_fields.append((self.ui.CampoValor, "Valor"))
        if hasattr(self.ui, 'CampoTotalProducido'):
            required_fields.append((self.ui.CampoTotalProducido, "Total Producido"))
        
        # Validate required fields
        for field, name in required_fields:
            if not field.text().strip():
                QMessageBox.warning(self, "Campos Incompletos", f"El campo {name} es obligatorio.")
                field.setFocus()
                return
        
        # Get values from UI
        tipo_trabajo = getattr(self.ui, 'CampoTipoTrabajo').text() if hasattr(self.ui, 'CampoTipoTrabajo') else ""
        referencia = getattr(self.ui, 'CampoReferenciaTrabajo').text() if hasattr(self.ui, 'CampoReferenciaTrabajo') else ""
        ticket_number = getattr(self.ui, 'CampoNumeroTicket').text() if hasattr(self.ui, 'CampoNumeroTicket') else ""
        talla = getattr(self.ui, 'CampoTalla').text() if hasattr(self.ui, 'CampoTalla') else ""
        color = getattr(self.ui, 'CampoColor').text() if hasattr(self.ui, 'CampoColor') else ""
        valor = getattr(self.ui, 'CampoValor').text() if hasattr(self.ui, 'CampoValor') else ""
        total_producido = getattr(self.ui, 'CampoTotalProducido').text() if hasattr(self.ui, 'CampoTotalProducido') else "0"
        
        # Generate unique serial code
        serial_code = f"{ticket_number}-{referencia}-{talla}"
        
        # Generate the selected code type
        if self.current_code_type == "barcode":
            code_path = self.generate_barcode(serial_code)
        else:
            code_path = self.generate_qr_code(serial_code)
        
        if not code_path:
            QMessageBox.critical(self, "Error", "Error al generar el código.")
            return
        
        # Display the code
        if not self.display_code_image(code_path):
            return
        
        # Save data to Excel
        if self.save_to_excel(serial_code, code_path):
            # Create a dictionary of tallas and cantidades (for the PDF)
            tallas_cantidades = {talla: int(total_producido) if total_producido.isdigit() else 0}
            
            # Generate PDF with QR code
            pdf_path = self.generate_vale_pdf(
                ticket_number, tipo_trabajo, referencia,
                tallas_cantidades, color, total_producido, serial_code
            )
            
            # Show success message with both paths
            QMessageBox.information(
                self,
                "Operación Exitosa",
                f"Código generado: {serial_code}\nTipo: {self.current_code_type.upper()}\n"
                f"Los datos se han guardado correctamente.\n"
                f"PDF generado en: {pdf_path}"
            )
            
            # Clear fields
            for field, _ in required_fields:
                field.clear()
            
            # Focus first field
            if required_fields:
                required_fields[0][0].setFocus()
    
    def generate_qr_code(self, data):
        """
        Generate QR code image
        
        Args:
            data: Data to encode in the QR code
            
        Returns:
            str: Path to the generated QR code image or None if generation fails
        """
        try:
            # Create filename
            filename = f"qr_{data.replace('/', '_').replace(' ', '_')}.png"
            filepath = os.path.join("codes", filename)
            
            # Generate QR code
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(data)
            qr.make(fit=True)
            
            qr_img = qr.make_image(fill_color="black", back_color="white")
            qr_img.save(filepath)
            
            return filepath
        except Exception as e:
            print(f"Error generating QR code: {e}")
            return None
    
    def display_code_image(self, image_path):
        """Display the code image in the QGraphicsView"""
        try:
            if not hasattr(self.ui, 'PreviwImage'):
                print("Error: PreviwImage not found in UI")
                return False
                
            # Load image
            pixmap = QPixmap(image_path)
            
            # Get scene from QGraphicsView
            scene = self.ui.PreviwImage.scene()
            if scene is None:
                scene = QtWidgets.QGraphicsScene()
                self.ui.PreviwImage.setScene(scene)
            else:
                scene.clear()
            
            # Add image to scene
            scene.addPixmap(pixmap)
            
            # Fit view
            self.ui.PreviwImage.fitInView(scene.sceneRect(), Qt.KeepAspectRatio)
            
            return True
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al mostrar la imagen: {str(e)}")
            return False
    
    def save_to_excel(self, serial_code, code_path):
        """
        Save data to Excel file
        
        Args:
            serial_code: Serial code for the generated barcode/QR
            code_path: Path to the generated code image
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Get values from UI
            tipo_trabajo = getattr(self.ui, 'CampoTipoTrabajo').text() if hasattr(self.ui, 'CampoTipoTrabajo') else ""
            referencia = getattr(self.ui, 'CampoReferenciaTrabajo').text() if hasattr(self.ui, 'CampoReferenciaTrabajo') else ""
            ticket_number = getattr(self.ui, 'CampoNumeroTicket').text() if hasattr(self.ui, 'CampoNumeroTicket') else ""
            talla = getattr(self.ui, 'CampoTalla').text() if hasattr(self.ui, 'CampoTalla') else ""
            color = getattr(self.ui, 'CampoColor').text() if hasattr(self.ui, 'CampoColor') else ""
            valor = getattr(self.ui, 'CampoValor').text() if hasattr(self.ui, 'CampoValor') else ""
            total_producido = getattr(self.ui, 'CampoTotalProducido').text() if hasattr(self.ui, 'CampoTotalProducido') else "0"
            
            # Load workbook
            wb = load_workbook(self.excel_path)
            ws = wb["Trabajos"]
            
            # Add new row
            ws.append([
                serial_code, tipo_trabajo, referencia, ticket_number,
                talla, color, valor, total_producido, 
                self.current_code_type.upper(), code_path
            ])
            
            # Save workbook
            wb.save(self.excel_path)
            return True
        except Exception as e:
            print(f"Error saving to Excel: {e}")
            QMessageBox.critical(self, "Error", f"Error al guardar los datos: {e}")
            return False
    
    def find_code_data(self, serial_code):
        """Find data related to a specific serial code in the Excel file"""
        try:
            # Load Excel file
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            # Search for the serial code in the first column
            for row in ws.iter_rows(min_row=2):  # Start from row 2 to skip header
                if row[0].value == serial_code:
                    # Return all data for this row
                    return {
                        "serial_code": row[0].value,
                        "tipo_trabajo": row[1].value,
                        "referencia": row[2].value,
                        "num_ticket": row[3].value,
                        "talla": row[4].value,
                        "color": row[5].value,
                        "valor": row[6].value,
                        "total_producido": row[7].value,
                        "tipo_codigo": row[8].value,
                        "ruta_imagen": row[9].value
                    }
            
            return None
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al buscar datos: {str(e)}")
            return None
    
    def on_code_scanned(self):
        """Handle barcode/QR code scanning"""
        # Get scanned code
        scanned_code = self.ui.codeReaderInput.text().strip()
        if not scanned_code:
            return
        
        # Find data for this code
        data = self.find_code_data(scanned_code)
        if not data:
            QMessageBox.warning(self, "Código No Encontrado", 
                               f"No se encontraron datos para el código: {scanned_code}")
            self.ui.codeReaderInput.clear()
            return
        
        # Generate a unique ID for this vale
        vale_id = f"V-{datetime.datetime.now().strftime('%Y%m%d')}-{str(uuid.uuid4())[:6]}"
        
        # Add data to table view
        self.add_to_table_view(vale_id, data)
        
        # Clear input field
        self.ui.codeReaderInput.clear()
        
        # Display code image if available
        if data["ruta_imagen"] and os.path.exists(data["ruta_imagen"]):
            self.display_code_image(data["ruta_imagen"])
            
        # Show success message
        QMessageBox.information(self, "Código Encontrado", 
                               f"Código escaneado: {scanned_code}\nTipo de trabajo: {data['tipo_trabajo']}\nVale ID: {vale_id}")
    
    def add_to_table_view(self, vale_id, data):
        """Add scanned code data to the TableView"""
        # Get employee info
        employee_text = self.ui.EmpleadosBox.currentText()
        employee_id = self.ui.EmpleadosBox.currentData()
        
        # Create row items
        items = [
            QtGui.QStandardItem(vale_id),
            QtGui.QStandardItem(data["serial_code"]),
            QtGui.QStandardItem(str(data["tipo_trabajo"])),
            QtGui.QStandardItem(str(data["referencia"])),
            QtGui.QStandardItem(str(data["num_ticket"])),
            QtGui.QStandardItem(str(data["talla"])),
            QtGui.QStandardItem(str(data["color"])),
            QtGui.QStandardItem(str(data["valor"])),
            QtGui.QStandardItem(str(data["total_producido"])),
            QtGui.QStandardItem(employee_text)
        ]
        
        # Add row to model
        self.table_model.appendRow(items)
        
        # Store data for later registration
        data["vale_id"] = vale_id
        data["employee_id"] = employee_id
        data["employee_text"] = employee_text
        
        # Add to pending vales list if not already exists
        if not hasattr(self, 'pending_vales'):
            self.pending_vales = []
        
        self.pending_vales.append(data)
    
    def register_vale(self):
        """Register the scanned vales to Excel"""
        if not hasattr(self, 'pending_vales') or not self.pending_vales:
            QMessageBox.warning(self, "Sin Vales", "No hay vales pendientes para registrar.")
            return
        
        try:
            # Get current date
            current_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Load Excel file
            wb = load_workbook(self.excel_path)
            
            # Get vales sheet or create if not exists
            if self.vales_sheet_name in wb.sheetnames:
                ws = wb[self.vales_sheet_name]
            else:
                ws = wb.create_sheet(title=self.vales_sheet_name)
                ws.append(["ID", "EmpleadoID", "Fecha", "Tipo Trabajo", "Número Ticket",
                          "Referencia", "Talla", "Color", "Valor", "Total Producido", "Código Serial"])
            
            # Add all pending vales
            for vale in self.pending_vales:
                ws.append([
                    vale["vale_id"],
                    vale["employee_id"],
                    current_date,
                    vale["tipo_trabajo"],
                    vale["num_ticket"],
                    vale["referencia"],
                    vale["talla"],
                    vale["color"],
                    vale["valor"],
                    vale["total_producido"],
                    vale["serial_code"]
                ])
            
            # Save Excel file
            wb.save(self.excel_path)
            
            # Clear pending vales and table view
            self.pending_vales = []
            self.table_model.removeRows(0, self.table_model.rowCount())
            
            # Show success message
            QMessageBox.information(self, "Vales Registrados", 
                                   f"Se han registrado {len(self.pending_vales)} vales correctamente.")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al registrar vales: {str(e)}")
    
    def on_save_button_clicked(self):
        """Handler for save button click"""
        # Define required fields and their UI elements
        required_fields = []
        
        # Only add fields that exist in the UI
        if hasattr(self.ui, 'CampoTipoTrabajo'):
            required_fields.append((self.ui.CampoTipoTrabajo, "Tipo de Trabajo"))
        if hasattr(self.ui, 'CampoReferenciaTrabajo'):
            required_fields.append((self.ui.CampoReferenciaTrabajo, "Referencia"))
        if hasattr(self.ui, 'CampoNumeroTicket'):
            required_fields.append((self.ui.CampoNumeroTicket, "Número de Ticket"))
        if hasattr(self.ui, 'CampoTalla'):
            required_fields.append((self.ui.CampoTalla, "Talla"))
        if hasattr(self.ui, 'CampoColor'):
            required_fields.append((self.ui.CampoColor, "Color"))
        if hasattr(self.ui, 'CampoValor'):
            required_fields.append((self.ui.CampoValor, "Valor"))
        if hasattr(self.ui, 'CampoTotalProducido'):
            required_fields.append((self.ui.CampoTotalProducido, "Total Producido"))
        
        # Validate required fields
        for field, name in required_fields:
            if not field.text().strip():
                QMessageBox.warning(self, "Campos Incompletos", f"El campo {name} es obligatorio.")
                field.setFocus()
                return
        
        # Get values from UI
        tipo_trabajo = getattr(self.ui, 'CampoTipoTrabajo').text() if hasattr(self.ui, 'CampoTipoTrabajo') else ""
        referencia = getattr(self.ui, 'CampoReferenciaTrabajo').text() if hasattr(self.ui, 'CampoReferenciaTrabajo') else ""
        ticket_number = getattr(self.ui, 'CampoNumeroTicket').text() if hasattr(self.ui, 'CampoNumeroTicket') else ""
        talla = getattr(self.ui, 'CampoTalla').text() if hasattr(self.ui, 'CampoTalla') else ""
        color = getattr(self.ui, 'CampoColor').text() if hasattr(self.ui, 'CampoColor') else ""
        valor = getattr(self.ui, 'CampoValor').text() if hasattr(self.ui, 'CampoValor') else ""
        total_producido = getattr(self.ui, 'CampoTotalProducido').text() if hasattr(self.ui, 'CampoTotalProducido') else "0"
        
        # Generate unique serial code
        serial_code = f"{ticket_number}-{referencia}-{talla}"
        
        # Generate the selected code type
        if self.current_code_type == "barcode":
            code_path = self.generate_barcode(serial_code)
        else:
            code_path = self.generate_qr_code(serial_code)
        
        if not code_path:
            QMessageBox.critical(self, "Error", "Error al generar el código no hay code_path.")
            return
        
        # Display the code
        if not self.display_code_image(code_path):
            return
        
        # Save data to Excel
        if self.save_to_excel(serial_code, code_path):
            # Create a dictionary of tallas and cantidades (for the PDF)
            tallas_cantidades = {talla: int(total_producido) if total_producido.isdigit() else 0}
            
            # Generate PDF with QR code
            pdf_path = self.generate_vale_pdf(
                ticket_number, tipo_trabajo, referencia,
                tallas_cantidades, color, total_producido, code_path
            )
            
            # Show success message with both paths
            QMessageBox.information(
                self,
                "Operación Exitosa",
                f"Código generado: {serial_code}\nTipo: {self.current_code_type.upper()}\n"
                f"Los datos se han guardado correctamente.\n"
                f"PDF generado en: {pdf_path}"
            )
            
            # Clear fields
            for field, _ in required_fields:
                field.clear()
            
            # Focus first field
            if required_fields:
                required_fields[0][0].setFocus()

    def Button(self):
        """Handler for menu button clicks"""
        # Get clicked button
        btnWidget = self.sender()

        # Handle different pages
        if btnWidget.objectName() == "btn_home":
            self.ui.stackedWidget.setCurrentWidget(self.ui.page_home)
            UIFunctions.resetStyle(self, "btn_home")
            btnWidget.setStyleSheet(UIFunctions.selectMenu(btnWidget.styleSheet()))
        elif btnWidget.objectName() == "btn_new_user":
            self.ui.stackedWidget.setCurrentWidget(self.ui.page_widgets)
            UIFunctions.resetStyle(self, "btn_widgets")
            btnWidget.setStyleSheet(UIFunctions.selectMenu(btnWidget.styleSheet()))
        elif btnWidget.objectName() == "btn_widgets":
            self.ui.stackedWidget.setCurrentWidget(self.ui.create_user)
            UIFunctions.resetStyle(self, "create_user")      
            btnWidget.setStyleSheet(UIFunctions.selectMenu(btnWidget.styleSheet()))

    # Window event handlers
    def moveWindow(self, event):
        """Handle window movement"""
        # If maximized change to normal
        if UIFunctions.returStatus() == 1:
            UIFunctions.maximize_restore(self)

        # Move window
        if event.buttons() == Qt.LeftButton:
            self.move(self.pos() + event.globalPos() - self.dragPos)
            self.dragPos = event.globalPos()
            event.accept()

    def eventFilter(self, watched, event):
        """Event filter for double-click events"""
        if hasattr(self, 'le') and watched == self.le and event.type() == QtCore.QEvent.MouseButtonDblClick:
            print("pos: ", event.pos())
        return super().eventFilter(watched, event)

    def mousePressEvent(self, event):
        """Handle mouse press events"""
        self.dragPos = event.globalPos()
        if event.buttons() == Qt.LeftButton:
            print('Mouse click: LEFT CLICK')
        if event.buttons() == Qt.RightButton:
            print('Mouse click: RIGHT CLICK')
        if event.buttons() == Qt.MidButton:
            print('Mouse click: MIDDLE BUTTON')

    def keyPressEvent(self, event):
        """Handle key press events"""
        print('Key: ' + str(event.key()) + ' | Text Press: ' + str(event.text()))

    def resizeEvent(self, event):
        """Handle resize events"""
        self.resizeFunction()
        return super(MainWindow, self).resizeEvent(event)

    def resizeFunction(self):
        """Log window size on resize"""
        print('Height: ' + str(self.height()) + ' | Width: ' + str(self.width()))

    def generate_vale_pdf(self, ticket_numbers, tipo_trabajo, referencia, tallas_cantidades, color, total_producido, serial_code_path):
        """
        Generate a PDF work voucher with multiple tickets on a single page.
        """
        # Create PDF filename and path
        pdf_filename = f"vale_{tipo_trabajo}_{'_'.join(map(str, ticket_numbers))}.pdf" # Asegurarse que ticket_numbers sean strings para el join si son ints
        pdf_path = os.path.join("codes", pdf_filename)

        # Create the PDF document
        doc = SimpleDocTemplate(pdf_path, pagesize=letter,
                                leftMargin=30, rightMargin=30,
                                topMargin=30, bottomMargin=30)
        styles = getSampleStyleSheet()
        elements = []

        # Define sizes (columns: 33 to 48 inclusive)
        sizes = list(range(33, 49))
        size_row = [str(size) for size in sizes]

        # Calculate how many vales we need to fill the PDF
        vales_per_page = 5
        total_pages_needed = 1
        total_vales_needed = vales_per_page * total_pages_needed
        
        extended_ticket_numbers = []
        ticket_count = len(ticket_numbers)
        
        for i in range(total_vales_needed):
            ticket_index = i % ticket_count
            extended_ticket_numbers.append(ticket_numbers[ticket_index])
        
        for i, ticket_number in enumerate(extended_ticket_numbers):
            # --- MODIFICACIÓN AQUÍ ---
            # Prepare quantities row
            # Para cada talla en 'sizes', si esa talla (convertida a string) está como clave en 'tallas_cantidades',
            # entonces usa 'total_producido'. De lo contrario, usa '0'.
            qty_row = [str(total_producido) if str(size) in tallas_cantidades else '' for size in sizes]
            # --- FIN DE LA MODIFICACIÓN ---

            # Prepare barcode image
            barcode_img = Image(serial_code_path, width=100, height=40) # Ajusta width/height según necesites

            # Create main container table
            container_data = [[
                Paragraph(f"{tipo_trabajo.upper()}", styles['Heading2']),
                "",
                barcode_img
            ]]
            container_table = Table(container_data, colWidths=[300, 50, 180]) # Ajustado para que sume cerca de 530
            container_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                # ('BOX', (0, 0), (-1, -1), 2, colors.green) # Comentado para que la tabla completa tenga el borde
            ]))

            # Ticket details table
            details_data = [
                ["Referencia:", referencia, "Color:", color, f"N° {ticket_number}"]
            ]
            # Ajustar colWidths para que sumen el ancho de la tabla contenedora (530)
            details_table = Table(details_data, colWidths=[80, 150, 50, 100, 150])
            details_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                # ('BOX', (0, 0), (-1, -1), 1, colors.green)
            ]))

            # Sizes and quantities table
            # El ancho de cada columna es (ancho total / número de tallas)
            # Ancho total disponible es 530. Número de tallas es len(size_row)
            col_width_size = 530 / len(size_row)
            sizes_table = Table([size_row, qty_row], colWidths=[col_width_size] * len(size_row))
            sizes_table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 1, colors.green),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')
            ]))

            # Footer table
            footer_data = [["Firma:", "________________", "Total:", str(total_producido)]]
            # Ajustar colWidths para que sumen el ancho de la tabla contenedora (530)
            footer_table = Table(footer_data, colWidths=[50, 250, 50, 180])
            footer_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('ALIGN', (2,0), (2,0), 'RIGHT'), # Total a la derecha
                ('ALIGN', (3,0), (3,0), 'RIGHT'), # Valor del total a la derecha
                # ('BOX', (0, 0), (-1, -1), 1, colors.green)
            ]))

            # Agrupar todos los componentes en una sola tabla contenedora
            complete_vale_data = [
                [container_table],
                [details_table],
                [sizes_table],
                [footer_table]
            ]
            
            complete_vale_table = Table(complete_vale_data, colWidths=[530]) # Ancho total del vale
            complete_vale_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2), # Pequeño espacio entre elementos internos
                ('BOX', (0,0), (-1,-1), 1, colors.green) # Borde verde para todo el vale
            ]))

            elements.append(KeepTogether([complete_vale_table]))
            
            if i < len(extended_ticket_numbers) - 1:
                elements.append(Spacer(1, 10)) # 10 puntos de espacio vertical

            if (i + 1) % vales_per_page == 0 and (i + 1) < len(extended_ticket_numbers):
                elements.append(PageBreak())

        doc.build(elements)
        return pdf_path
    
    def generate_vale_pdf_funcional(self, ticket_numbers, tipo_trabajo, referencia, tallas_cantidades, color, total_producido, serial_code_path):
        """
        Generate a PDF work voucher with multiple tickets on a single page.
        """
        # Create PDF filename and path
        pdf_filename = f"vale_{tipo_trabajo}_{ticket_numbers}.pdf"
        pdf_path = os.path.join("codes", pdf_filename)

        # Create the PDF document
        doc = SimpleDocTemplate(pdf_path, pagesize=letter,
                                leftMargin=30, rightMargin=30,
                                topMargin=30, bottomMargin=30)
        styles = getSampleStyleSheet()
        elements = []

        # Define sizes (columns: 33 to 48 inclusive)
        sizes = list(range(33, 49))
        size_row = [str(size) for size in sizes]

        # Calculate how many vales we need to fill the PDF
        vales_per_page = 5  # Based on your PageBreak logic
        total_pages_needed = 1  # You can adjust this or make it a parameter
        total_vales_needed = vales_per_page * total_pages_needed
        
        # If we have fewer ticket numbers than needed, repeat them
        extended_ticket_numbers = []
        ticket_count = len(ticket_numbers)
        
        for i in range(total_vales_needed):
            # Cycle through the ticket numbers if we need more vales than tickets
            ticket_index = i % ticket_count
            extended_ticket_numbers.append(ticket_numbers[ticket_index])
        
        # Create a frame for multiple tickets
        for i, ticket_number in enumerate(extended_ticket_numbers):
            # Prepare quantities row
            qty_row = [str(tallas_cantidades.get(str(size), 0)) for size in sizes]

            # Prepare barcode image
            barcode_img = Image(serial_code_path, width=100, height=40)

            # Create main container table
            container_data = [[
                Paragraph(f"{tipo_trabajo.upper()}", styles['Heading2']),
                "",
                barcode_img
            ]]
            container_table = Table(container_data, colWidths=[300, 50, 200])
            container_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOX', (0, 0), (-1, -1), 2, colors.green)
            ]))

            # Ticket details table
            details_data = [
                ["Referencia:", referencia, "Color:", color, f"N° {ticket_number}"]
            ]
            details_table = Table(details_data, colWidths=[100, 150, 80, 100, 100])
            details_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('BOX', (0, 0), (-1, -1), 1, colors.green)
            ]))

            # Sizes and quantities table
            sizes_table = Table([size_row, qty_row], colWidths=[30] * len(size_row))
            sizes_table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 1, colors.green),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')
            ]))

            # Footer table
            footer_data = [["Firma:", "", "", "Total:", str(total_producido)]]
            footer_table = Table(footer_data, colWidths=[100, 150, 100, 80, 100])
            footer_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('BOX', (0, 0), (-1, -1), 1, colors.green)
            ]))

            # **SOLUCIÓN**: Agrupar todos los componentes en una sola tabla contenedora
            complete_vale_data = [
                [container_table],
                [details_table],
                [sizes_table],
                [footer_table]
            ]
            
            complete_vale_table = Table(complete_vale_data, colWidths=[530])
            complete_vale_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ]))

            # Agregar el vale completo como una unidad
            elements.append(KeepTogether([complete_vale_table]))
            
            # Agregar espaciado entre vales
            if i < len(extended_ticket_numbers) - 1:
                elements.append(Spacer(1, 10))

            # Insert page break every 3 tickets
            if (i + 1) % 5 == 0 and (i + 1) < len(extended_ticket_numbers):
                elements.append(PageBreak())

        # Build PDF
        doc.build(elements)
        return pdf_path




if __name__ == "__main__":
    app = QApplication(sys.argv)
    QtGui.QFontDatabase.addApplicationFont('fonts/segoeui.ttf')
    QtGui.QFontDatabase.addApplicationFont('fonts/segoeuib.ttf')
    window = MainWindow()
    sys.exit(app.exec_())