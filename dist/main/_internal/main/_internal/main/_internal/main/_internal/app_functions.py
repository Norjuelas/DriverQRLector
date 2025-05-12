################################################################################
##
## PROJECT MADE WITH: Qt Designer and PySide2
## V: 1.0.0
##
## This project can be used freely for all uses, as long as they maintain the
## respective credits only in the Python scripts, any information in the visual
## interface (GUI) can be modified without any implication.
##
## There are limitations on Qt licenses if you want to use your products
## commercially, I recommend reading them on the official website:
## https://doc.qt.io/qtforpython/licenses.html
##
################################################################################

## ==> GUI FILE
from main import *
import uuid
import os
import barcode
from barcode.writer import ImageWriter
from openpyxl import Workbook, load_workbook
from PySide2.QtCore import QBuffer, QIODevice
from PySide2.QtGui import QPixmap, QImage
from PySide2.QtWidgets import QFileDialog, QMessageBox

class Functions(MainWindow):
    
    def setupBarCodeGenerator(self):
        """
        Configura la funcionalidad para generar códigos de barras, mostrarlos en un preview
        y guardar la información en Excel.
        """
        # Ruta para el archivo Excel que servirá como base de datos
        excel_path = "trabajos_database.xlsx"
        
        # Carpeta para guardar las imágenes de códigos de barras
        if not os.path.exists("barcodes"):
            os.makedirs("barcodes")
        
        # Crear el archivo Excel si no existe
        if not os.path.exists(excel_path):
            wb = Workbook()
            ws = wb.active
            ws.append(["Código Serial", "Tipo Trabajo", "Referencia", "Número Ticket", 
                    "Talla", "Color", "Valor", "Total Producido", "Ruta Imagen"])
            wb.save(excel_path)
        
        # Función para generar código serial único
        def generate_serial_code():
            # Combina los valores de los campos para crear un código único
            tipo_trabajo = self.CampoTipoTrabajo.text().strip()
            referencia = self.CampoReferenciaTrabajo.text().strip()
            num_ticket = self.CampoNumeroTicket.text().strip()
            talla = self.CampoTalla.text().strip()
            color = self.CampoColor.text().strip()
            
            # Genera un UUID único y toma los primeros 8 caracteres
            unique_id = str(uuid.uuid4())[:8]
            
            # Crea el código serial combinando la información
            serial_code = f"{tipo_trabajo[:2]}{referencia[:3]}{num_ticket}{talla}{color[:1]}-{unique_id}"
            return serial_code.upper()
        
        # Función para generar y guardar el código de barras
        def generate_barcode(serial_code):
            try:
                # Genera el código de barras (Code128 es un formato común y versátil)
                barcode_class = barcode.get_barcode_class('code128')
                barcode_instance = barcode_class(serial_code, writer=ImageWriter())
                
                # Guarda el código de barras como imagen PNG
                barcode_path = f"barcodes/{serial_code}"
                barcode_filename = barcode_instance.save(barcode_path)
                
                return barcode_filename
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al generar el código de barras: {str(e)}")
                return None
        
        # Función para mostrar la imagen en el QGraphicsView
        def display_barcode(barcode_path):
            try:
                # Cargar la imagen
                pixmap = QPixmap(barcode_path)
                
                # Obtener la escena del QGraphicsView
                scene = self.PreviwImage.scene()
                if scene is None:
                    scene = QGraphicsScene()
                    self.PreviwImage.setScene(scene)
                else:
                    scene.clear()
                
                # Añadir la imagen a la escena
                scene.addPixmap(pixmap)
                
                # Ajustar la vista
                self.PreviwImage.fitInView(scene.sceneRect(), Qt.KeepAspectRatio)
                
                return True
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al mostrar la imagen: {str(e)}")
                return False
        
        # Función para guardar los datos en Excel
        def save_to_excel(serial_code, barcode_path):
            try:
                # Obtener los valores de los campos
                tipo_trabajo = self.CampoTipoTrabajo.text().strip()
                referencia = self.CampoReferenciaTrabajo.text().strip()
                num_ticket = self.CampoNumeroTicket.text().strip()
                talla = self.CampoTalla.text().strip()
                color = self.CampoColor.text().strip()
                valor = self.CampoValor.text().strip()
                total_producido = self.CampoTotaProducido.text().strip()
                
                # Cargar el archivo Excel existente
                wb = load_workbook(excel_path)
                ws = wb.active
                
                # Añadir nueva fila con los datos
                ws.append([serial_code, tipo_trabajo, referencia, num_ticket, 
                        talla, color, valor, total_producido, barcode_path])
                
                # Guardar el archivo Excel
                wb.save(excel_path)
                
                return True
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al guardar en Excel: {str(e)}")
                return False
        
        # Función principal que se ejecuta al hacer clic en el botón guardar
        def on_save_button_clicked():
            # Validar que todos los campos obligatorios estén completos
            required_fields = [
                (self.CampoTipoTrabajo, "Tipo de Trabajo"),
                (self.CampoReferenciaTrabajo, "Referencia"),
                (self.CampoNumeroTicket, "Número de Ticket"),
                (self.CampoTalla, "Talla"),
                (self.CampoColor, "Color"),
                (self.CampoValor, "Valor"),
                (self.CampoTotaProducido, "Total Producido")
            ]
            
            for field, name in required_fields:
                if not field.text().strip():
                    QMessageBox.warning(self, "Campos Incompletos", f"El campo {name} es obligatorio.")
                    field.setFocus()
                    return
            
            # Generar el código serial único
            serial_code = generate_serial_code()
            
            # Generar y guardar el código de barras
            barcode_path = generate_barcode(serial_code)
            if not barcode_path:
                return
            
            # Mostrar el código de barras en el QGraphicsView
            if not display_barcode(barcode_path):
                return
            
            # Guardar los datos en Excel
            if save_to_excel(serial_code, barcode_path):
                QMessageBox.information(
                    self, 
                    "Operación Exitosa", 
                    f"Código generado: {serial_code}\nLos datos se han guardado correctamente."
                )
                
                # Limpiar los campos después de guardar
                for field, _ in required_fields:
                    field.clear()
                
                # Enfocar el primer campo para un nuevo registro
                self.CampoTipoTrabajo.setFocus()
        
        # Conectar el botón guardar con la función
        self.pushButtonGuardar.clicked.connect(on_save_button_clicked)
        
        # Asegurarse de que el QGraphicsView tenga una escena
        if self.PreviwImage.scene() is None:
            self.PreviwImage.setScene(QGraphicsScene())