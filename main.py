import sys
import os
import platform
import uuid
import traceback
import datetime as dt
import time
import pandas as pd

from openpyxl import Workbook, load_workbook

from PySide2 import QtCore, QtGui, QtWidgets
from PySide2.QtCore import (
    QCoreApplication, QPropertyAnimation, QDate, QDateTime, QMetaObject, 
    QObject, QPoint, QRect, QSize, QTime, QThread, QUrl, Qt, QEvent
)
from PySide2.QtGui import (
    QBrush, QColor, QConicalGradient, QCursor, QFont, QFontDatabase, 
    QIcon, QKeySequence, QLinearGradient, QPalette, QPainter, QPixmap, QRadialGradient
)
from PySide2.QtWidgets import *
from openpyxl.styles import Font, PatternFill, Alignment

from PySide2.QtWidgets import QTableView, QVBoxLayout
from PySide2.QtGui import QStandardItemModel

from generate_pdf import generate_vale_pdf

from utils import validate_cedula, display_code_image

from config import TIPOS_DE_TRABAJO, WORK_TYPE_ABBREVIATIONS, CAMPOS_VALOR_TRABAJO_MAP    
# GUI FILE
from app_modules import *
from vales_card import ValeCard, DropColumnWidget

from PySide2.QtWidgets import QWidget, QVBoxLayout, QLabel, QTreeWidget, QTreeWidgetItem, QHeaderView
from PySide2.QtCore import Qt

DATOS_VALES = [
    {"ticket": "T001", "referencia": "REF-A", "color": "Rojo", "total": 100, "estado": "impreso", "Satelite": False},
    {"ticket": "T002", "referencia": "REF-B", "color": "Azul", "total": 150, "estado": "impreso", "Satelite": True},
    {"ticket": "T003", "referencia": "REF-C", "color": "Verde", "total": 200, "estado": "salido", "Satelite": True},
    {"ticket": "T004", "referencia": "REF-D", "color": "Negro", "total": 50, "estado": "entro", "Satelite": True},
    {"ticket": "T005", "referencia": "REF-E", "color": "Blanco", "total": 300, "estado": "pagado", "Satelite": False},
    {"ticket": "T006", "referencia": "REF-F", "color": "Amarillo", "total": 120, "estado": "pagado", "Satelite": True},
    {"ticket": "T007", "referencia": "REF-G", "color": "Gris", "total": 90, "estado": "impreso", "Satelite": False},
]
class MainWindow(QMainWindow):
    def __init__(self):
        QMainWindow.__init__(self)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.dragPos = None

        # Current code type (barcode or qr)
        self.current_code_type = "barcode"  # Default to barcode

        # Print system information
        print('System: ' + platform.system())
        print('Version: ' + platform.release())

        # Setup the window
        self.setup_window()
        
        # Setup menus
        self.setup_menus()
        
        # Setup barcode generator functionality
        self.setup_code_generator()
        
        # Setup code reader functionality
        self.setup_code_reader()
        
        # Setup employee management
        self.setup_employee_management()

        # Conectar el botón de eliminarTODO (agregar esta línea)
        self.setup_eliminar_todo_button()

        # Setup update reports button
        self.setup_update_button()
        
        # Setup add employee button
        self.setup_add_employee_button()

        self.autocompletado_manager = AutocompletadoManager(excel_path=self.excel_path, sheet_name="Trabajos")
        self.datos_vales = DATOS_VALES # O cárgalos desde tu Excel aquí

        self.setup_autocompletado_fields()
        self.setup_kanban_board()
        self.actualizar_dashboard_pagos()
        #self.ui.btnRefrescarPagos.clicked.connect(self.actualizar_dashboard_pagos)
        # Show t    he window
        self.show()


    def setup_code_reader(self):
        """Setup the code reader functionality and TableView inside WidgetTabla"""
        
        # El codeReaderInput YA existe en tu UI, solo necesitamos conectar el signal
        if hasattr(self.ui, 'codeReaderInput'):
            print("codeReaderInput encontrado en UI")
            # Configurar el input existente
            self.ui.codeReaderInput.setPlaceholderText("Escanear o ingresar código aquí...")
            self.ui.codeReaderInput.setMinimumWidth(250)
            
            # IMPORTANTE: Conectar el signal returnPressed
            self.ui.codeReaderInput.returnPressed.connect(self.on_code_scanned)
            print("Signal returnPressed conectado")
        else:
            print("ERROR: codeReaderInput no encontrado en UI")
            return
        
        # Access the WidgetTabla from UI and setup table
        if hasattr(self.ui, 'WidgetTabla'):
            print("WidgetTabla encontrado")
            
            # Solo crear el layout si WidgetTabla no tiene uno
            if self.ui.WidgetTabla.layout() is None:
                table_layout = QVBoxLayout(self.ui.WidgetTabla)
                self.ui.WidgetTabla.setLayout(table_layout)
                print("Layout creado para WidgetTabla")
            else:
                table_layout = self.ui.WidgetTabla.layout()
                print("Usando layout existente de WidgetTabla")
            
            # Solo crear la tabla si no existe
            if not hasattr(self.ui, 'tableViewVale'):
                print("Creando tableViewVale")
                # Create TableView
                self.ui.tableViewVale = QTableView(self.ui.WidgetTabla)
                self.ui.tableViewVale.setMinimumHeight(200)
                
                # Create the model for the table
                self.table_model = QStandardItemModel()
                # Definir encabezados: fijos + dinámicos basados en WORK_TYPE_ABBREVIATIONS
                fixed_headers = ["Código Serial", "Número Ticket", "Referencia", "Color", "Total Producido"]
                valor_headers = [f"Valor {work_type}" for work_type in WORK_TYPE_ABBREVIATIONS.keys()]
                self.table_model.setHorizontalHeaderLabels(fixed_headers + valor_headers)
                
                # Set model and adjust view
                self.ui.tableViewVale.setModel(self.table_model)
                # Apply stretch mode to all columns
                header = self.ui.tableViewVale.horizontalHeader()
                for i in range(self.table_model.columnCount()):
                    header.setSectionResizeMode(i, QHeaderView.Stretch)
                
                # Add TableView to WidgetTabla's layout
                table_layout.addWidget(self.ui.tableViewVale)
                print("tableViewVale creado y agregado")
            else:
                print("tableViewVale ya existe")
        else:
            print("Warning: WidgetTabla not found in UI")


    def setup_code_generator(self):
        """
        Configura la funcionalidad del generador de códigos y el archivo Excel.
        (Actualizado para usar los nuevos tipos de trabajo)
        """
        self.excel_path = "trabajos_database.xlsx"
        self.vales_sheet_name = "Vales"

        if not os.path.exists("codes"):
            os.makedirs("codes")

        # Usar los nuevos tipos de trabajo definidos globalmente o en la clase
        work_types = list(WORK_TYPE_ABBREVIATIONS.keys())

        # Definir las cabeceras para la hoja "Trabajos"
        trabajos_headers = ["Código Serial", "Número Ticket", "Referencia", "Color"]
        for i in range(33, 49):
            trabajos_headers.append(f"Cant_T{i}")
        trabajos_headers.append("Total Producido")

        # Añadir cabeceras para los valores de los nuevos tipos de trabajo
        trabajos_headers.extend([f"Valor {wt}" for wt in work_types])

        trabajos_headers.extend(["Tipo Código", "Ruta Imagen"])

        # Añadir columnas para los códigos seriales por TIPO NUEVO de trabajo
        trabajos_headers.extend([f"Código_{wt}" for wt in work_types])

        # Definir las cabeceras para la hoja "Vales" (Mantener si no cambia)
        vales_headers = [
            "ID_Vale", "EmpleadoID", "FechaHora_Generacion", "Numero_Ticket_Asociado",
            "Referencia_Asociada", "Color_Trabajo", "Resumen_Tallas_Cantidades",
            "Total_Producido_Trabajo", "Suma_Valores_Trabajos", "Codigo_Serial_Trabajo_Asociado",
            "WorkTypeDetected"
        ]

        # Crear/Cargar Excel y verificar/actualizar cabeceras (Tu lógica existente)
        if not os.path.exists(self.excel_path):
            wb = Workbook()
            ws_trabajos = wb.active
            ws_trabajos.title = "Trabajos"
            ws_trabajos.append(trabajos_headers)
            ws_vales = wb.create_sheet(title=self.vales_sheet_name)
            ws_vales.append(vales_headers)
            wb.save(self.excel_path)
        else:
            wb = load_workbook(self.excel_path)
            ws_trabajos = wb["Trabajos"]
            current_headers = [cell.value for cell in ws_trabajos[1]]
            if current_headers != trabajos_headers:
                print("Actualizando cabeceras de la hoja 'Trabajos'...")
                # Considera una migración más segura si ya hay datos
                ws_trabajos.delete_rows(1)
                ws_trabajos.insert_rows(1)
                for col_idx, header in enumerate(trabajos_headers, 1):
                    ws_trabajos.cell(row=1, column=col_idx).value = header
                wb.save(self.excel_path)

            if self.vales_sheet_name not in wb.sheetnames:
                ws_vales = wb.create_sheet(title=self.vales_sheet_name)
                ws_vales.append(vales_headers)
                wb.save(self.excel_path)

        # Conectar botón (Tu lógica existente)
        if hasattr(self.ui, 'pushButtonGuardar'):
            self.ui.pushButtonGuardar.clicked.connect(self.on_save_button_clicked)
        else:
            print("Advertencia: self.ui.pushButtonGuardar no encontrado.")

        # Configurar escena (Tu lógica existente)
        if hasattr(self.ui, 'PreviwImage'):
            if self.ui.PreviwImage.scene() is None:
                self.ui.PreviwImage.setScene(QtWidgets.QGraphicsScene(self))
        else:
            print("Advertencia: self.ui.PreviwImage no encontrado.")

    def save_to_excel(self, serial_codes, code_path, ticket_number, referencia, color, tallas_cantidades, total_producido_calculado, valores_trabajo):
        """
        Guarda los datos en la hoja 'Trabajos' del archivo Excel.
        """
        try:
            wb = load_workbook(self.excel_path)
            ws_trabajos = wb["Trabajos"]

            # Crear una nueva fila con los datos
            row_data = [
                list(serial_codes.values())[0] if serial_codes else "",  # Código Serial (primer código generado)
                ticket_number,
                referencia,
                color
            ]
            # Añadir cantidades por talla (33 a 48)
            for i in range(33, 49):
                row_data.append(tallas_cantidades.get(str(i), 0))
            # Añadir total producido y valores por tipo de trabajo
            row_data.append(total_producido_calculado)
            # Añadir valores para cada tipo de trabajo según WORK_TYPE_ABBREVIATIONS
            for work_type in WORK_TYPE_ABBREVIATIONS.keys():
                row_data.append(valores_trabajo.get(work_type, 0))
            # Añadir tipo de código y ruta de la imagen
            row_data.extend([
                self.current_code_type.upper(),  # Tipo Código
                code_path  # Ruta Imagen
            ])
            # Añadir códigos seriales por tipo de trabajo
            for work_type in WORK_TYPE_ABBREVIATIONS.keys():
                row_data.append(serial_codes.get(work_type, ""))

            # Añadir la fila a la hoja "Trabajos"
            ws_trabajos.append(row_data)
            wb.save(self.excel_path)
            return True
        except PermissionError as e:
            if e.errno == 13:  # Errno 13 is Permission Denied
                QMessageBox.critical(self, "Error al Guardar",
                                    f"No se pudo guardar en '{self.excel_path}'.\n\n"
                                    "Por favor, asegúrese de que el archivo Excel no esté abierto en otro programa e inténtelo de nuevo.")
            else:
                QMessageBox.critical(self, "Error de Permiso", f"Error de permiso al guardar en Excel: {e}")
            return False
        except Exception as e:
            print(f"Error al guardar en Excel: {e}")
            QMessageBox.critical(self, "Error", f"Error al guardar en Excel: {e}")
            return False

    def find_code_data(self, serial_code):
        """Find data related to a specific serial code in the Excel file"""
        try:
            # Load Excel file
            wb = load_workbook(self.excel_path)
            ws = wb["Trabajos"]  # Explicitly use "Trabajos" sheet

            # Obtener encabezados para buscar índices de columnas
            headers = [cell.value for cell in ws[1]]
            try:
                serial_code_idx = headers.index("Código Serial")
                num_ticket_idx = headers.index("Número Ticket")
                referencia_idx = headers.index("Referencia")
                color_idx = headers.index("Color")
                total_producido_idx = headers.index("Total Producido")
                # Mapear índices de columnas para valores de trabajo
                valor_indices = {}
                for work_type in WORK_TYPE_ABBREVIATIONS.keys():
                    valor_column = f"Valor {work_type}"
                    if valor_column in headers:
                        valor_indices[work_type] = headers.index(valor_column)
            except ValueError as e:
                QMessageBox.critical(self, "Error", f"Columna faltante en 'Trabajos': {e}")
                return None

            # Search for the serial code in the first column
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[serial_code_idx] == serial_code:
                    # Construir diccionario de datos
                    data = {
                        "serial_code": row[serial_code_idx],
                        "num_ticket": row[num_ticket_idx],
                        "referencia": row[referencia_idx],
                        "color": row[color_idx],
                        "total_producido": row[total_producido_idx]
                    }
                    # Añadir valores de trabajo
                    for work_type, idx in valor_indices.items():
                        key = f"valor_{work_type.lower().replace(' ', '_')}"
                        data[key] = row[idx] if idx < len(row) else 0
                    return data

            return None
        except Exception as e:
            print(f"Error al buscar datos: {e}")
            QMessageBox.critical(self, "Error", f"Error al buscar datos: {str(e)}")
            return None

    def setup_update_button(self):
        """Conecta el botón ActualizarDB para generar reportes de empleados."""
        if hasattr(self.ui, 'btnActualizarDB'):
            try:
                self.ui.btnActualizarDB.clicked.disconnect()
            except RuntimeError:
                pass
            self.ui.btnActualizarDB.clicked.connect(self.update_employee_reports)
            self.ui.btnActualizarDB.setEnabled(True)
            print("Botón 'btnActualizarDB' conectado correctamente.")
        else:
            print("ERROR: El QPushButton 'btnActualizarDB' no se encontró en la UI.")

    def setup_autocompletado_fields(self):
        """
        Configura el autocompletado para los campos QLineEdit relevantes.
        """
        # Define los campos QLineEdit de la UI y las columnas de Excel correspondientes
        # Formato: 'identificador_unico': {'line_edit': self.ui.NombreLineEdit, 'columna': 'NombreColumnaEnExcel'}
        campos_a_configurar = {}

        # Campos de texto obligatorios según on_save_button_clicked
        if hasattr(self.ui, 'CampoReferenciaTrabajo'):
            campos_a_configurar['referencia'] = {
                'line_edit': self.ui.CampoReferenciaTrabajo,
                'columna': 'Referencia'
            }
        if hasattr(self.ui, 'CampoNumeroTicket'):
            campos_a_configurar['numero_ticket'] = {
                'line_edit': self.ui.CampoNumeroTicket,
                'columna': 'Número Ticket'
            }
        if hasattr(self.ui, 'CampoColor'):
            campos_a_configurar['color'] = {
                'line_edit': self.ui.CampoColor,
                'columna': 'Color'
            }

        # Opcional: Configurar autocompletado para tallas individuales
        # Asumimos que las tallas están en columnas separadas (Cant_T33 a Cant_T48) en el Excel
        for i in range(33, 49):  # Tallas de 33 a 48
            field_name = f'CampoTalla_{i}'
            if hasattr(self.ui, field_name):
                campos_a_configurar[f'talla_{i}'] = {
                    'line_edit': getattr(self.ui, field_name),
                    'columna': f'Cant_T{i}'  # Nombre de columna en el Excel (Cant_T33, Cant_T34, etc.)
                }

        # Configurar autocompletado para valores de trabajo usando CAMPOS_VALOR_TRABAJO_MAP
        for work_type, field_name in CAMPOS_VALOR_TRABAJO_MAP.items():
            if hasattr(self.ui, field_name):
                campos_a_configurar[field_name.lower()] = {
                    'line_edit': getattr(self.ui, field_name),
                    'columna': f'Valor {work_type}'
                }

        if campos_a_configurar:
            self.autocompletado_manager.configurar_multiples_campos(campos_a_configurar)
            print("Autocompletado configurado para los campos.")
        else:
            print("No se encontraron campos para configurar el autocompletado.")


# -----------------------------------------------------------------------
    def setup_kanban_board(self):
        """
        Configura las columnas de la UI como zonas de destino para el Drag & Drop.
        """
        # 1. Crear las instancias de nuestras columnas lógicas
        self.col_impresos_empleados = DropColumnWidget("impreso")
        self.col_impresos_satelite = DropColumnWidget("impreso")
        self.col_salidos = DropColumnWidget("salido")
        self.col_entraron = DropColumnWidget("entro")
        self.col_pagos = DropColumnWidget("pagado")
        # Para la columna de "Impresos > Empleados"
        layout_impresos_emp = QVBoxLayout(self.ui.widget_impresos_empleado_container)
        layout_impresos_emp.addWidget(self.col_impresos_empleados)

        # Para la columna de "Impresos > Satélite"
        layout_impresos_sat = QVBoxLayout(self.ui.widget_impresos_satelite_container)
        layout_impresos_sat.addWidget(self.col_impresos_satelite)
        
        # Para la columna de "Salidos"
        layout_salidos = QVBoxLayout(self.ui.Vales_salidos)
        layout_salidos.addWidget(self.col_salidos)

        # Para la columna de "Entraron"
        layout_entraron = QVBoxLayout(self.ui.Vales_entraron)
        layout_entraron.addWidget(self.col_entraron)

        # Para la columna de "Pagados"
        layout_pagos = QVBoxLayout(self.ui.Vales_pagos)
        layout_pagos.addWidget(self.col_pagos)
        
        # 3. Conectar las señales de drop al manejador
        self.col_impresos_empleados.card_dropped.connect(self.handle_card_drop)
        self.col_impresos_satelite.card_dropped.connect(self.handle_card_drop)
        self.col_salidos.card_dropped.connect(self.handle_card_drop)
        self.col_entraron.card_dropped.connect(self.handle_card_drop)
        self.col_pagos.card_dropped.connect(self.handle_card_drop)

        # 4. Poblar el tablero por primera vez
        self.update_kanban_view()

    def handle_card_drop(self, ticket_id, nuevo_estado):
        """
        Maneja la lógica de actualizar el estado de un vale cuando se suelta.
        """
        vale_a_mover = next((vale for vale in self.datos_vales if vale["ticket"] == ticket_id), None)
        
        if vale_a_mover:
            es_satelite = vale_a_mover.get("Satelite", False)
            if nuevo_estado in ["salido", "entro"] and not es_satelite:
                print(f"Movimiento no permitido: El vale {ticket_id} (Empleado) no puede ir a '{nuevo_estado}'.")
                return # El movimiento se cancela

            vale_a_mover["estado"] = nuevo_estado
            print(f"Ticket {ticket_id} actualizado al estado '{nuevo_estado}'")
            self.update_kanban_view() # Redibuja todo el tablero

    def clear_kanban_board(self):
        """Limpia todas las tarjetas de todas las columnas."""
        self.col_impresos_empleados.clear()
        self.col_impresos_satelite.clear()
        self.col_salidos.clear()
        self.col_entraron.clear()
        self.col_pagos.clear()

    def update_kanban_view(self):
        """Limpia el tablero y lo vuelve a poblar con los datos de self.datos_vales."""
        self.clear_kanban_board()

        for vale in self.datos_vales:
            card = ValeCard(vale)
            estado = vale.get("estado", "").lower()
            es_satelite = vale.get("Satelite", False)

            if estado == "impreso":
                if es_satelite:
                    self.col_impresos_satelite.add_card(card)
                else:
                    self.col_impresos_empleados.add_card(card)
            elif estado == "salido":
                self.col_salidos.add_card(card)
            elif estado == "entro":
                self.col_entraron.add_card(card)
            elif estado == "pagado":
                self.col_pagos.add_card(card)
    def create_impresos_column(self):
        """ Crea la columna especial "Vales Impresos" con sus dos subgrupos. """
        container = QGroupBox("Vales Impresos")
        self.apply_groupbox_style(container)
        
        layout = QVBoxLayout(container)
        
        # Subgrupo Empleados
        group_empleados = QGroupBox("Empleados")
        layout_empleados = QVBoxLayout(group_empleados)
        layout_empleados.addWidget(self.create_scroll_area(self.col_impresos_empleados))
        layout.addWidget(group_empleados)
        
        # Subgrupo Satélite
        group_satelite = QGroupBox("Satélite")
        layout_satelite = QVBoxLayout(group_satelite)
        layout_satelite.addWidget(self.create_scroll_area(self.col_impresos_satelite))
        layout.addWidget(group_satelite)
        
        return container

    def create_scrollable_column(self, title, content_widget):
        """Crea una columna genérica con título y área de scroll."""
        column_container = QGroupBox(title)
        self.apply_groupbox_style(column_container)
        main_vbox = QVBoxLayout(column_container)
        main_vbox.addWidget(self.create_scroll_area(content_widget))
        return column_container

    def create_scroll_area(self, content_widget):
        """Crea un QScrollArea para un widget de contenido."""
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(content_widget)
        scroll.setStyleSheet("QScrollArea { border: none; background-color: transparent; }")
        return scroll

    def apply_groupbox_style(self, groupbox):
        """Aplica un estilo consistente a los QGroupBox de las columnas."""
        groupbox.setStyleSheet("""
            QGroupBox {
                font-size: 16px; font-weight: bold; background-color: #F5F5F5;
                border: 1px solid #D0D0D0; border-radius: 8px; margin-top: 1ex;
            }
            QGroupBox::title {
                subcontrol-origin: margin; subcontrol-position: top center; padding: 0 10px;
            }
        """)

    def handle_card_drop(self, ticket_id, nuevo_estado):
        """
        Esta es la función principal que actualiza los datos cuando una tarjeta se mueve.
        """
        print(f"Moviendo ticket {ticket_id} al estado '{nuevo_estado}'")
        
        # 1. Encontrar el vale en nuestra lista de datos
        vale_a_mover = None
        for vale in DATOS_VALES:
            if vale["ticket"] == ticket_id:
                vale_a_mover = vale
                break
        
        if vale_a_mover:
            # 2. Validar si el movimiento es permitido
            es_satelite = vale_a_mover.get("Satelite", False)
            if nuevo_estado in ["salido", "entro"] and not es_satelite:
                print(f"Movimiento no permitido: El vale {ticket_id} es de Empleado y no puede ir a '{nuevo_estado}'.")
                # No hacemos nada y el tablero no se actualizará, la tarjeta volverá a su sitio.
                return

            # 3. Actualizar el estado en la lista de datos
            vale_a_mover["estado"] = nuevo_estado
            
            # 4. Volver a dibujar todo el tablero con los datos actualizados
            self.update_kanban_view(DATOS_VALES)

    def clear_kanban_board(self):
        """Limpia todas las tarjetas de todas las columnas."""
        self.col_impresos_empleados.clear()
        self.col_impresos_satelite.clear()
        self.col_salidos.clear()
        self.col_entraron.clear()
        self.col_pagos.clear()

    def update_kanban_view(self):
        """Limpia el tablero y lo vuelve a poblar con los datos actualizados."""
        self.clear_kanban_board()

        for vale in self.datos_vales:
            card = ValeCard(vale)
            
            estado = vale.get("estado", "").lower()
            es_satelite = vale.get("Satelite", False)

            if estado == "impreso":
                if es_satelite:
                    self.col_impresos_satelite.add_card(card)
                else:
                    self.col_impresos_empleados.add_card(card)
            elif estado == "salido":
                self.col_salidos.add_card(card)
            elif estado == "entro":
                self.col_entraron.add_card(card)
            elif estado == "pagado":
                self.col_pagos.add_card(card)

#--------------------------------------------------------------------------- 


#--------------------------------------------------------------------------- 
# Asegúrate de tener esta importación al inicio de tu archivo

    def actualizar_dashboard_pagos(self):
        """
        Calcula y muestra los consolidados de pagos para empleados y satélites.
        """
        print("Actualizando dashboard de pagos...")
        datos_de_prueba = [
        {'ticket': 'T101', 'referencia': 'REF-A', 'total': 150, 'estado': 'pagado', 'Satelite': False, 'responsable': 'Ana Gomez'},
        {'ticket': 'T102', 'referencia': 'REF-A', 'total': 50, 'estado': 'pagado', 'Satelite': False, 'responsable': 'Ana Gomez'},
        {'ticket': 'T103', 'referencia': 'REF-B', 'total': 200, 'estado': 'pagado', 'Satelite': False, 'responsable': 'Carlos Diaz'},
        {'ticket': 'T201', 'referencia': 'REF-C', 'total': 500, 'estado': 'pagado', 'Satelite': True, 'responsable': 'Taller Externo 1'},
        {'ticket': 'T202', 'referencia': 'REF-C', 'total': 450, 'estado': 'pagado', 'Satelite': True, 'responsable': 'Taller Externo 1'},
        {'ticket': 'T301', 'referencia': 'REF-D', 'total': 100, 'estado': 'impreso', 'Satelite': False, 'responsable': 'Ana Gomez'},
    ]
        # 1. PREPARACIÓN DE DATOS CON PANDAS
        # ------------------------------------
        if not self.datos_vales:
            print("No hay datos de vales para procesar.")
            return

        # Convertir la lista de diccionarios a un DataFrame de Pandas
        df = pd.DataFrame(datos_de_prueba)
        # Filtrar solo los vales que han sido pagados
        pagos_df = df[df['estado'].str.lower() == 'pagado'].copy()
        
        if pagos_df.empty:
            print("No se encontraron vales pagados.")
            # Aquí podrías limpiar los widgets si lo deseas
            return

        # Convertir 'total' a numérico, manejando errores
        pagos_df['total'] = pd.to_numeric(pagos_df['total'], errors='coerce').fillna(0)

        # Separar dataframes para empleados y satélites
        pagos_empleados_df = pagos_df[pagos_df['Satelite'] == False]
        pagos_satelites_df = pagos_df[pagos_df['Satelite'] == True]

        # 2. POBLAR DASHBOARDS DE RESUMEN (Widget_Pagos_*)
        # -----------------------------------------------
        self._crear_panel_resumen(
            parent_widget=self.ui.Widget_Pagos_empleados,
            df_datos=pagos_empleados_df,
            titulo="Consolidado Empleados"
        )
        self._crear_panel_resumen(
            parent_widget=self.ui.Widget_Pagos_Satelites,
            df_datos=pagos_satelites_df,
            titulo="Consolidado Satélites"
        )

        # 3. POBLAR VISTAS DETALLADAS (scrollAreaWidget_Pagos_*)
        # -----------------------------------------------------
        self._poblar_arbol_pagos(
            scroll_area=self.ui.scrollAreaWidget_Pagos_empleados,
            df_datos=pagos_empleados_df
        )
        self._poblar_arbol_pagos(
            scroll_area=self.ui.scrollAreaWidget_Pagos_Satelites,
            df_datos=pagos_satelites_df
        )

    def _crear_panel_resumen(self, parent_widget, df_datos, titulo):
        """
        Función auxiliar para crear los widgets del dashboard de resumen.
        """
        # Limpiar widget anterior
        for i in reversed(range(parent_widget.layout().count())): 
            parent_widget.layout().itemAt(i).widget().setParent(None)

        # Cálculos
        total_pagado = df_datos['total'].sum()
        cantidad_tickets = len(df_datos)
        # ⚠️ ADAPTAR 'responsable' si tu clave se llama diferente
        responsables_unicos = df_datos['responsable'].nunique()

        # Crear etiquetas y añadirlas al layout
        layout = parent_widget.layout()
        
        titulo_label = QLabel(f"<b>{titulo}</b>")
        titulo_label.setStyleSheet("font-size: 16px; margin-bottom: 10px;")
        
        total_label = QLabel(f"<b>Total Pagado:</b> ${total_pagado:,.0f}")
        tickets_label = QLabel(f"<b>Tickets Pagados:</b> {cantidad_tickets}")
        responsables_label = QLabel(f"<b>Total Empleados/Satélites:</b> {responsables_unicos}")

        layout.addWidget(titulo_label)
        layout.addWidget(total_label)
        layout.addWidget(tickets_label)
        layout.addWidget(responsables_label)
        layout.addStretch() # Empuja todo hacia arriba

    def _poblar_arbol_pagos(self, scroll_area, df_datos):
        """
        Función auxiliar para crear y poblar el QTreeWidget con datos agrupados.
        """
        if df_datos.empty:
            if scroll_area.widget():
                scroll_area.widget().clear()
            return
            
        datos_agrupados = df_datos.groupby(['responsable', 'referencia'])['total'].agg(['sum', 'count'])

        tree = QTreeWidget()
        tree.setHeaderLabels(["Responsable / Referencia", "Total Pagado", "Nº Tickets"])
        tree.header().setSectionResizeMode(0, QHeaderView.Stretch)

        responsable_actual = ""
        parent_item = None

        for (responsable, referencia), group in datos_agrupados.iterrows():
            if responsable != responsable_actual:
                responsable_actual = responsable
                parent_item = QTreeWidgetItem(tree, [responsable])
                
                # --- LÍNEA CORREGIDA ---
                # En lugar de setStyleSheet, modificamos la fuente del ítem
                font = parent_item.font(0)
                font.setBold(True)
                parent_item.setFont(0, font)
                # ----------------------

            child_item = QTreeWidgetItem(parent_item, [
                f"    └─ {referencia}", 
                f"${group['sum']:,.0f}", 
                str(group['count'])
            ])

        tree.expandAll()
        
        scroll_area.setWidget(tree)
        scroll_area.setWidgetResizable(True)
#--------------------------------------------------------------------------- 


    def update_employee_reports(self):
        """
        Función combinada que:
        1. Genera un reporte consolidado por empleado y lo muestra en tableViewVale
        2. Crea hojas individuales por empleado con detalles completos y consolidados
        3. Maneja la nueva estructura donde cada vale tiene valores diferenciados por tipo de trabajo
        """
        try:
            print("Actualizando reportes de empleados...")
            
            # Cargar el archivo Excel
            wb = load_workbook(self.excel_path)
            
            # Verificar si existen las hojas necesarias
            if "Vales" not in wb.sheetnames or "Empleados" not in wb.sheetnames:
                QMessageBox.warning(self, "Error", "No se encontraron las hojas 'Vales' o 'Empleados' en el archivo Excel.")
                return
            
            ws_vales = wb[self.vales_sheet_name]
            ws_empleados = wb["Empleados"]

            # Obtener información de empleados
            empleados = {}
            for row in ws_empleados.iter_rows(min_row=2, max_col=5, values_only=True):
                if len(row) >= 5 and row[4]:  # EmpleadoID existe
                    empleados[row[4]] = {
                        "Nombre": row[0] or "Sin Nombre",
                        "Cedula": row[1] or "",
                        "Celular": row[2] or "",
                        "Correo": row[3] or ""
                    }

            if not empleados:
                QMessageBox.information(self, "Info", "No hay empleados registrados en la base de datos.")
                return

            # LEER Y PROCESAR DATOS DE VALES
            vales_data = []
            headers = []
            
            # Obtener encabezados
            for cell in ws_vales[1]:
                headers.append(cell.value)
            
            # Buscar índices de columnas importantes
            try:
                empleado_id_idx = headers.index("EmpleadoID")
                fecha_idx = headers.index("FechaHora_Generacion")
                valor_idx = headers.index("Suma_Valores_Trabajos")

                id_col_name = "ID_Vale"
                id_idx = headers.index(id_col_name) if id_col_name in headers else None
                
                num_ticket_col_name = "Numero_Ticket_Asociado"
                num_ticket_idx = headers.index(num_ticket_col_name) if num_ticket_col_name in headers else None
                
                referencia_col_name = "Referencia_Asociada"
                referencia_idx = headers.index(referencia_col_name) if referencia_col_name in headers else None
                
                talla_col_name = "Resumen_Tallas_Cantidades"
                talla_idx = headers.index(talla_col_name) if talla_col_name in headers else None
                
                color_col_name = "Color_Trabajo"
                color_idx = headers.index(color_col_name) if color_col_name in headers else None
                
                total_producido_col_name = "Total_Producido_Trabajo"
                total_producido_idx = headers.index(total_producido_col_name) if total_producido_col_name in headers else None
                
                codigo_serial_col_name = "Codigo_Serial_Trabajo_Asociado"
                codigo_serial_idx = headers.index(codigo_serial_col_name) if codigo_serial_col_name in headers else None
                
                work_type_col_name = "WorkTypeDetected"
                work_type_idx = headers.index(work_type_col_name) if work_type_col_name in headers else None

            except ValueError as e:
                QMessageBox.warning(self, "Error", f"Falta una columna esencial en 'Vales' para generar reportes: {e}")
                return
            
            # Leer datos de vales
            for row in ws_vales.iter_rows(min_row=2, values_only=True):
                if len(row) > empleado_id_idx and row[empleado_id_idx]:
                    vale_dict = {
                        "empleado_id": row[empleado_id_idx],
                        "fecha": row[fecha_idx] if fecha_idx < len(row) else None,
                        "valor": row[valor_idx] if valor_idx < len(row) else None,
                        "id": row[id_idx] if id_idx is not None and id_idx < len(row) else None,
                        "num_ticket": row[num_ticket_idx] if num_ticket_idx is not None and num_ticket_idx < len(row) else None,
                        "referencia": row[referencia_idx] if referencia_idx is not None and referencia_idx < len(row) else None,
                        "talla": row[talla_idx] if talla_idx is not None and talla_idx < len(row) else None,
                        "color": row[color_idx] if color_idx is not None and color_idx < len(row) else None,
                        "total_producido": row[total_producido_idx] if total_producido_idx is not None and total_producido_idx < len(row) else None,
                        "codigo_serial": row[codigo_serial_idx] if codigo_serial_idx is not None and codigo_serial_idx < len(row) else None,
                        "work_type_detected": row[work_type_idx] if work_type_idx is not None and work_type_idx < len(row) else None
                    }
                    vales_data.append(vale_dict)

            # Invertir WORK_TYPE_ABBREVIATIONS para mapear códigos a nombres completos
            work_type_mapping = {v: k for k, v in WORK_TYPE_ABBREVIATIONS.items()}

            # Inicializar datos del reporte consolidado
            report_data = {}
            for emp_id in empleados:
                report_data[emp_id] = {
                    "Nombre": empleados[emp_id]["Nombre"],
                    "Total_Vales": 0,
                    "Total_Valor": 0.0,
                    "Trabajos": {work_type: 0 for work_type in WORK_TYPE_ABBREVIATIONS.keys()}
                }

            # Procesar vales para el reporte consolidado
            for vale in vales_data:
                emp_id = vale["empleado_id"]
                if emp_id in report_data:
                    report_data[emp_id]["Total_Vales"] += 1
                    
                    if vale["valor"] is not None:
                        try:
                            valor = float(vale["valor"])
                            report_data[emp_id]["Total_Valor"] += valor
                        except (ValueError, TypeError):
                            pass
                    
                    work_type = vale["work_type_detected"]
                    if work_type and work_type in work_type_mapping:
                        mapped_work_type = work_type_mapping[work_type]
                        if mapped_work_type in report_data[emp_id]["Trabajos"]:
                            report_data[emp_id]["Trabajos"][mapped_work_type] += 1

            # PARTE 1: Actualizar la tabla de previsualización (reporte consolidado)
            if hasattr(self.ui, 'tableViewVale') and hasattr(self, 'table_model'):
                self.table_model.clear()
                # Definir encabezados: fijos + dinámicos basados en WORK_TYPE_ABBREVIATIONS
                fixed_headers = ["EmpleadoID", "Nombre", "Total Vales", "Total Valor"]
                work_type_headers = list(WORK_TYPE_ABBREVIATIONS.keys())
                self.table_model.setHorizontalHeaderLabels(fixed_headers + work_type_headers)
                
                for emp_id, data in report_data.items():
                    row_data = [
                        emp_id,
                        data["Nombre"],
                        data["Total_Vales"],
                        round(data["Total_Valor"], 2)
                    ] + [data["Trabajos"][work_type] for work_type in WORK_TYPE_ABBREVIATIONS.keys()]
                    items = [QtGui.QStandardItem(str(value)) for value in row_data]
                    self.table_model.appendRow(items)
                
                self.ui.tableViewVale.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                print("Reporte consolidado generado y mostrado en 'tableViewVale'.")
            else:
                print("Advertencia: No se pudo actualizar la tabla 'tableViewVale'.")

            # PARTE 2: Crear hojas individuales por empleado
            for empleado_id, empleado_info in empleados.items():
                vales_empleado = [vale for vale in vales_data if vale["empleado_id"] == empleado_id]
                
                sheet_name = f"Empleado_{empleado_id}"
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
                
                emp_ws = wb.create_sheet(title=sheet_name)
                
                emp_ws['A1'] = "INFORMACIÓN DEL EMPLEADO"
                emp_ws['A1'].font = Font(bold=True, size=14)
                emp_ws.merge_cells('A1:F1')
                
                emp_ws['A2'] = "Nombre:"
                emp_ws['B2'] = empleado_info["Nombre"]
                emp_ws['A3'] = "Cédula:"
                emp_ws['B3'] = empleado_info["Cedula"]
                emp_ws['A4'] = "Celular:"
                emp_ws['B4'] = empleado_info["Celular"]
                emp_ws['A5'] = "Correo:"
                emp_ws['B5'] = empleado_info["Correo"]
                emp_ws['A6'] = "ID Empleado:"
                emp_ws['B6'] = empleado_id
                
                for row in range(2, 7):
                    emp_ws[f'A{row}'].font = Font(bold=True)
                
                emp_ws['A8'] = "RESUMEN POR TIPO DE TRABAJO"
                emp_ws['A8'].font = Font(bold=True, size=12)
                emp_ws.merge_cells('A8:D8')
                
                trabajo_resumen = {}
                valor_total_general = 0
                
                for vale in vales_empleado:
                    if vale["work_type_detected"] and vale["valor"] is not None:
                        trabajo = work_type_mapping.get(vale["work_type_detected"], vale["work_type_detected"])
                        if trabajo not in trabajo_resumen:
                            trabajo_resumen[trabajo] = {"cantidad": 0, "valor_total": 0}
                        
                        trabajo_resumen[trabajo]["cantidad"] += 1
                        try:
                            valor = float(vale["valor"])
                            trabajo_resumen[trabajo]["valor_total"] += valor
                            valor_total_general += valor
                        except (ValueError, TypeError):
                            print(f"Error al convertir valor '{vale['valor']}'")
                
                headers_resumen = ["Tipo de Trabajo", "Cantidad de Vales", "Valor Total", "Promedio por Vale"]
                for col, header in enumerate(headers_resumen, start=1):
                    cell = emp_ws.cell(row=9, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                row_idx = 10
                for trabajo, datos in sorted(trabajo_resumen.items()):
                    emp_ws.cell(row=row_idx, column=1).value = trabajo
                    emp_ws.cell(row=row_idx, column=2).value = datos["cantidad"]
                    
                    valor_cell = emp_ws.cell(row=row_idx, column=3)
                    valor_cell.value = datos["valor_total"]
                    valor_cell.number_format = '#,##0'
                    
                    promedio_cell = emp_ws.cell(row=row_idx, column=4)
                    promedio = datos["valor_total"] / datos["cantidad"] if datos["cantidad"] > 0 else 0
                    promedio_cell.value = promedio
                    promedio_cell.number_format = '#,##0'
                    
                    row_idx += 1
                
                emp_ws.cell(row=row_idx, column=1).value = "TOTAL GENERAL:"
                emp_ws.cell(row=row_idx, column=1).font = Font(bold=True)
                emp_ws.cell(row=row_idx, column=2).value = sum(d["cantidad"] for d in trabajo_resumen.values())
                emp_ws.cell(row=row_idx, column=2).font = Font(bold=True)
                
                total_general_cell = emp_ws.cell(row=row_idx, column=3)
                total_general_cell.value = valor_total_general
                total_general_cell.font = Font(bold=True)
                total_general_cell.number_format = '#,##0'
                
                detalle_start_row = row_idx + 3
                
                emp_ws.cell(row=detalle_start_row, column=1).value = "DETALLE DE VALES"
                emp_ws.cell(row=detalle_start_row, column=1).font = Font(bold=True, size=12)
                emp_ws.merge_cells(f'A{detalle_start_row}:J{detalle_start_row}')
                
                headers_detalle = ["ID", "Fecha", "# Ticket", "Referencia", "Talla", 
                                "Color", "Valor", "Total Producido", "Código Serial", "Trabajo"]
                detalle_start_row += 1
                for col, header in enumerate(headers_detalle, start=1):
                    cell = emp_ws.cell(row=detalle_start_row, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                row_idx = detalle_start_row + 1
                for vale in vales_empleado:
                    emp_ws.cell(row=row_idx, column=1).value = vale["id"]
                    
                    if vale["fecha"] is not None:
                        if isinstance(vale["fecha"], dt.datetime):
                            emp_ws.cell(row=row_idx, column=2).value = vale["fecha"].strftime("%Y-%m-%d")
                        else:
                            emp_ws.cell(row=row_idx, column=2).value = str(vale["fecha"])
                    else:
                        emp_ws.cell(row=row_idx, column=2).value = "Sin fecha"
                    
                    emp_ws.cell(row=row_idx, column=3).value = vale["num_ticket"]
                    emp_ws.cell(row=row_idx, column=4).value = vale["referencia"]
                    emp_ws.cell(row=row_idx, column=5).value = vale["talla"]
                    emp_ws.cell(row=row_idx, column=6).value = vale["color"]
                    
                    valor_cell = emp_ws.cell(row=row_idx, column=7)
                    valor_cell.value = vale["valor"]
                    valor_cell.number_format = '#,##0'
                    
                    emp_ws.cell(row=row_idx, column=8).value = vale["total_producido"]
                    emp_ws.cell(row=row_idx, column=9).value = vale["codigo_serial"]
                    emp_ws.cell(row=row_idx, column=10).value = work_type_mapping.get(vale["work_type_detected"], vale["work_type_detected"])
                    
                    row_idx += 1
                
                last_row = row_idx
                
                for col in range(1, 11):
                    emp_ws.column_dimensions[chr(64 + col)].width = 15
                
                self._create_consolidados(emp_ws, vales_empleado, last_row, work_type_mapping)

            wb.save(self.excel_path)
            
            QMessageBox.information(
                self, 
                "Reportes Actualizados", 
                f"Se han actualizado exitosamente:\n"
                f"• Reporte consolidado mostrado en la tabla\n"
                f"• {len(empleados)} hojas individuales de empleados\n"
                f"• Consolidados semanales, mensuales y anuales"
            )
            
            print("Reportes de empleados actualizados con éxito (reporte consolidado + hojas individuales).")
            
        except Exception as e:
            print(f"Error al generar los reportes de empleados: {e}")
            traceback.print_exc()
            QMessageBox.critical(self, "Error", f"No se pudieron generar los reportes: {e}")

    def _create_consolidados(self, emp_ws, vales_empleado, start_row, work_type_mapping):
        """
        Método auxiliar para crear las tablas de consolidados (semanal, mensual, anual)
        """
        try:
            consolidado_start_row = start_row + 3
            emp_ws.cell(row=consolidado_start_row, column=1).value = "CONSOLIDADO DE PAGOS"
            emp_ws.cell(row=consolidado_start_row, column=1).font = Font(bold=True, size=12)
            emp_ws.merge_cells(f'A{consolidado_start_row}:F{consolidado_start_row}')
            
            consolidado_start_row += 2
            
            emp_ws.cell(row=consolidado_start_row, column=1).value = "CONSOLIDADO SEMANAL"
            emp_ws.cell(row=consolidado_start_row, column=1).font = Font(bold=True)
            emp_ws.merge_cells(f'A{consolidado_start_row}:D{consolidado_start_row}')
            
            consolidado_start_row += 1
            headers_semanal = ["Semana", "Fecha Inicio", "Fecha Fin", "Valor Total"]
            for col, header in enumerate(headers_semanal, start=1):
                cell = emp_ws.cell(row=consolidado_start_row, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            weekly_data = {}
            for vale in vales_empleado:
                if vale["fecha"] is not None and isinstance(vale["fecha"], dt.datetime) and vale["valor"] is not None:
                    try:
                        year = vale["fecha"].year
                        week_num = vale["fecha"].isocalendar()[1]
                        week_key = f"{year}-W{week_num:02d}"
                        
                        start_of_week = vale["fecha"] - dt.timedelta(days=vale["fecha"].weekday())
                        end_of_week = start_of_week + dt.timedelta(days=6)
                        
                        if week_key not in weekly_data:
                            weekly_data[week_key] = {
                                "start_date": start_of_week,
                                "end_date": end_of_week,
                                "total": 0
                            }
                        
                        weekly_data[week_key]["total"] += float(vale["valor"])
                    except Exception as e:
                        print(f"Error procesando fecha semanal: {e}")
            
            row_idx = consolidado_start_row + 1
            for week_key, data in sorted(weekly_data.items()):
                emp_ws.cell(row=row_idx, column=1).value = week_key
                emp_ws.cell(row=row_idx, column=2).value = data["start_date"].strftime("%Y-%m-%d")
                emp_ws.cell(row=row_idx, column=3).value = data["end_date"].strftime("%Y-%m-%d")
                valor_cell = emp_ws.cell(row=row_idx, column=4)
                valor_cell.value = data["total"]
                valor_cell.number_format = '#,##0'
                row_idx += 1
            
            row_idx += 2
            emp_ws.cell(row=row_idx, column=1).value = "CONSOLIDADO MENSUAL"
            emp_ws.cell(row=row_idx, column=1).font = Font(bold=True)
            emp_ws.merge_cells(f'A{row_idx}:C{row_idx}')
            
            row_idx += 1
            headers_mensual = ["Año", "Mes", "Valor Total"]
            for col, header in enumerate(headers_mensual, start=1):
                cell = emp_ws.cell(row=row_idx, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            monthly_data = {}
            for vale in vales_empleado:
                if vale["fecha"] is not None and isinstance(vale["fecha"], dt.datetime) and vale["valor"] is not None:
                    try:
                        year = vale["fecha"].year
                        month = vale["fecha"].month
                        month_key = f"{year}-{month:02d}"
                        
                        if month_key not in monthly_data:
                            monthly_data[month_key] = {"year": year, "month": month, "total": 0}
                        
                        monthly_data[month_key]["total"] += float(vale["valor"])
                    except Exception as e:
                        print(f"Error procesando fecha mensual: {e}")
            
            month_names = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
            
            row_idx += 1
            for month_key, data in sorted(monthly_data.items()):
                emp_ws.cell(row=row_idx, column=1).value = data["year"]
                emp_ws.cell(row=row_idx, column=2).value = month_names[data["month"] - 1]
                valor_cell = emp_ws.cell(row=row_idx, column=3)
                valor_cell.value = data["total"]
                valor_cell.number_format = '#,##0'
                row_idx += 1
            
            row_idx += 2
            emp_ws.cell(row=row_idx, column=1).value = "CONSOLIDADO ANUAL"
            emp_ws.cell(row=row_idx, column=1).font = Font(bold=True)
            emp_ws.merge_cells(f'A{row_idx}:B{row_idx}')
            
            row_idx += 1
            headers_anual = ["Año", "Valor Total"]
            for col, header in enumerate(headers_anual, start=1):
                cell = emp_ws.cell(row=row_idx, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            yearly_data = {}
            for vale in vales_empleado:
                if vale["fecha"] is not None and isinstance(vale["fecha"], dt.datetime) and vale["valor"] is not None:
                    try:
                        year = vale["fecha"].year
                        if year not in yearly_data:
                            yearly_data[year] = 0
                        yearly_data[year] += float(vale["valor"])
                    except Exception as e:
                        print(f"Error procesando fecha anual: {e}")
            
            row_idx += 1
            for year, total in sorted(yearly_data.items()):
                emp_ws.cell(row=row_idx, column=1).value = year
                valor_cell = emp_ws.cell(row=row_idx, column=2)
                valor_cell.value = total
                valor_cell.number_format = '#,##0'
                row_idx += 1
                
        except Exception as e:
            print(f"Error creando consolidados: {e}")
    def setup_employee_management(self):
        """Carga empleados en el ComboBox, conecta el botón Registrar Vale y limpia la tabla de previsualización."""
        try:
            # --- Manejo de Archivo Excel ---
            try:
                wb = load_workbook(self.excel_path)
            except FileNotFoundError:
                print(f"Advertencia: Archivo Excel no encontrado en {self.excel_path}. Creando uno nuevo.")
                wb = Workbook()
                # Usa la hoja activa si es un libro nuevo
                if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
                    empleados_ws = wb.active
                    empleados_ws.title = "Empleados"
                else:
                    empleados_ws = wb.create_sheet(title="Empleados")
                # Añadir encabezados
                empleados_ws.append(["Nombre", "Cedula", "Celular", "Correo", "EmpleadoId"])
                # Añadir un empleado de ejemplo
                empleados_ws.append(["Juan Pérez (Ejemplo)", "1234567890", "3001234567", "juan@example.com", "E001"])
                wb.save(self.excel_path)
                print(f"Archivo Excel '{self.excel_path}' creado con hoja 'Empleados'.")

            if "Empleados" not in wb.sheetnames:
                print(f"Creando hoja 'Empleados' en archivo existente: {self.excel_path}")
                empleados_ws = wb.create_sheet(title="Empleados")
                empleados_ws.append(["Nombre", "Cedula", "Celular", "Correo", "EmpleadoId"])
                empleados_ws.append(["Juan Pérez (Ejemplo)", "1234567890", "3001234567", "juan@example.com", "E001"])
                wb.save(self.excel_path)
            else:
                empleados_ws = wb["Empleados"]
                # Asegurarse de que la hoja Empleados tenga las cabeceras básicas si está vacía o corrupta
                if empleados_ws.max_row == 0 or not all(empleados_ws.cell(row=1, column=c+1).value for c in range(5)):
                    empleados_ws.delete_rows(1, empleados_ws.max_row)  # Limpiar por si acaso
                    empleados_ws.append(["Nombre", "Cedula", "Celular", "Correo", "EmpleadoId"])
                    wb.save(self.excel_path)

            # --- Configuración del ComboBox ---
            if hasattr(self.ui, 'EmpleadosBox'):
                self.ui.EmpleadosBox.clear()  # Limpiar items previos
                self.ui.EmpleadosBox.setMinimumWidth(200)

                employees_loaded = False
                # Leer las 5 columnas estándar
                for row in empleados_ws.iter_rows(min_row=2, max_col=5):
                    # Asegurarse de que la fila tiene al menos las columnas de Nombre (0) y EmpleadoId (4)
                    if len(row) >= 5 and row[0].value and row[4].value:
                        name = str(row[0].value)
                        emp_id = str(row[4].value) if row[4].value else "SIN-ID"
                        self.ui.EmpleadosBox.addItem(f"{name} ({emp_id})", emp_id)
                        employees_loaded = True

                if not employees_loaded:
                    self.ui.EmpleadosBox.addItem("Sin empleados registrados", "")
                    print("No se cargaron empleados desde la hoja 'Empleados'.")
                else:
                    print(f"Se cargaron {self.ui.EmpleadosBox.count()} empleados.")
            else:
                print("ERROR: El QComboBox 'EmpleadosBox' no se encontró en la UI.")
                if hasattr(self.ui, 'btnRegisterVale'):
                    self.ui.btnRegisterVale.setEnabled(False)
                return

            # --- Limpiar la tabla de previsualización ---
            if hasattr(self.ui, 'tableViewVale') and hasattr(self, 'table_model'):
                self.table_model.clear()  # Limpiar todas las filas
                # Definir los encabezados fijos
                fixed_headers = ["Código Serial", "Número Ticket", "Referencia", "Tipo Trabajo", "Color", "Total Producido"]
                # Generar los encabezados dinámicos para los valores de trabajo
                valor_headers = [f"Valor {work_type}" for work_type in WORK_TYPE_ABBREVIATIONS.keys()]
                # Combinar encabezados fijos y dinámicos
                self.table_model.setHorizontalHeaderLabels(fixed_headers + valor_headers)
                self.ui.tableViewVale.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                print("Tabla de previsualización 'tableViewVale' limpiada y encabezados actualizados.")
            else:
                print("Advertencia: No se pudo limpiar la tabla 'tableViewVale' porque no existe o no tiene modelo.")

            # --- Conexión del Botón Registrar Vale ---
            if hasattr(self.ui, 'btnRegisterVale'):
                if hasattr(self, 'register_vale'):
                    try:
                        self.ui.btnRegisterVale.clicked.disconnect()
                    except RuntimeError:
                        pass
                    self.ui.btnRegisterVale.clicked.connect(self.register_vale)
                    self.ui.btnRegisterVale.setEnabled(True)
                    print("Botón 'btnRegisterVale' conectado correctamente.")
                else:
                    print("ERROR: El método 'register_vale' no está definido en esta clase.")
                    self.ui.btnRegisterVale.setEnabled(False)
            else:
                print("ERROR: El QPushButton 'btnRegisterVale' no se encontró en la UI.")

        except FileNotFoundError:
            print(f"ERROR CRÍTICO: No se pudo encontrar ni crear el archivo Excel en {self.excel_path}")
        except Exception as e:
            print(f"Error inesperado durante setup_employee_management: {e}")
            traceback.print_exc()

    def _collect_ticket_data(self, suffix="", ticket_label="A"):
        """
        Función auxiliar para recolectar y validar los datos de un tiquete.
        Reutiliza la lógica para el tiquete A (suffix="") y B (suffix="_2").
        Retorna un diccionario con los datos o None si hay un error.
        """
        ui = self.ui
        fields_to_clear = []
        
        # 1. Recolectar campos básicos
        try:
            referencia_field = getattr(ui, f"CampoReferenciaTrabajo{suffix}")
            ticket_field = getattr(ui, f"CampoNumeroTicket{suffix}")
            color_field = getattr(ui, f"CampoColor{suffix}")
        except AttributeError as e:
            QMessageBox.critical(self, "Error de UI", f"No se encontró un campo esperado: {e}. Revisa los nombres en tu archivo .ui.")
            return None, None

        fields_to_clear.extend([referencia_field, ticket_field, color_field])

        referencia = referencia_field.text().strip()
        ticket_number = ticket_field.text().strip()
        color = color_field.text().strip()

        if not all([referencia, ticket_number, color]):
            QMessageBox.warning(self, "Campos Incompletos", f"Debes llenar Referencia, N° de Tiquete y Color para el Tiquete {ticket_label}.")
            return None, None

        # 2. Recolectar tallas
        tallas_cantidades = {}
        has_any_talla = False
        for i in range(33, 49):
            talla_field_name = f"CampoTalla_{i}{suffix}"
            if hasattr(ui, talla_field_name):
                talla_field = getattr(ui, talla_field_name)
                fields_to_clear.append(talla_field)
                cantidad_text = talla_field.text().strip()
                if cantidad_text:
                    if not cantidad_text.isdigit() or int(cantidad_text) <= 0:
                        QMessageBox.warning(self, "Entrada Inválida", f"La cantidad para Talla {i} (Tiquete {ticket_label}) debe ser un número positivo.")
                        talla_field.setFocus()
                        return None, None
                    tallas_cantidades[str(i)] = int(cantidad_text)
                    has_any_talla = True
        
        if not has_any_talla:
            QMessageBox.warning(self, "Campos Incompletos", f"Debes ingresar cantidad para al menos una talla en el Tiquete {ticket_label}.")
            return None, None
            
        # 3. Empaquetar y retornar datos
        ticket_data = {
            "ticket_number": ticket_number,
            "referencia": referencia,
            "color": color,
            "tallas_cantidades": tallas_cantidades,
        }
        return ticket_data, fields_to_clear


    def on_save_button_clicked(self):
        """
        Handler para el botón de guardar, actualizado para manejar dos tiquetes
        y llamar al nuevo generador de PDF.
        """
        # --- Recolección de Datos para Ambos Tiquetes ---
        print("Recolectando datos del Tiquete A (Izquierda)...")
        ticket_A_data, fields_A = self._collect_ticket_data(suffix="", ticket_label="A")
        if ticket_A_data is None:
            return # La validación falló y el mensaje ya se mostró

        print("Recolectando datos del Tiquete B (Derecha)...")
        ticket_B_data, fields_B = self._collect_ticket_data(suffix="", ticket_label="B")
        if ticket_B_data is None:
            return # La validación falló

        all_fields_to_clear = fields_A + fields_B

        # --- Generación de PDF ---
        # Crea un nombre de archivo descriptivo
        output_filename = f"codes/vales_{ticket_A_data['ticket_number']}_{ticket_B_data['ticket_number']}.pdf"

        # Llama a la función importada que ahora maneja dos tiquetes
        pdf_path = generate_vale_pdf(
            left_ticket_info=ticket_A_data,
            right_ticket_info=ticket_B_data,
            output_filename=output_filename
        )

        if not pdf_path:
            QMessageBox.critical(self, "Error", "No se pudo generar el archivo PDF.")
            return

        # --- Feedback al Usuario ---
        QMessageBox.information(
            self, "Operación Exitosa",
            f"PDF para tiquetes {ticket_A_data['ticket_number']} y {ticket_B_data['ticket_number']} generado exitosamente.\n\n"
            f"Archivo guardado en: {pdf_path}"
        )

        # --- Limpieza de Campos ---
        print("Limpiando campos de la UI...")
        for field_widget in all_fields_to_clear:
            field_widget.clear()
        
        # Devolver el foco al primer campo
        if hasattr(self.ui, 'CampoReferenciaTrabajo'):
            self.ui.CampoReferenciaTrabajo.setFocus()

        # Opcional: Abrir el PDF generado
        try:
            if sys.platform == "win32":
                os.startfile(os.path.abspath(pdf_path))
            else:
                opener = "open" if sys.platform == "darwin" else "xdg-open"
                subprocess.call([opener, os.path.abspath(pdf_path)])
        except Exception as e:
            print(f"No se pudo abrir el PDF automáticamente: {e}")
        
        # Itera sobre el diccionario de abreviaturas actualizado
        for work_type, abbr in WORK_TYPE_ABBREVIATIONS.items():
            # SOLO genera código si se ingresó un valor para ese tipo de trabajo
            if work_type in valores_trabajo:
                serial_code = generate_serial_code(ticket_number, referencia, color, tallas_cantidades, abbr)
                # Llama a generate_barcode (que genera QR)
                # PASA el work_type (nombre legible) como clave a barcode_paths
                barcode_path = generate_barcode(serial_code, subfolder_name=subfolder_name_for_codes)
                if not barcode_path:
                    QMessageBox.critical(self, "Error", f"Error al generar el código QR para {work_type}.")
                    return
                serial_codes[work_type] = serial_code
                barcode_paths[work_type] = barcode_path # Clave = Nombre del trabajo ("Corte", "Empaque", etc.)
        # --- Fin Generación de Códigos ---

        # --- Guardado en Excel (Tu lógica, pero asegúrate que use valores_trabajo y serial_codes actualizados) ---
        first_work_type_with_code = next(iter(barcode_paths), None)
        image_path_for_excel = barcode_paths.get(first_work_type_with_code, "")

        if not self.save_to_excel( # Debes actualizar save_to_excel para manejar las nuevas columnas
            serial_codes=serial_codes,
            code_path=image_path_for_excel,
            ticket_number=ticket_number,
            referencia=referencia,
            color=color,
            tallas_cantidades=tallas_cantidades,
            total_producido_calculado=total_producido_calculado,
            valores_trabajo=valores_trabajo
        ):
            QMessageBox.critical(self, "Error de Guardado", "No se pudieron guardar los datos en Excel.")
            return

        # --- Generación de PDF (Actualizado) ---
        # ¡IMPORTANTE! Asegúrate que generate_vale_pdf ahora usa barcode_paths
        # con las claves correctas ("Corte", "Empaque", etc.) y que recibe
        # TIPOS_DE_TRABAJO si lo necesita para iterar.
        pdf_path = generate_vale_pdf(
            ticket_number, referencia, tallas_cantidades, color,
            total_producido_calculado, barcode_paths, valores_trabajo,
            TIPOS_DE_TRABAJO # Pasa el diccionario con los tipos
        )

        if not pdf_path:
            QMessageBox.critical(self, "Error", "Datos guardados, pero error al generar PDF.")
            return

        # Muestra imagen y mensaje (Tu lógica)
        if image_path_for_excel:
            display_code_image(self.ui, image_path_for_excel)
        else:
            if hasattr(self.ui, 'PreviwImage') and self.ui.PreviwImage.scene():
                self.ui.PreviwImage.scene().clear()

        # Asumiendo que self.current_code_type sigue existiendo o se adapta
        current_code_type = getattr(self, 'current_code_type', 'QR') 

        QMessageBox.information(
            self, "Operación Exitosa",
            f"Códigos generados.\n"
            f"Tipo: {current_code_type.upper()}\n"
            f"Total Producido: {total_producido_calculado} unidades.\n"
            f"PDF generado en: {pdf_path}"
        )

        # Limpieza (Tu lógica)
        for field_widget in all_fields_to_clear:
            field_widget.clear()
        if hasattr(self.ui, 'CampoReferenciaTrabajo'):
            self.ui.CampoReferenciaTrabajo.setFocus()


    def setup_window(self):
        """Setup basic window properties and appearance"""
        # Remove standard title bar
        UIFunctions.removeTitleBar(True)
        
        # Set window title
        self.setWindowTitle('Gestor de Vales')
        UIFunctions.labelTitle(self, 'Thimoty')
        
        # Set window size
        startSize = QSize(1300, 720)
        self.resize(startSize)
        self.setMinimumSize(startSize)
        # UIFunctions.enableMaximumSize(self, 500, 720)

    def setup_menus(self):
        """Setup menu functionality and navigation"""
        # Toggle menu size button
        self.ui.btn_toggle_menu.clicked.connect(lambda: UIFunctions.toggleMenu(self, 220, True))
        
        # Add custom menus
        self.ui.stackedWidget.setMinimumWidth(20)
        UIFunctions.addNewMenu(self, "Leer Vales", "btn_home", "url(:/16x16/icons/16x16/cil-home.png)", True)
        UIFunctions.addNewMenu(self, "Crear Vales", "btn_new_user", "url(:/16x16/icons/16x16/cil-user-follow.png)", True)
        
        # ==> NUEVOS MENÚS AÑADIDOS AQUÍ
        UIFunctions.addNewMenu(self, "Tablas", "btn_tables", "url(:/16x16/icons/16x16/cil-grid.png)", True)
        UIFunctions.addNewMenu(self, "Consolidado", "btn_consolidado", "url(:/16x16/icons/16x16/cil-chart-pie.png)", True)
        
        # Menú de configuración (si lo necesitas, si no, puedes eliminarlo)
        UIFunctions.addNewMenu(self, "Configuracion", "btn_widgets", "url(:/16x16/icons/16x16/cil-equalizer.png)", False)

        # Select starting menu
        UIFunctions.selectStandardMenu(self, "btn_home")
        
        # Set starting page
        self.ui.stackedWidget.setCurrentWidget(self.ui.page_home)

    def setup_eliminar_todo_button(self):
        """Conecta el botón EliminarTODO con la función de eliminación"""
        self.ui.EliminarTODO.clicked.connect(self.eliminar_todo_con_backup)

    def eliminar_todo_con_backup(self):
        """
        Elimina la base de datos Excel y el directorio codes después de crear un backup
        Versión con threading para UI más responsiva
        """
        try:
            # Confirmar acción con el usuario
            reply = QMessageBox.question(
                self, 
                'Confirmar Eliminación', 
                '¿Está seguro de que desea eliminar toda la base de datos?\n\n'
                'Se creará un backup automáticamente antes de eliminar.',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
                
            # Obtener rutas usando BackupManager
            excel_path, codes_dir = BackupManager.get_paths()
            
            # Verificar si hay algo que respaldar
            has_files, excel_exists, codes_exists = BackupManager.check_files_exist(excel_path, codes_dir)
            
            if not has_files:
                QMessageBox.information(
                    self, 
                    'Información', 
                    'No hay archivos para eliminar. La base de datos ya está limpia.'
                )
                return
            
            # Deshabilitar el botón mientras se procesa
            self.ui.EliminarTODO.setEnabled(False)
            self.ui.EliminarTODO.setText("Procesando...")
            QApplication.processEvents()
            
            # Crear backup en thread separado
            self.backup_thread = BackupThread(excel_path, codes_dir)
            self.backup_thread.finished.connect(self.on_backup_finished)
            self.backup_thread.progress.connect(self.on_backup_progress)
            self.backup_thread.start()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al iniciar eliminación: {str(e)}")
            self.ui.EliminarTODO.setEnabled(True)
            self.ui.EliminarTODO.setText("Eliminar TODO")
    def eliminar_todo_simple(self):
        """
        Versión simple sin threading - buena para pocos archivos
        """
        try:
            # Confirmar acción
            reply = QMessageBox.question(
                self, 
                'Confirmar Eliminación', 
                '¿Está seguro de que desea eliminar toda la base de datos?\n\n'
                'Se creará un backup automáticamente.',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
            
            # Cambiar botón para mostrar progreso
            self.ui.EliminarTODO.setEnabled(False)
            original_text = self.ui.EliminarTODO.text()
            
            def update_progress(message):
                self.ui.EliminarTODO.setText(message)
                QApplication.processEvents()
            
            # Obtener rutas
            excel_path, codes_dir = BackupManager.get_paths()
            
            # Verificar archivos
            has_files, excel_exists, codes_exists = BackupManager.check_files_exist(excel_path, codes_dir)
            
            if not has_files:
                QMessageBox.information(self, 'Información', 'No hay archivos para eliminar.')
                return
            
            # Crear backup
            update_progress("Creando backup...")
            backup_name = BackupManager.create_backup_sync(excel_path, codes_dir, update_progress)
            
            # Eliminar archivos originales
            update_progress("Eliminando archivos...")
            deleted_files = BackupManager.delete_files(excel_path, codes_dir)
            
            # Recrear estructura
            update_progress("Recreando estructura...")
            self.setup_code_generator()
            
            QMessageBox.information(
                self, 
                'Completado', 
                f'Backup creado: {backup_name}\n'
                f'Archivos eliminados: {", ".join(deleted_files)}'
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error durante la eliminación: {str(e)}")
        
        finally:
            # Restaurar botón
            self.ui.EliminarTODO.setEnabled(True)
            self.ui.EliminarTODO.setText("Eliminar TODO")
    
    def on_backup_progress(self, message):
        """Callback para actualizar progreso del backup"""
        self.ui.EliminarTODO.setText(message)
    
    def on_backup_finished(self, success, message):
        """Callback cuando termina el backup"""
        try:
            if success:
                # Backup exitoso, proceder con eliminación
                self.ui.EliminarTODO.setText("Eliminando archivos...")
                QApplication.processEvents()
                
                excel_path, codes_dir = BackupManager.get_paths()
                deleted_files = BackupManager.delete_files(excel_path, codes_dir)
                
                # Recrear estructura
                self.ui.EliminarTODO.setText("Recreando estructura...")
                QApplication.processEvents()
                self.setup_code_generator()
                
                QMessageBox.information(
                    self, 
                    'Completado', 
                    f'{message}\n'
                    f'Archivos eliminados: {", ".join(deleted_files)}'
                )
            else:
                # Error en backup
                QMessageBox.critical(self, "Error en Backup", message)
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error en proceso de eliminación: {str(e)}")
        
        finally:
            # Rehabilitar botón
            self.ui.EliminarTODO.setEnabled(True)
            self.ui.EliminarTODO.setText("Eliminar TODO")





    def on_code_scanned(self):
        """
        Maneja el evento cuando se escanea o ingresa un código.
        (Actualizado para usar los nuevos tipos de trabajo)
        """
        if not hasattr(self.ui, 'codeReaderInput'):
            print("ERROR: codeReaderInput no encontrado en UI")
            return

        scanned_code = self.ui.codeReaderInput.text().strip()
        if not scanned_code:
            QMessageBox.warning(self, "Entrada Vacía", "Por favor, escanee o ingrese un código válido.")
            return

        # Cargar el archivo Excel
        try:
            wb = load_workbook(self.excel_path)
            ws_trabajos = wb["Trabajos"]
            ws_vales = wb[self.vales_sheet_name]
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo abrir el archivo Excel: {e}")
            return

        # --- Búsqueda Actualizada ---
        # Usa los NUEVOS tipos de trabajo para buscar
        work_types = list(WORK_TYPE_ABBREVIATIONS.keys())
        work_type_columns = {work_type: f"Código_{work_type}" for work_type in work_types}
        found_row = None
        work_type_found = None

        headers = [cell.value for cell in ws_trabajos[1]]
        # --- Fin Búsqueda Actualizada ---

        # VALIDACIÓN 1: Tabla visual (Tu lógica existente)
        if hasattr(self.ui, 'tableViewVale') and hasattr(self, 'table_model') and self.table_model.rowCount() > 0:
            for row_idx in range(self.table_model.rowCount()):
                item = self.table_model.item(row_idx, 0) # Columna 0 es Código Serial
                if item and item.text() == scanned_code:
                    QMessageBox.warning(self, "Vale Duplicado", f"El código '{scanned_code}' ya está en la tabla actual.")
                    self.ui.codeReaderInput.clear()
                    self.ui.codeReaderInput.setFocus()
                    return

        # VALIDACIÓN 2: Hoja "Vales" (Tu lógica existente)
        if ws_vales.max_row > 1:
            vales_headers_list = [cell.value for cell in ws_vales[1]]
            codigo_serial_trabajo_col_name = "Codigo_Serial_Trabajo_Asociado"
            try:
                codigo_serial_vales_idx = vales_headers_list.index(codigo_serial_trabajo_col_name)
                for row_num in range(2, ws_vales.max_row + 1):
                    cell_value = ws_vales.cell(row=row_num, column=codigo_serial_vales_idx + 1).value
                    if cell_value == scanned_code:
                        QMessageBox.warning(self, "Vale Ya Registrado", f"El vale '{scanned_code}' ya fue registrado en Excel.")
                        self.ui.codeReaderInput.clear()
                        self.ui.codeReaderInput.setFocus()
                        return
            except ValueError:
                QMessageBox.critical(self, "Error de Configuración", f"La columna '{codigo_serial_trabajo_col_name}' no se encuentra en 'Vales'.")
                self.ui.codeReaderInput.clear()
                self.ui.codeReaderInput.setFocus()
                return

        # Búsqueda en "Trabajos" (Tu lógica existente, pero ahora usa work_type_columns actualizado)
        for row_num, current_row_values in enumerate(ws_trabajos.iter_rows(min_row=2, values_only=True), start=2):
            for work_type, column_name in work_type_columns.items():
                try:
                    col_idx = headers.index(column_name)
                    if current_row_values[col_idx] == scanned_code:
                        found_row = current_row_values
                        work_type_found = work_type
                        break
                except (ValueError, IndexError):
                    # La columna no existe o hay un problema, lo manejamos más adelante
                    continue
            if found_row:
                break

        if not found_row:
            QMessageBox.warning(self, "Código No Encontrado", f"El código '{scanned_code}' no se encontró en 'Trabajos'.")
            self.ui.codeReaderInput.clear()
            self.ui.codeReaderInput.setFocus()
            return

        # --- Extracción de Datos Actualizada ---
        try:
            def get_value_from_row_safe(header_name, row_data, header_list):
                try:
                    idx = header_list.index(header_name)
                    return row_data[idx]
                except (ValueError, IndexError):
                    print(f"Advertencia: Columna '{header_name}' no encontrada.")
                    return None

            ticket_number = get_value_from_row_safe("Número Ticket", found_row, headers)
            referencia = get_value_from_row_safe("Referencia", found_row, headers)
            color = get_value_from_row_safe("Color", found_row, headers)
            total_producido = get_value_from_row_safe("Total Producido", found_row, headers)
            # Obtener el valor del trabajo específico encontrado
            valor_trabajo_especifico = get_value_from_row_safe(f"Valor {work_type_found}", found_row, headers)

            if any(v is None for v in [ticket_number, referencia, color, total_producido, valor_trabajo_especifico]):
                QMessageBox.critical(self, "Error de Datos", f"Faltan datos esenciales para '{scanned_code}'. Verifique la hoja 'Trabajos'.")
                return

            tallas_cantidades = {}
            for i in range(33, 49):
                cantidad = get_value_from_row_safe(f"Cant_T{i}", found_row, headers)
                if cantidad and isinstance(cantidad, (int, float)) and cantidad > 0:
                    tallas_cantidades[str(i)] = int(cantidad)
            resumen_tallas = "; ".join([f"T{k}:{v}" for k, v in tallas_cantidades.items()]) or "N/A"

        except Exception as e:
            QMessageBox.critical(self, "Error de Procesamiento", f"Error al procesar datos para '{scanned_code}': {str(e)}")
            traceback.print_exc()
            return
        # --- Fin Extracción de Datos ---

        # Obtener empleado (Tu lógica existente)
        if not hasattr(self.ui, 'EmpleadosBox'):
            QMessageBox.critical(self, "Error de UI", "No se encontró el ComboBox de empleados.")
            return
        empleado_id = self.ui.EmpleadosBox.currentData()
        if not empleado_id:
            QMessageBox.warning(self, "Empleado No Seleccionado", "Por favor, seleccione un empleado.")
            self.ui.EmpleadosBox.setFocus()
            return

        # Crear fila para "Vales" (Tu lógica existente)
        id_vale = f"V{int(time.time())}{str(uuid.uuid4())[:4]}"
        fecha_hora_actual = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        vale_row = [
            id_vale, empleado_id, fecha_hora_actual, ticket_number, referencia,
            color, resumen_tallas, total_producido, valor_trabajo_especifico,
            scanned_code, work_type_found
        ]

        # Guardar en "Vales" (Tu lógica existente)
        try:
            ws_vales.append(vale_row)
            wb.save(self.excel_path)
            QMessageBox.information(self, "Vale Registrado", f"Vale {id_vale} registrado para {work_type_found}.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo guardar el vale en Excel: {e}")
            return

        # --- Actualizar Tabla Visual (Actualizado) ---
        if hasattr(self.ui, 'tableViewVale'):
            # Crear la fila con los datos en el ORDEN CORRECTO
            # (Asegúrate que este orden coincida con el de setup_code_reader)
            row_data = [
                scanned_code,
                ticket_number,
                referencia,
                work_type_found, # Tipo de Trabajo encontrado
                color,
                total_producido
            ]
            # Añadir todos los valores de trabajo, estén o no en la fila
            for wt in WORK_TYPE_ABBREVIATIONS.keys():
                row_data.append(get_value_from_row_safe(f"Valor {wt}", found_row, headers))

            items = [QtGui.QStandardItem(str(value) if value is not None else "") for value in row_data]
            self.table_model.appendRow(items)
        # --- Fin Actualizar Tabla Visual ---

        # Limpiar campo (Tu lógica existente)
        self.ui.codeReaderInput.clear()
        self.ui.codeReaderInput.setFocus()



    def setup_add_employee_button(self):
        """Conecta el botón de agregar empleado (si lo tienes)"""
        if hasattr(self.ui, 'btnAgregarEmpleado'):
            self.ui.btnAgregarEmpleado.clicked.connect(self.add_employee)
            print("Botón 'btnAgregarEmpleado' conectado correctamente.")
        else:
            print("Botón 'btnAgregarEmpleado' no encontrado en la UI.")



    # Versión mejorada con validación
    def add_employee(self):
        """Agrega empleado con validaciones adicionales"""
        try:
            # Obtener valores de los campos
            nombre = self.ui.Nombre_Empleado.text().strip() if hasattr(self.ui, 'Nombre_Empleado') else ""
            cedula = self.ui.Cedula_Empleado.text().strip() if hasattr(self.ui, 'Cedula_Empleado') else ""
            celular = self.ui.Celular_Empleado.text().strip() if hasattr(self.ui, 'Celular_Empleado') else ""
            correo = self.ui.Correo_Empleado.text().strip() if hasattr(self.ui, 'Correo_Empleado') else ""
            
            # Validaciones
            if not nombre:
                QMessageBox.warning(self, "Campo Requerido", "El nombre es obligatorio.")
                self.ui.Nombre_Empleado.setFocus()
                return
                
            if not cedula:
                QMessageBox.warning(self, "Campo Requerido", "La cédula es obligatoria.")
                self.ui.Cedula_Empleado.setFocus()
                return
            
            # Validar formato de cédula
            if not validate_cedula(cedula):
                QMessageBox.warning(self, "Cédula Inválida", "La cédula debe contener solo números (6-12 dígitos).")
                self.ui.Cedula_Empleado.setFocus()
                return
            
            # Validar email si se proporciona
            if correo and "@" not in correo:
                QMessageBox.warning(self, "Email Inválido", "El formato del correo electrónico no es válido.")
                self.ui.Correo_Empleado.setFocus()
                return
            
            # Generar ID único del empleado
            unique_id = str(uuid.uuid4())[:4]
            empleado_id = f"E{cedula[-4:]}{unique_id}".upper()
            
            # Cargar Excel y verificar duplicados
            wb = load_workbook(self.excel_path)
            
            if "Empleados" not in wb.sheetnames:
                empleados_ws = wb.create_sheet(title="Empleados")
                empleados_ws.append(["Nombre", "Cedula", "Celular", "Correo", "EmpleadoId"])
            else:
                empleados_ws = wb["Empleados"]
            
            # Verificar si la cédula ya existe
            for row in empleados_ws.iter_rows(min_row=2, max_col=5):
                if row[1].value and str(row[1].value) == cedula:
                    QMessageBox.warning(self, "Empleado Existente", f"Ya existe un empleado con cédula {cedula}")
                    return
            
            # Agregar empleado
            empleados_ws.append([nombre, cedula, celular, correo, empleado_id])
            wb.save(self.excel_path)
            
            # Limpiar campos
            self.ui.Nombre_Empleado.clear()
            self.ui.Cedula_Empleado.clear()
            self.ui.Celular_Empleado.clear()
            self.ui.Correo_Empleado.clear()
            
            # Actualizar ComboBox
            if hasattr(self, 'setup_employee_management'):
                self.setup_employee_management()
            
            # Mensaje de éxito
            QMessageBox.information(self, "Empleado Agregado", f"Empleado: {nombre}\nCédula: {cedula}\nID: {empleado_id}")
            
            print(f"Empleado agregado exitosamente: {nombre} ({empleado_id})")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al agregar empleado: {str(e)}")
            print(f"Error en add_employee_with_validation: {e}")


        #TODOok



    #TODOok
    def register_vale(self):
        """Registra los vales escaneados y limpia la tabla."""
        try:
            if not self.table_model or self.table_model.rowCount() == 0:
                QMessageBox.warning(self, "Sin Vales", "No hay vales pendientes para registrar.")
                return

            total_vales = self.table_model.rowCount()
            self.table_model.removeRows(0, total_vales)

            QMessageBox.information(
                self, "Vales Registrados",
                f"Se han limpiado {total_vales} registros de la vista previa."
            )
        except Exception as e:
            print(f"Error al registrar vales: {e}")
            QMessageBox.critical(self, "Error", f"Error al registrar vales: {str(e)}")

    def Button(self):
        """Handler for menu button clicks"""
        # Get clicked button
        btnWidget = self.sender()

        # Handle different pages
        if btnWidget.objectName() == "btn_home":
            self.ui.stackedWidget.setCurrentWidget(self.ui.page_home)
            UIFunctions.resetStyle(self, "btn_home")
            btnWidget.setStyleSheet(UIFunctions.selectMenu(btnWidget.styleSheet()))
            
        elif btnWidget.objectName() == "btn_new_user":
            self.ui.stackedWidget.setCurrentWidget(self.ui.page_widgets) # Asumo que esta es la página para "Crear Vales"
            UIFunctions.resetStyle(self, "btn_new_user") # Corregido: antes decía "btn_widgets"
            btnWidget.setStyleSheet(UIFunctions.selectMenu(btnWidget.styleSheet()))

        # ==> NUEVOS HANDLERS AÑADIDOS AQUÍ
        elif btnWidget.objectName() == "btn_tables":
            self.ui.stackedWidget.setCurrentWidget(self.ui.page_tables)
            UIFunctions.resetStyle(self, "btn_tables")
            btnWidget.setStyleSheet(UIFunctions.selectMenu(btnWidget.styleSheet()))

        elif btnWidget.objectName() == "btn_consolidado":
            self.ui.stackedWidget.setCurrentWidget(self.ui.page_consolidado)
            UIFunctions.resetStyle(self, "btn_consolidado")
            btnWidget.setStyleSheet(UIFunctions.selectMenu(btnWidget.styleSheet()))

        # Handler para el botón de configuración (si lo usas)
        elif btnWidget.objectName() == "btn_widgets":
            # Asegúrate de que la página sea la correcta, por ejemplo self.ui.page_settings
            self.ui.stackedWidget.setCurrentWidget(self.ui.create_user) # Cambia create_user por tu página de configuración
            UIFunctions.resetStyle(self, "btn_widgets") # Corregido: antes decía "create_user"
            btnWidget.setStyleSheet(UIFunctions.selectMenu(btnWidget.styleSheet()))
    # Window event handlers
    def moveWindow(self, event):
        """Handle window movement"""
        # If maximized change to normal
        if UIFunctions.returStatus() == 1:
            UIFunctions.maximize_restore(self)

        # Move window
        if event.buttons() == Qt.LeftButton:
            self.move(self.pos() + event.globalPos() - self.dragPos)
            self.dragPos = event.globalPos()
            event.accept()

    def eventFilter(self, watched, event):
        """Event filter for double-click events"""
        if hasattr(self, 'le') and watched == self.le and event.type() == QtCore.QEvent.MouseButtonDblClick:
            print("pos: ", event.pos())
        return super().eventFilter(watched, event)

    def mousePressEvent(self, event):
        """Handle mouse press events"""
        self.dragPos = event.globalPos()
        if event.buttons() == Qt.LeftButton:
            print('Mouse click: LEFT CLICK')
        if event.buttons() == Qt.RightButton:
            print('Mouse click: RIGHT CLICK')
        if event.buttons() == Qt.MidButton:
            print('Mouse click: MIDDLE BUTTON')

    def keyPressEvent(self, event):
        """Handle key press events"""
        print('Key: ' + str(event.key()) + ' | Text Press: ' + str(event.text()))

    def resizeEvent(self, event):
        """Handle resize events"""
        self.resizeFunction()
        return super(MainWindow, self).resizeEvent(event)

    def resizeFunction(self):
        """Log window size on resize"""
        print('Height: ' + str(self.height()) + ' | Width: ' + str(self.width()))

if __name__ == "__main__":
    app = QApplication(sys.argv)
    QtGui.QFontDatabase.addApplicationFont('fonts/segoeui.ttf')
    QtGui.QFontDatabase.addApplicationFont('fonts/segoeuib.ttf')
    window = MainWindow()

    sys.exit(app.exec_())